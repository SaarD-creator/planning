# Last minute planning is vaak niet top
# overschakeling compleettt
# splitsing volgens ideaalmomenten
# samengevoegde attracties als 1 --> fix in plaatsing! maar nog vaak 2x drie uur bij 1 attractie
# niet meer laden na elke verandering
# Werkblad heropleidingen! Plus op de planning
# LM: afkappingen is in orde (automatisch op planning) Maar: lange blokken (5 uur aan een stuk) kunnen precies voorkomen...
# post-processing laat geen switches toe die een 1-uursblok achterlaten (ook niet in LM)
# max van vier uur aaneensluitend is weg, maar lossere regels in post-processing (wel vaak vier uursblokken)
# last minute afwezigen ziet er al goed uit!
# uitschakelen pauzevlinderuren werkt!
# volgorde verdeling met pauzevlinders laatst!
# pauzes kloppen!! 1h15 min pauze voor minderjarige lange werkers --> alleen nog niet top op korte dagen (mogelijks gefixt)
# nieuw werkblad analyse
# zelfde versie als 3.5 maar pauzevlinders zijn ook volgens volgorde uit gekozen nummertje
#betere verdeling 3 uur blokken, maar te veel 6 uur bij zelfde attractie & 1+3 logica voor 9u30 ipv 3+1 & 2+2 logica voor 4 uur opt einde

#uitschakelen attracties op bepaalde uren lijkt te werken!
#samenvoegen attracties per uur werkttttt!!! Kleine bug is er uit gehaald
#hele dag bij attractie werkt
# probleem met twee


import streamlit as st
import random
from collections import defaultdict
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from io import BytesIO
import datetime


# -----------------------------
# Excelbestand uploaden
# -----------------------------
uploaded_file = st.file_uploader("Upload Excel bestand", type=["xlsx"])

if not uploaded_file:
    st.session_state.pop("lm_base_bytes", None)   # ← reset bij nieuw bestand
    st.warning("Upload eerst het Excelbestand met de gegevens om verder te gaan.")
    st.stop()

file_bytes = uploaded_file.read()

# Workbook met berekende waarden voor de gewone planning
wb = load_workbook(BytesIO(file_bytes), data_only=True)
ws = wb["Input"]


ws_speciaal = wb["Input_"]
ws_aanpassingen = wb["Aanpassingen"]

# -----------------------------
# Datum op basis van W4 in Input_
# -----------------------------
_vandaag_datum = datetime.date.today()
_morgen_datum = _vandaag_datum + datetime.timedelta(days=1)
_w4 = str(ws_speciaal.cell(4, 23).value or "").strip().lower()
if _w4 == "morgen":
    vandaag = _morgen_datum.strftime("%d-%m-%Y")
else:
    vandaag = _vandaag_datum.strftime("%d-%m-%Y")
vandaag_altijd_vandaag = _vandaag_datum.strftime("%d-%m-%Y")  # altijd vandaag, voor last-minute


def parse_uur_waarde(val):
    """
    Zet een celwaarde om naar een uur als float.
    Werkt met: int (10), float (17.5), str ('10', '17,5', '17:30'),
               datetime.time (10:00:00 → 10.0, 17:30:00 → 17.5)
    Geeft None terug als de waarde niet parseerbaar is.
    """
    if val is None:
        return None
    import datetime as _dt
    if isinstance(val, _dt.time):
        return val.hour + val.minute / 60 + val.second / 3600
    if isinstance(val, (int, float)):
        return float(val)
    s = str(val).strip()
    if ":" in s:
        parts = s.split(":")
        try:
            return int(parts[0]) + int(parts[1]) / 60
        except:
            return None
    s = s.replace(",", ".").replace("u", "").strip()
    try:
        return float(s)
    except:
        return None
        

def formatteer_uur(u):
    """10 → '10u', 13.5 → '13u30', 9.25 → '9u15'"""
    uren = int(u)
    minuten = round((u - uren) * 60)
    if minuten == 0:
        return f"{uren}u"
    return f"{uren}u{minuten:02d}"


# Kolom→uur mapping vanuit Input_ rij 2 (I2:S2)
# Nodig voor samenvoeg-attracties en dichte uren, die vóór open_uren gelezen worden
col_to_uur_speciaal = {}
for _kol in range(9, 20):
    _uur = parse_uur_waarde(ws_speciaal.cell(2, _kol).value)
    if _uur is not None and _uur != 0:
        col_to_uur_speciaal[_kol] = _uur
        

def parse_blok_duur(label_str):
    """Parst '0,5h', '0,75h', '1,5h' → float. Geeft 1.0 als niet parseerbaar."""
    if not label_str:
        return 1.0
    s = str(label_str).strip().replace(",", ".").replace("h", "").strip()
    try:
        return float(s)
    except ValueError:
        return 1.0



# Blokduur per uur (uit I1:S1 van Input_)
# Kolommen zonder label krijgen standaard 1.0 uur
blok_durations = {}
for _kol in range(9, 20):  # kolom I t/m S
    _uur = col_to_uur_speciaal.get(_kol)
    if _uur is None:
        continue
    _label = ws_speciaal.cell(1, _kol).value
    blok_durations[_uur] = parse_blok_duur(_label)


ws_studenten = wb["Studenten"]


# -----------------------------

# Hulpfuncties
# -----------------------------
def max_consecutive_hours(urenlijst):
    if not urenlijst:
        return 0
    urenlijst = sorted(set(urenlijst))
    maxr = huidig = 1
    for i in range(1, len(urenlijst)):
        huidig = huidig + 1 if urenlijst[i] == urenlijst[i-1] + 1 else 1
        maxr = max(maxr, huidig)
    return maxr



def compute_ideal_moments(open_uren):
    """
    Ideaalmomenten vallen na elke 3 blokken (op basis van positie, niet uren).
    Als het totaal aantal blokken deelbaar is door 3, zijn er geen ideaalmomenten nodig.

    Voorbeeld: 9 blokken → deelbaar door 3 → lege set
    Voorbeeld: 8 blokken → ideaalmomenten op blok 4 en blok 7 (index 3 en 6)
    """
    if not open_uren:
        return set()
    blokken = sorted(open_uren)
    if len(blokken) < 4:
        return set()
    return {blokken[i] for i in range(3, len(blokken), 3)}


def partition_run_lengths(run_hours, ideal_moments=None):
    """
    Splits een run van blokken op in stukken van [3, 2, 4, 1].
    Gebruikt blokposities (via run_hours lijst), niet uuraritmethiek.
    """
    L = len(run_hours)
    blocks = [3, 2, 4, 1]
    use_ideal = bool(ideal_moments) and L % 3 != 0

    if use_ideal:
        INF = 10 ** 9
        dp = [(INF, INF, [])] * (L + 1)
        dp[0] = (0, 0, [])
        for i in range(1, L + 1):
            best = (INF, INF, [])
            for b in blocks:
                if i - b < 0:
                    continue
                prev_neg, prev_ones, prev_blks = dp[i - b]
                if prev_neg == INF:
                    continue
                # Startuur van dit blok rechtstreeks uit run_hours
                blok_start_uur = run_hours[i - b]
                is_ideal = (i - b > 0) and (blok_start_uur in ideal_moments)
                cand = (
                    prev_neg - (1 if is_ideal else 0),
                    prev_ones + (1 if b == 1 else 0),
                    prev_blks + [b],
                )
                if (cand[0], cand[1]) < (best[0], best[1]):
                    best = cand
            dp[i] = best
        return dp[L][2]
    else:
        dp = [(10 ** 9, [])] * (L + 1)
        dp[0] = (0, [])
        for i in range(1, L + 1):
            best = (10 ** 9, [])
            for b in blocks:
                if i - b < 0:
                    continue
                prev_ones, prev_blks = dp[i - b]
                cand = (prev_ones + (1 if b == 1 else 0), prev_blks + [b])
                if cand < best:
                    best = cand
            dp[i] = best
        return dp[L][1]


def contiguous_runs(sorted_hours):
    runs = []
    if not sorted_hours:
        return runs
    # Gebruik positie in open_uren, niet uur+1
    blok_index = {u: i for i, u in enumerate(sorted(open_uren))}
    run = [sorted_hours[0]]
    for h in sorted_hours[1:]:
        if blok_index.get(h, -99) == blok_index.get(run[-1], -98) + 1:
            run.append(h)
        else:
            runs.append(run)
            run = [h]
    runs.append(run)
    return runs

# Helpers die in meerdere delen gebruikt worden
def normalize_attr(name):
    """Normaliseer attractienaam zodat 'X 2' telt als 'X'; trim & lower-case voor vergelijking."""
    if not name:
        return ""
    s = str(name).strip()
    parts = s.rsplit(" ", 1)
    if len(parts) == 2 and parts[1].isdigit():
        s = parts[0]
    return s.strip().lower()

def parse_header_uur(header):
    """Map headertekst (bv. '14u', '14:00', '14:30') naar het hele uur (14)."""
    if not header:
        return None
    s = str(header).strip()
    try:
        if "u" in s:
            return int(s.split("u")[0])
        if ":" in s:
            uur, _min = s.split(":")
            return int(uur)
        return int(s)
    except:
        return None

# -----------------------------
# Studenten inlezen
# -----------------------------
studenten = []

# Lees attractienamen uit rij 1 van 'Studenten', kolommen G t/m X (7-24)
attractie_namen_studenten = [
    ws_studenten.cell(1, kol).value
    for kol in range(7, 25)
]

for rij in range(2, 500):
    naam = ws_studenten.cell(rij, 5).value  # kolom E
    if not naam:
        continue

    # Werkuren uit C (beginuur) en D (einduur)
    _begin = parse_uur_waarde(ws_studenten.cell(rij, 3).value)
    _eind  = parse_uur_waarde(ws_studenten.cell(rij, 4).value)
    if _begin is None or _eind is None:
        uren_beschikbaar = []
    else:
        uren_beschikbaar = [
            u for u in col_to_uur_speciaal.values()
            if _begin <= u and u + blok_durations.get(u, 1.0) <= _eind
        ]

    # Attracties uit kolommen G-X (7-24), namen staan in rij 1
    attracties = [
        attractie_namen_studenten[kol - 7]
        for kol in range(7, 25)
        if ws_studenten.cell(rij, kol).value in [1, True, "WAAR", "X"]
        and attractie_namen_studenten[kol - 7]
    ]

    # Aantal attracties uit kolom Z (26)
    try:
        aantal_attracties = int(ws_studenten.cell(rij, 26).value) if ws_studenten.cell(rij, 26).value else len(attracties)
    except:
        aantal_attracties = len(attracties)

    studenten.append({
        "naam": naam,
        "uren_beschikbaar": sorted(uren_beschikbaar),
        "attracties": attracties,
        "aantal_attracties": aantal_attracties,
        "is_pauzevlinder": False,
        "pv_number": None,
        "assigned_attracties": set(),
        "assigned_hours": []
    })

dichte_uren_per_attr = defaultdict(set)
# Input_: rijen 17 t/m 22, vakjes in I-S (kol 9-19), attractienaam in T (kol 20)

for rij in range(17, 23):  # rij 17 t/m 22
    attr_naam_raw = ws_speciaal.cell(rij, 20).value  # kolom T
    if attr_naam_raw:
        attr_naam = normalize_attr(attr_naam_raw)
        for col_idx in range(9, 20):  # kolom I t/m S
            val = ws_speciaal.cell(rij, col_idx).value
            if val in [1, True, "WAAR", "X"]:
                uur = col_to_uur_speciaal.get(col_idx)
                if uur:
                    dichte_uren_per_attr[attr_naam].add(uur)

# -----------------------------
# Samenvoeg-attracties (per uur)
# -----------------------------


# In DEEL 1 bij "Samenvoeg-attracties (per uur)"
uur_samenvoegingen = defaultdict(list)
# Input_: rijen 10 t/m 15, vakjes in I-S (kol 9-19), attractienamen in T-U-V (kol 20-22)

for rij in range(10, 16):  # rij 10 t/m 15
    groep = []
    for col in range(20, 23):  # kolom T, U, V
        val = ws_speciaal.cell(rij, col).value
        if val:
            groep.append(str(val).strip())

    if len(groep) > 1:
        for col_idx in range(9, 20):  # kolom I t/m S
            if ws_speciaal.cell(rij, col_idx).value in [1, True, "WAAR", "X"]:
                uur = col_to_uur_speciaal.get(col_idx)
                if uur:
                    uur_samenvoegingen[uur].append(groep)


# -----------------------------
# Alle mogelijke samengevoegde attracties (namen)
# -----------------------------

samengevoegde_attracties = set()

for groepen in uur_samenvoegingen.values():
    for groep in groepen:
        samengevoegde_attracties.add(" + ".join(groep))



# -----------------------------
# Voeg samengestelde attracties toe aan individuele studenten
# -----------------------------
for s in studenten:
    huidige = set(s["attracties"])
    for sameng in samengevoegde_attracties:
        onderdelen = [a.strip() for a in sameng.split("+")]
        if all(o in huidige for o in onderdelen):
            s["attracties"].append(sameng)  # voeg de samengestelde attractie toe




# -----------------------------
# Openingsuren
# -----------------------------
# Openingsuren lezen vanuit Input_, I2:S2 (kolommen 9-19)
open_uren = []
uur_labels = {}

for kol in range(9, 20):
    val = ws_speciaal.cell(2, kol).value
    uur = parse_uur_waarde(val)
    if uur is None or uur == 0:
        continue
    open_uren.append(uur)
    label = ws_speciaal.cell(1, kol).value
    if label and str(label).strip():
        uur_labels[uur] = str(label).strip()

if not open_uren:
    open_uren = list(range(10, 19))
open_uren = sorted(set(open_uren))
        

ideaalmomenten = compute_ideal_moments(open_uren)

# -----------------------------
# Sorteervolgorde studenten
# Eerst op aantal attracties,
# daarna op vaste tie-break regel uit BU2
# -----------------------------
bu2_waarde = ws_speciaal.cell(3, 23).value  # W3 in Input_
try:
    tie_break_mode = int(bu2_waarde)
except:
    tie_break_mode = 1

if tie_break_mode not in [1, 2, 3, 4, 5]:
    tie_break_mode = 1


def naam_tie_break_key(naam_raw):
    naam = str(naam_raw).strip().lower()

    if tie_break_mode == 1:
        # gewone alfabetische volgorde
        return naam

    elif tie_break_mode == 2:
        # omgekeerde alfabetische volgorde
        return "".join(chr(255 - ord(c)) for c in naam)

    elif tie_break_mode == 3:
        # eerst op aantal letters, daarna alfabetisch
        return (len(naam), naam)

    elif tie_break_mode == 4:
        # alfabetisch op basis van laatste letters
        return naam[::-1]

    elif tie_break_mode == 5:
        # omgekeerde van mode 4
        return "".join(chr(255 - ord(c)) for c in naam[::-1])

    return naam




# -----------------------------

# Pauzevlinders
# -----------------------------
# Input_: C14 t/m C18 (kolom 3, rijen 14-18)
pauzevlinder_namen = [
    ws_speciaal.cell(rij, 3).value
    for rij in range(14, 19)
    if ws_speciaal.cell(rij, 3).value
]

# Pauzevlinderuren lezen vanuit Input_, I3:S3 (kolommen 9-19)
required_pauze_hours = []
for kol in range(9, 20):
    val = ws_speciaal.cell(3, kol).value
    try:
        uur = int(val)
    except (TypeError, ValueError):
        continue
    if uur and uur in open_uren:
        required_pauze_hours.append(uur)
required_pauze_hours = sorted(set(required_pauze_hours))


for idx,pvnaam in enumerate(pauzevlinder_namen,start=1):
    for s in studenten:
        if s["naam"]==pvnaam:
            s["is_pauzevlinder"]=True
            s["pv_number"]=idx
            s["uren_beschikbaar"]=[u for u in s["uren_beschikbaar"] if u not in required_pauze_hours]
            break

# Maak 'selected' lijst van pauzevlinders (dicts met naam en attracties)
selected = [s for s in studenten if s.get("is_pauzevlinder")]
selected = sorted(selected, key=lambda s: naam_tie_break_key(s["naam"]))

# -----------------------------
# Attracties & aantallen (raw)
# -----------------------------
aantallen_raw = {}
attracties_te_plannen = []
for rij in range(3, 21):  # E3:F20 in Aanpassingen
    naam = ws_aanpassingen.cell(rij, 5).value  # kolom E
    if naam:
        try:
            aantal = int(ws_aanpassingen.cell(rij, 6).value)  # kolom F
        except:
            aantal = 0
        aantallen_raw[naam] = max(0, min(2, aantal))
        if aantallen_raw[naam] >= 1:
            attracties_te_plannen.append(naam)

# Priority order for second spots (column BA, rows 5-11)
second_priority_order = [
    ws_aanpassingen.cell(rij, 9).value  # kolom I
    for rij in range(3, 13)             # I3:I12
    if ws_aanpassingen.cell(rij, 9).value
]


# -----------------------------
# Attractielijst uitbreiden met samengevoegde attracties (globaal)
# -----------------------------

for nieuwe in samengevoegde_attracties:
    if nieuwe not in attracties_te_plannen:
        attracties_te_plannen.append(nieuwe)
    aantallen_raw[nieuwe] = 1


# -----------------------------
# Actieve attracties per uur (ivm samenvoegingen)
# -----------------------------

actieve_attracties_per_uur = {}
# Gebruik de raw aantallen als basis
aantallen = {uur: {a: aantallen_raw.get(a, 1) for a in attracties_te_plannen} for uur in open_uren}

for uur in open_uren:
    actief = set()
    # 1. Voeg eerst alle individuele attracties toe die NIET dicht zijn
    for a in attracties_te_plannen:
        if " + " in a: continue # Sla samengevoegde namen hier nog even over
        
        if uur in dichte_uren_per_attr.get(normalize_attr(a), set()):
            aantallen[uur][a] = 0
        else:
            actief.add(a)

    # 2. Verwerk de samenvoegingen voor dit specifieke uur
    huidige_groepen = uur_samenvoegingen.get(uur, [])
    for groep in huidige_groepen:
        samengevoegde_naam = " + ".join(groep)
        
        # Voeg de samengevoegde attractie toe aan de planning
        actief.add(samengevoegde_naam)
        aantallen[uur][samengevoegde_naam] = 1
        
        # VERWIJDER de onderdelen uit de actieve lijst (voorkomt dubbele telling)
        for onderdeel in groep:
            if onderdeel in actief:
                actief.remove(onderdeel)
            aantallen[uur][onderdeel] = 0

    actieve_attracties_per_uur[uur] = actief



### -----------------------------
### Compute aantallen per hour + red spots (GEÏNTEGREERD)
### -----------------------------
red_spots = {uur: set() for uur in open_uren}          
second_spot_blocked = {uur: set() for uur in open_uren}  

for uur in open_uren:
    # 1. Hoeveel studenten zijn er dit uur echt beschikbaar? [1]
    student_count = sum(
        1 for s in studenten
        if uur in s["uren_beschikbaar"] and not (
            s["is_pauzevlinder"] and uur in required_pauze_hours
        )
    )
    
    # 2. Hoeveel attracties moeten dit uur minimaal 1 persoon hebben? [1]
    # We kijken naar de actieve lijst van dat uur (rekening houdend met uitschakelingen/samenvoegingen)
    base_spots = sum(1 for a in actieve_attracties_per_uur[uur] if aantallen[uur].get(a, 0) >= 1)
    
    # 3. Bereken het overschot
    extra_spots = student_count - base_spots

    # 4. Verdeel de tweede plekken op basis van de prioriteitslijst uit Excel (BA5:BA11) [2]
    for attr in second_priority_order:
        # Check of de attractie dit uur actief is én of hij normaal 2 personen nodig heeft [2, 3]
        if attr in actieve_attracties_per_uur[uur] and aantallen_raw.get(attr) == 2:
            if extra_spots > 0:
                # Er is nog een student over voor een tweede plek
                aantallen[uur][attr] = 2
                extra_spots -= 1
            else:
                # Geen studenten meer over? Blokkeer de tweede plek voor dit uur
                second_spot_blocked[uur].add(attr)
                aantallen[uur][attr] = 1  # Forceer het aantal voor dit uur naar 1


# -----------------------------
# Red spots for samengestelde attracties
# -----------------------------

for uur in open_uren:
    # Groepen die dit uur samengevoegd zijn
    groepen = uur_samenvoegingen.get(uur, [])

    # Samengestelde attracties die DIT uur actief zijn
    samengestelde = set(" + ".join(g) for g in groepen)

    # Losse attracties die in een samenvoeging zitten
    losse_in_samenvoeging = set(a for g in groepen for a in g)

    # 1️⃣ Samenvoeging actief → losse attracties verbieden
    for attr in losse_in_samenvoeging:
        red_spots[uur].add(attr)

    # 2️⃣ Samenvoeging NIET actief → samenvoeging verbieden
    for samengestelde_attr in samengevoegde_attracties:
        if samengestelde_attr not in samengestelde:
            red_spots[uur].add(samengestelde_attr)


# -----------------------------
# Studenten die effectief inzetbaar zijn
# -----------------------------
studenten_workend = [
    s for s in studenten if any(u in open_uren for u in s["uren_beschikbaar"])
]


# -----------------------------
# Blacklist van attracties per student (BB16:BG79)
# -----------------------------
student_blacklist = defaultdict(set)

for rij in range(3, 26):  # O3:P25 in Aanpassingen
    naam = ws_aanpassingen.cell(rij, 15).value  # kolom O
    if not naam:
        continue
    naam = str(naam).strip()
    attr = ws_aanpassingen.cell(rij, 16).value  # kolom P (1 attractie)
    if attr:
        student_blacklist[naam].add(str(attr).strip().lower())


# Sorteer attracties op "kritieke score" (hoeveel studenten ze kunnen doen)
def kritieke_score(attr, studenten_list):
    return sum(1 for s in studenten_list if attr in s["attracties"])

attracties_te_plannen.sort(key=lambda a: kritieke_score(a, studenten_workend))



# -----------------------------
# Assign per student
# -----------------------------
assigned_map = defaultdict(list)  # (uur, attr) -> list of student-names
per_hour_assigned_counts = {uur: {a: 0 for a in attracties_te_plannen} for uur in open_uren}
extra_assignments = defaultdict(list)


# -----------------------------
# Overbodige pauzevlinderuren → laatste PV naar extra
# -----------------------------
import math

aantal_pv = len(selected)
aantal_pauze_uren = len(required_pauze_hours)
afgekapte_pv_uren = set()  # nieuw: bijhouden welke uren afgekapt worden

if aantal_pv > 0 and aantal_pauze_uren > 0:
    plaatsen_pauzeplanning = (aantal_pauze_uren * 4 - 1) * aantal_pv

    try:
        lange_pauzes = int(ws_speciaal.cell(15, 6).value) if ws_speciaal.cell(15, 6).value else 0  # F15
    except:
        lange_pauzes = 0
    try:
        korte_pauzes = int(ws_speciaal.cell(16, 6).value) if ws_speciaal.cell(16, 6).value else 0  # F16
    except:
        korte_pauzes = 0

    pauze_kwartieren = 2 * lange_pauzes + korte_pauzes
    open_spots = plaatsen_pauzeplanning - pauze_kwartieren
    overbodige_uren = max(0, math.floor((open_spots - aantal_pv * 3) / 4))

    if overbodige_uren > 0:
        laatste_pv = selected[-1]
        pv_pauze_uren = sorted(required_pauze_hours, reverse=True)
        uren_te_verschuiven = min(overbodige_uren, len(pv_pauze_uren))

        for i in range(uren_te_verschuiven):
            uur = pv_pauze_uren[i]
            extra_assignments[uur].append(laatste_pv["naam"])
            afgekapte_pv_uren.add(uur)  # nieuw: registreer dit uur als afgeknipt


def herbereken_afgekapte_pv_uren(absentees_set=None, base_maps=None):
    """
    Herleidt afgekapte_pv_uren op basis van de huidige selected-lijst.
    Raakt extra_assignments niet aan.
    In last-minute context: geef absentees_set en base_maps mee voor
    correcte pauzetelling (i.p.v. stale BP2/BQ2 cellen).
    """
    global afgekapte_pv_uren
    afgekapte_pv_uren = set()

    _aantal_pv = len(selected)
    _aantal_pauze_uren = len(required_pauze_hours)
    if _aantal_pv == 0 or _aantal_pauze_uren == 0:
        return

    _plaatsen = (_aantal_pauze_uren * 4 - 1) * _aantal_pv

    if absentees_set is not None and base_maps is not None:
        _lange, _korte = lm5_bereken_pauze_counts(absentees_set, base_maps)
    else:
        try:
            _lange = int(ws_speciaal.cell(15, 6).value) if ws_speciaal.cell(15, 6).value else 0
        except:
            _lange = 0
        try:
            _korte = int(ws_speciaal.cell(16, 6).value) if ws_speciaal.cell(16, 6).value else 0
        except:
            _korte = 0

    _open_spots = _plaatsen - (2 * _lange + _korte)
    _overbodige = max(0, math.floor((_open_spots - _aantal_pv * 3) / 4))

    if _overbodige > 0:
        _pv_pauze_uren = sorted(required_pauze_hours, reverse=True)
        for uur in _pv_pauze_uren[:min(_overbodige, len(_pv_pauze_uren))]:
            afgekapte_pv_uren.add(uur)




MAX_CONSEC = 4
MAX_PER_STUDENT_ATTR = 6


# -----------------------------
# Vaste dag-attracties (BG–BI)
# -----------------------------

vaste_plaatsingen = []  # lijst van dicts: {naam, attractie}

for rij in range(3, 6):  # R3:T5 in Aanpassingen
    if ws_aanpassingen.cell(rij, 18).value in [1, True, "WAAR", "X"]:  # kolom R
        naam = ws_aanpassingen.cell(rij, 19).value       # kolom S
        attractie = ws_aanpassingen.cell(rij, 20).value  # kolom T
        if naam and attractie:
            vaste_plaatsingen.append({
                "naam": str(naam).strip(),
                "attractie": str(attractie).strip()
            })

# -----------------------------
# Vaste plaatsingen toepassen
# -----------------------------

for vp in vaste_plaatsingen:
    student = next((s for s in studenten if s["naam"] == vp["naam"]), None)
    if not student:
        continue

    attr = vp["attractie"]

    # effectieve werkuren van deze student
    uren = [
        u for u in student["uren_beschikbaar"]
        if u in open_uren
        and not (student["is_pauzevlinder"] and u in required_pauze_hours)
    ]

    for uur in uren:
        # attractie moet dit uur actief zijn
        if attr not in actieve_attracties_per_uur.get(uur, set()):
            continue

        # rode attracties overslaan
        if attr in red_spots.get(uur, set()):
            continue

        # capaciteit check
        max_spots = aantallen[uur].get(attr, 1)
        if attr in second_spot_blocked.get(uur, set()):
            max_spots = 1

        if per_hour_assigned_counts[uur][attr] >= max_spots:
            continue

        # plaats student
        assigned_map[(uur, attr)].append(student["naam"])
        per_hour_assigned_counts[uur][attr] += 1
        student["assigned_hours"].append(uur)
        student["assigned_attracties"].add(attr)

    # student mag niet meer door de normale planner
    student["uren_beschikbaar"] = []





def student_tie_break_key(student):
    return naam_tie_break_key(student["naam"])

def is_werkende_pauzevlinder(s):
    """Geeft 1 als de student een pauzevlinder is die nog wél werkt, anders 0."""
    if s["is_pauzevlinder"] and any(u in open_uren for u in s["uren_beschikbaar"]):
        return 1
    return 0

studenten_sorted = sorted(
    studenten_workend,
    key=lambda s: (is_werkende_pauzevlinder(s), s["aantal_attracties"], student_tie_break_key(s))
)

# -----------------------------
# Voorbereiden: expand naar posities per uur
# -----------------------------
positions_per_hour = {uur: [] for uur in open_uren}
for uur in open_uren:
    for attr in actieve_attracties_per_uur[uur]:
        max_pos = aantallen[uur].get(attr, 1)
        for pos in range(1, max_pos+1):
            # sla rode posities over
            if attr in second_spot_blocked[uur] and pos == 2:
                continue
            positions_per_hour[uur].append((attr, pos))
# -----------------------------
# occupied_positions vullen op basis van bestaande assigned_map
# -----------------------------
occupied_positions = {uur: {} for uur in open_uren}

for (uur, attr), namen in assigned_map.items():
    for idx, naam in enumerate(namen, start=1):
        occupied_positions[uur][(attr, idx)] = naam


# -----------------------------
# Hulpfunctie: kan blok geplaatst worden?
# -----------------------------
def can_place_block(student, block_hours, attr):
    for h in block_hours:
        # check of attractie beschikbaar is in dit uur
        if (attr, 1) not in positions_per_hour[h] and (attr, 2) not in positions_per_hour[h]:
            return False
        # alle posities bezet?
        taken1 = (attr,1) in occupied_positions[h]
        taken2 = (attr,2) in occupied_positions[h]
        if taken1 and taken2:
            return False
    return True

# -----------------------------
# Plaats blok
# -----------------------------
def place_block(student, block_hours, attr):
    for h in block_hours:
        # kies positie: eerst pos1, anders pos2
        if (attr,1) in positions_per_hour[h] and (attr,1) not in occupied_positions[h]:
            pos = 1
        else:
            pos = 2
        occupied_positions[h][(attr,pos)] = student["naam"]
        assigned_map[(h, attr)].append(student["naam"])
        student["assigned_hours"].append(h)
    student["assigned_attracties"].add(attr)


# =============================
# FLEXIBELE BLOKKEN & PLAATSLOGICA
# =============================

def student_kan_attr(student, attr):
    if " + " not in attr:
        # check blacklist
        if attr.lower() in student_blacklist.get(student["naam"], set()):
            return False
        return attr in student["attracties"]
    onderdelen = [a.strip() for a in attr.split("+")]
    # check elk onderdeel tegen blacklist
    for o in onderdelen:
        if o.lower() in student_blacklist.get(student["naam"], set()):
            return False
    return all(o in student["attracties"] for o in onderdelen)


def _max_spots_for(attr, uur):
    """Houd rekening met red_spots: 2e plek dicht als het rood is."""
    max_spots = aantallen[uur].get(attr, 1)
    if attr in second_spot_blocked.get(uur, set()):
        max_spots = 1
    return max_spots

def _has_capacity(attr, uur):
    if attr in red_spots.get(uur, set()):
        return False
    return per_hour_assigned_counts[uur][attr] < _max_spots_for(attr, uur)


def uren_bij_basis_attr(student, attr):
    """
    Tel alle uren die een student bij een basisattractie heeft gestaan,
    inclusief uren bij de samengevoegde versie of losse onderdelen.
    - "A"   → telt ook uren bij "A+B"
    - "A+B" → telt ook uren bij "A" en bij "B"
    """
    gerelateerd = {attr}

    if " + " in attr:
        for onderdeel in attr.split(" + "):
            gerelateerd.add(onderdeel.strip())
    else:
        for sameng in samengevoegde_attracties:
            onderdelen = [o.strip() for o in sameng.split(" + ")]
            if attr.strip() in onderdelen:
                gerelateerd.add(sameng)

    uren = set()
    for h in student["assigned_hours"]:
        for rel_attr in gerelateerd:
            if student["naam"] in assigned_map.get((h, rel_attr), []):
                uren.add(h)
                break
    return uren


def _try_place_block_on_attr(student, block_hours, attr):
    # Capaciteit check
    for h in block_hours:
        if not _has_capacity(attr, h):
            return False

    # ── AANGEPAST: tel ook uren bij verwante attracties mee ──
    uren_bij_attr = uren_bij_basis_attr(student, attr)

    # Check max 6 unieke uren per basisattractie per dag
    nieuwe_uren = set(block_hours)
    totaal_uren = uren_bij_attr | nieuwe_uren
    if len(totaal_uren) > 6:
        return False

    # Plaatsen
    for h in block_hours:
        assigned_map[(h, attr)].append(student["naam"])
        per_hour_assigned_counts[h][attr] += 1
        student["assigned_hours"].append(h)

    student["assigned_attracties"].add(attr)
    return True



def _try_place_block_any_attr(student, block_hours):
    def candidate_score(attr):
        schaarste = sum(1 for s in studenten_workend if attr in s["attracties"])

        # ── AANGEPAST: tel ook uren bij verwante attracties mee ──
        bestaande_uren = uren_bij_basis_attr(student, attr)
        totaal_na_plaatsing = len(bestaande_uren | set(block_hours))
        reeds_gebruikt = attr in student["assigned_attracties"] or bool(bestaande_uren)

        breedte_profiel = student.get("aantal_attracties", len(student.get("attracties", [])))

        fairness_straf = 0
        if totaal_na_plaatsing > 4:
            if breedte_profiel >= 6:    fairness_straf = 100
            elif breedte_profiel >= 5:  fairness_straf = 60
            elif breedte_profiel >= 4:  fairness_straf = 25

        hergebruik_straf = 1 if reeds_gebruikt else 0
        huidige_uren_op_attr = len(bestaande_uren)

        return (fairness_straf, hergebruik_straf, huidige_uren_op_attr, schaarste, attr)

    candidate_attrs = [
        a for a in attracties_te_plannen
        if student_kan_attr(student, a)
    ]
    candidate_attrs.sort(key=candidate_score)

    for attr in candidate_attrs:
        if _try_place_block_on_attr(student, block_hours, attr):
            return True

    return False
    


def _try_place_block_samenvoeging_transitie(student, block_hours, respecteer_fairness=True):
    if len(block_hours) < 2:
        return False

    breedte_profiel = student.get("aantal_attracties", len(student.get("attracties", [])))

    def fairness_ok(basis_attr, nieuwe_uren):
        if not respecteer_fairness:
            return True
        bestaande = uren_bij_basis_attr(student, basis_attr)
        totaal = len(bestaande | set(nieuwe_uren))
        if totaal <= 4:
            return True
        # Zelfde drempels als candidate_score
        if breedte_profiel >= 6: return False
        if breedte_profiel >= 5: return False
        if breedte_profiel >= 4: return False
        return True  # weinig opties → toch toestaan

    # ── Geval 1: eerste uur/uren zijn samenvoeging ──
    eerste_uur = block_hours[0]
    for sameng_attr in actieve_attracties_per_uur.get(eerste_uur, set()):
        if " + " not in sameng_attr:
            continue
        if not student_kan_attr(student, sameng_attr):
            continue

        # Hoeveel opeenvolgende uren is de samenvoeging actief vanaf het begin?
        sameng_uren = []
        for h in block_hours:
            if sameng_attr in actieve_attracties_per_uur.get(h, set()):
                sameng_uren.append(h)
            else:
                break

        rest_uren = [h for h in block_hours if h not in sameng_uren]
        if not rest_uren:
            continue

        if not all(_has_capacity(sameng_attr, h) for h in sameng_uren):
            continue

        onderdelen = [o.strip() for o in sameng_attr.split("+")]

        for onderdeel in onderdelen:
            if not student_kan_attr(student, onderdeel):
                continue
            if not all(_has_capacity(onderdeel, h) for h in rest_uren):
                continue

            if len(uren_bij_basis_attr(student, onderdeel) | set(block_hours)) > 6:
                continue
            if not fairness_ok(onderdeel, block_hours):
                continue

            for h in sameng_uren:
                assigned_map[(h, sameng_attr)].append(student["naam"])
                per_hour_assigned_counts[h][sameng_attr] += 1
                student["assigned_hours"].append(h)
            student["assigned_attracties"].add(sameng_attr)

            for h in rest_uren:
                assigned_map[(h, onderdeel)].append(student["naam"])
                per_hour_assigned_counts[h][onderdeel] += 1
                student["assigned_hours"].append(h)
            student["assigned_attracties"].add(onderdeel)

            return True

    # ── Geval 2: laatste uur/uren zijn samenvoeging ──
    laatste_uur = block_hours[-1]
    for sameng_attr in actieve_attracties_per_uur.get(laatste_uur, set()):
        if " + " not in sameng_attr:
            continue
        if not student_kan_attr(student, sameng_attr):
            continue

        sameng_uren = []
        for h in reversed(block_hours):
            if sameng_attr in actieve_attracties_per_uur.get(h, set()):
                sameng_uren.insert(0, h)
            else:
                break

        vroege_uren = [h for h in block_hours if h not in sameng_uren]
        if not vroege_uren:
            continue

        if not all(_has_capacity(sameng_attr, h) for h in sameng_uren):
            continue

        onderdelen = [o.strip() for o in sameng_attr.split("+")]

        for onderdeel in onderdelen:
            if not student_kan_attr(student, onderdeel):
                continue
            if not all(_has_capacity(onderdeel, h) for h in vroege_uren):
                continue

            # ── AANGEPAST: gebruik uren_bij_basis_attr + fairness check ──
            if len(uren_bij_basis_attr(student, onderdeel) | set(block_hours)) > 6:
                continue
            if not fairness_ok(onderdeel, block_hours):
                continue

            for h in vroege_uren:
                assigned_map[(h, onderdeel)].append(student["naam"])
                per_hour_assigned_counts[h][onderdeel] += 1
                student["assigned_hours"].append(h)
            student["assigned_attracties"].add(onderdeel)

            for h in sameng_uren:
                assigned_map[(h, sameng_attr)].append(student["naam"])
                per_hour_assigned_counts[h][sameng_attr] += 1
                student["assigned_hours"].append(h)
            student["assigned_attracties"].add(sameng_attr)

            return True

    return False


def _place_block_with_fallback(student, hours_seq, preferred_sizes=None, reset_sizes=None):
    if not hours_seq:
        return []

    if preferred_sizes is None:
        preferred_sizes = [3, 2, 4, 1]

    # Na het eerste blok terugvallen op reset_sizes (of gewone volgorde als geen reset opgegeven)
    next_sizes = reset_sizes if reset_sizes is not None else preferred_sizes

    for size in preferred_sizes:
        if len(hours_seq) < size:
            continue
        block = hours_seq[:size]

        eerste_uur = block[0]
        laatste_uur = block[-1]

        heeft_samenvoeging = size > 1 and (
            any(" + " in attr and student_kan_attr(student, attr)
                for attr in actieve_attracties_per_uur.get(eerste_uur, set()))
            or any(" + " in attr and student_kan_attr(student, attr)
                   for attr in actieve_attracties_per_uur.get(laatste_uur, set()))
        )

        if heeft_samenvoeging:
            sameng_kandidaten = [
                attr
                for uur in [eerste_uur, laatste_uur]
                for attr in actieve_attracties_per_uur.get(uur, set())
                if " + " in attr and student_kan_attr(student, attr)
            ]
            beste_sameng_score = (
                min(kritieke_score(a, studenten_workend) for a in sameng_kandidaten)
                if sameng_kandidaten else float("inf")
            )
            reguliere_kandidaten = [
                a for a in attracties_te_plannen
                if student_kan_attr(student, a)
                and all(_has_capacity(a, h) for h in block)
            ]
            beste_regulier_score = (
                min(kritieke_score(a, studenten_workend) for a in reguliere_kandidaten)
                if reguliere_kandidaten else float("inf")
            )

            if beste_sameng_score <= beste_regulier_score:
                # Stap 1: transitie mét fairness
                if _try_place_block_samenvoeging_transitie(student, block):
                    return _place_block_with_fallback(student, hours_seq[size:], preferred_sizes=next_sizes)
                # Stap 2: normaal mét fairness
                if _try_place_block_any_attr(student, block):
                    return _place_block_with_fallback(student, hours_seq[size:], preferred_sizes=next_sizes)
                # Stap 3: transitie zónder fairness (laatste redmiddel voor deze size)
                if _try_place_block_samenvoeging_transitie(student, block, respecteer_fairness=False):
                    return _place_block_with_fallback(student, hours_seq[size:], preferred_sizes=next_sizes)
            else:
                # Stap 1: normaal mét fairness
                if _try_place_block_any_attr(student, block):
                    return _place_block_with_fallback(student, hours_seq[size:], preferred_sizes=next_sizes)
                # Stap 2: transitie mét fairness
                if _try_place_block_samenvoeging_transitie(student, block):
                    return _place_block_with_fallback(student, hours_seq[size:], preferred_sizes=next_sizes)
                # Stap 3: transitie zónder fairness (laatste redmiddel voor deze size)
                if _try_place_block_samenvoeging_transitie(student, block, respecteer_fairness=False):
                    return _place_block_with_fallback(student, hours_seq[size:], preferred_sizes=next_sizes)
        else:
            # Stap 1: normaal mét fairness
            if _try_place_block_any_attr(student, block):
                return _place_block_with_fallback(student, hours_seq[size:], preferred_sizes=next_sizes)

    # Noodventiel: eerste uur tijdelijk overslaan
    return [hours_seq[0]] + _place_block_with_fallback(student, hours_seq[1:], preferred_sizes=next_sizes)



    
# -----------------------------
# Nieuwe assign_student
# -----------------------------


def assign_student(s):
    uren = sorted(u for u in s["uren_beschikbaar"] if u in open_uren)
    if s["is_pauzevlinder"]:
        uren = [u for u in uren if u not in required_pauze_hours]

    if not uren:
        return

    runs = contiguous_runs(uren)

    for run in runs:
        L = len(run)
        if L % 3 != 0 and ideaalmomenten:
            blokken = partition_run_lengths(run, ideal_moments=ideaalmomenten)
            seen = []
            for b in blokken:
                if b not in seen:
                    seen.append(b)
            for b in [3, 2, 4, 1]:
                if b not in seen:
                    seen.append(b)
            preferred_sizes = seen
        else:
            preferred_sizes = [3, 2, 4, 1]
        unplaced = _place_block_with_fallback(s, run, preferred_sizes=preferred_sizes, reset_sizes=[3, 2, 4, 1])

        for h in unplaced:
            extra_assignments[h].append(s["naam"])


for s in studenten_sorted:
    assign_student(s)
    
# -----------------------------
# Post-processing: lege plekken opvullen door doorschuiven
# -----------------------------

def doorschuif_leegplek(uur, attr, pos_idx, student_naam, stap, max_stappen=5):
    if stap > max_stappen:
        return False
    namen = assigned_map.get((uur, attr), [])
    naam = namen[pos_idx-1] if pos_idx-1 < len(namen) else ""
    if naam:
        return False

    kandidaten = []
    for b_attr in attracties_te_plannen:
        b_namen = assigned_map.get((uur, b_attr), [])
        for b_pos, b_naam in enumerate(b_namen):
            if not b_naam or b_naam == student_naam:
                continue
            cand_student = next((s for s in studenten_workend if s["naam"] == b_naam), None)
            if not cand_student:
                continue
            # Mag deze student de lege attractie doen?
            if attr not in cand_student["attracties"]:
                continue
            # Mag de extra de vrijgekomen plek doen?
            extra_student = next((s for s in studenten_workend if s["naam"] == student_naam), None)
            if not extra_student:
                continue
            if b_attr not in extra_student["attracties"]:
                continue
            # Score: zo min mogelijk 1-uursblokken creëren
            uren_cand = sorted([u for u in cand_student["assigned_hours"] if u != uur] + [uur])
            uren_extra = sorted(extra_student["assigned_hours"] + [uur])
            def count_1u_blokken(uren):
                if not uren:
                    return 0
                runs = contiguous_runs(uren)
                return sum(1 for r in runs if len(r) == 1)
            score = count_1u_blokken(uren_cand) + count_1u_blokken(uren_extra)
            kandidaten.append((score, b_attr, b_pos, b_naam, cand_student))
    kandidaten.sort()

    for score, b_attr, b_pos, b_naam, cand_student in kandidaten:
        extra_student = next((s for s in studenten_workend if s["naam"] == student_naam), None)
        if not extra_student:
            continue
        # Voer de swap uit
        assigned_map[(uur, b_attr)][b_pos] = student_naam
        extra_student["assigned_hours"].append(uur)
        extra_student["assigned_attracties"].add(b_attr)
        per_hour_assigned_counts[uur][b_attr] += 0  # netto gelijk
        assigned_map[(uur, attr)].insert(pos_idx-1, b_naam)
        assigned_map[(uur, attr)] = assigned_map[(uur, attr)][:aantallen[uur].get(attr, 1)]
        cand_student["assigned_hours"].remove(uur)
        cand_student["assigned_attracties"].discard(b_attr)
        cand_student["assigned_hours"].append(uur)
        cand_student["assigned_attracties"].add(attr)
        per_hour_assigned_counts[uur][attr] += 0  # netto gelijk
        # Check of alles klopt (geen dubbele, geen restricties overtreden)
        # (optioneel: extra checks toevoegen)
        return True
    return False

max_iterations = 10
for _ in range(max_iterations):
    changes_made = False
    for uur in open_uren:
        for attr in actieve_attracties_per_uur[uur]:
            max_pos = aantallen[uur].get(attr, 1)
            if attr in red_spots.get(uur, set()):
                max_pos = 1
            for pos_idx in range(1, max_pos+1):
                namen = assigned_map.get((uur, attr), [])
                naam = namen[pos_idx-1] if pos_idx-1 < len(namen) else ""
                if naam:
                    continue
                # Probeer voor alle extra's op dit uur
                extras_op_uur = list(extra_assignments[uur])  # kopie ivm mutatie
                for extra_naam in extras_op_uur:
                    extra_student = next((s for s in studenten_workend if s["naam"] == extra_naam), None)
                    if not extra_student:
                        continue
                    if attr in extra_student["attracties"]:
                        # Kan direct geplaatst worden, dus hoort niet bij dit scenario
                        continue
                    # Probeer doorschuiven
                    if doorschuif_leegplek(uur, attr, pos_idx, extra_naam, 1, max_iterations):
                        extra_assignments[uur].remove(extra_naam)
                        changes_made = True
                        break  # stop met deze plek, ga naar volgende lege plek
    if not changes_made:
        break



# -----------------------------
# Post-processing: wissel laatste blok van 2 of 3 uren
# als iemand 5 of 6 uur op 1 attractie staat
# -----------------------------

vaste_studenten = {vp["naam"] for vp in vaste_plaatsingen}

def get_student_by_name(naam):
    return next((s for s in studenten_workend if s["naam"] == naam), None)

def get_student_attr_on_hour(student_naam, uur):
    for attr in actieve_attracties_per_uur.get(uur, set()):
        if student_naam in assigned_map.get((uur, attr), []):
            return attr
    return None

def get_hours_on_attr(student, attr):
    uren = []
    for uur in sorted(set(student["assigned_hours"])):
        if student["naam"] in assigned_map.get((uur, attr), []):
            uren.append(uur)
    return sorted(uren)

def get_runs_on_attr(student, attr):
    uren = get_hours_on_attr(student, attr)
    return contiguous_runs(uren)

def count_attr_switches(student):
    uur_attr = []
    for uur in sorted(set(student["assigned_hours"])):
        attr = get_student_attr_on_hour(student["naam"], uur)
        if attr:
            uur_attr.append((uur, attr))

    if not uur_attr:
        return 0

    switches = 0
    prev_attr = uur_attr[0][1]
    for _, attr in uur_attr[1:]:
        if attr != prev_attr:
            switches += 1
        prev_attr = attr
    return switches

def remove_assignment(student, uur, attr):
    namen = assigned_map.get((uur, attr), [])
    if student["naam"] in namen:
        namen.remove(student["naam"])
    if uur in student["assigned_hours"]:
        student["assigned_hours"].remove(uur)

def add_assignment(student, uur, attr):
    assigned_map[(uur, attr)].append(student["naam"])
    student["assigned_hours"].append(uur)
    student["assigned_attracties"].add(attr)

def rebuild_student_attrs(student):
    attrs = set()
    for uur in sorted(set(student["assigned_hours"])):
        attr = get_student_attr_on_hour(student["naam"], uur)
        if attr:
            attrs.add(attr)
    student["assigned_attracties"] = attrs

def is_valid_attr_for_student_on_hours(student, attr, uren):
    # vaste dagplaatsingen niet aanpassen
    if student["naam"] in vaste_studenten:
        return False

    # student moet attractie kunnen doen
    if not student_kan_attr(student, attr):
        return False

    # attractie moet op al die uren actief en geldig zijn
    for uur in uren:
        if attr not in actieve_attracties_per_uur.get(uur, set()):
            return False
        if attr in red_spots.get(uur, set()):
            return False

    return True

def respects_student_attr_rules(student, attr):
    uren = get_hours_on_attr(student, attr)
    if len(uren) > 6:
        return False
    return True

def can_swap_exact_block(student_a, attr_a, block_hours, student_b, attr_b):
    # zelfde student of zelfde attractie heeft geen zin
    if student_a["naam"] == student_b["naam"]:
        return False
    if attr_a == attr_b:
        return False

    # beide richtingen moeten kunnen
    if not is_valid_attr_for_student_on_hours(student_a, attr_b, block_hours):
        return False
    if not is_valid_attr_for_student_on_hours(student_b, attr_a, block_hours):
        return False

    # student_b moet op exact deze uren ook éénzelfde blok hebben op attr_b
    for uur in block_hours:
        if student_b["naam"] not in assigned_map.get((uur, attr_b), []):
            return False
        # en niet tegelijk nog ergens anders zitten
        current_attr = get_student_attr_on_hour(student_b["naam"], uur)
        if current_attr != attr_b:
            return False

    # student_a moet natuurlijk ook exact daar staan
    for uur in block_hours:
        if student_a["naam"] not in assigned_map.get((uur, attr_a), []):
            return False
        current_attr = get_student_attr_on_hour(student_a["naam"], uur)
        if current_attr != attr_a:
            return False

    return True

def count_problem_attrs(student):
    """
    Tel voor hoeveel attracties deze student meer dan 4 uur ingepland staat.
    """
    count = 0
    for attr in list(student["assigned_attracties"]):
        if len(get_hours_on_attr(student, attr)) > 4:
            count += 1
    return count

def total_overflow_hours(student):
    """
    Tel hoeveel uren boven de limiet van 4 uur deze student in totaal heeft.
    Voorbeeld:
    - 5 uur op een attractie => +1
    - 6 uur op een attractie => +2
    """
    overflow = 0
    for attr in list(student["assigned_attracties"]):
        uren = len(get_hours_on_attr(student, attr))
        if uren > 4:
            overflow += (uren - 4)
    return overflow

def can_use_block_as_swap_target(student, attr, block_hours):
    """
    Check of student op exact deze uren op exact dezelfde attractie staat.
    """
    for uur in block_hours:
        if student["naam"] not in assigned_map.get((uur, attr), []):
            return False
        huidige_attr = get_student_attr_on_hour(student["naam"], uur)
        if huidige_attr != attr:
            return False
    return True

def try_swap_specific_block(student, attr, block_hours):
    """
    Probeer één specifiek blok (eerste OF laatste) van student/attr te wisselen.
    Alleen als:
    - het blok 2 of 3 uur lang is
    - de andere student op exact die uren ook één blok op één attractie heeft
    - alle regels geldig blijven
    - geen geïsoleerde 1-uursblokken ontstaan
    - max 2 extra wissels in totaal
    - het totaal aantal >4u-problemen niet stijgt
    - en liefst daalt
    """
    if len(block_hours) not in [2, 3]:
        return False

    orig_switches_a = count_attr_switches(student)
    orig_problem_count_a = count_problem_attrs(student)
    orig_overflow_a = total_overflow_hours(student)

    eerste_uur = block_hours[0]
    kandidaten = []

    for andere_student in studenten_workend:
        if andere_student["naam"] == student["naam"]:
            continue
        if andere_student["naam"] in vaste_studenten:
            continue

        attr_b = get_student_attr_on_hour(andere_student["naam"], eerste_uur)
        if not attr_b or attr_b == attr:
            continue

        # Andere student moet exact op dit hele blok op dezelfde attractie staan
        if not can_use_block_as_swap_target(andere_student, attr_b, block_hours):
            continue

        # Beide studenten moeten elkaars attractie op die uren mogen doen
        if not is_valid_attr_for_student_on_hours(student, attr_b, block_hours):
            continue
        if not is_valid_attr_for_student_on_hours(andere_student, attr, block_hours):
            continue

        kandidaten.append((andere_student["naam"], attr_b, andere_student))

    for _, attr_b, andere_student in kandidaten:
        orig_switches_b = count_attr_switches(andere_student)
        orig_problem_count_b = count_problem_attrs(andere_student)
        orig_overflow_b = total_overflow_hours(andere_student)

        # --- tijdelijke swap uitvoeren ---
        for uur in block_hours:
            remove_assignment(student, uur, attr)
            remove_assignment(andere_student, uur, attr_b)

        for uur in block_hours:
            add_assignment(student, uur, attr_b)
            add_assignment(andere_student, uur, attr)

        rebuild_student_attrs(student)
        rebuild_student_attrs(andere_student)

        valid = True

        # Geen geïsoleerde 1-uursblokken laten ontstaan
        def heeft_geisoleerd_uur(s, a):
            runs = get_runs_on_attr(s, a)
            return any(len(r) == 1 for r in runs)

        if heeft_geisoleerd_uur(student, attr):
            valid = False
        if heeft_geisoleerd_uur(student, attr_b):
            valid = False
        if heeft_geisoleerd_uur(andere_student, attr):
            valid = False
        if heeft_geisoleerd_uur(andere_student, attr_b):
            valid = False

        # Regels voor beide studenten / beide attracties
        for s, a in [
            (student, attr),
            (student, attr_b),
            (andere_student, attr),
            (andere_student, attr_b),
        ]:
            if not respects_student_attr_rules(s, a):
                valid = False

        # Max 2 extra wissels in totaal
        new_switches_a = count_attr_switches(student)
        new_switches_b = count_attr_switches(andere_student)
        extra_wissels = (new_switches_a - orig_switches_a) + (new_switches_b - orig_switches_b)

        if extra_wissels > 2:
            valid = False

        # Problemen na swap
        new_problem_count_a = count_problem_attrs(student)
        new_problem_count_b = count_problem_attrs(andere_student)
        new_overflow_a = total_overflow_hours(student)
        new_overflow_b = total_overflow_hours(andere_student)

        orig_total_problem_count = orig_problem_count_a + orig_problem_count_b
        new_total_problem_count = new_problem_count_a + new_problem_count_b

        orig_total_overflow = orig_overflow_a + orig_overflow_b
        new_total_overflow = new_overflow_a + new_overflow_b

        # Geen nieuw probleem creëren
        if new_total_problem_count > orig_total_problem_count:
            valid = False

        # Geen grotere overschrijding creëren
        if new_total_problem_count == orig_total_problem_count and new_total_overflow > orig_total_overflow:
            valid = False

        # Moet minstens iets verbeteren
        verbetering = (
            (new_total_problem_count < orig_total_problem_count)
            or (
                new_total_problem_count == orig_total_problem_count
                and new_total_overflow < orig_total_overflow
            )
        )

        if not verbetering:
            valid = False

        if valid:
            return True

        # --- rollback ---
        for uur in block_hours:
            remove_assignment(student, uur, attr_b)
            remove_assignment(andere_student, uur, attr)

        for uur in block_hours:
            add_assignment(student, uur, attr)
            add_assignment(andere_student, uur, attr_b)

        rebuild_student_attrs(student)
        rebuild_student_attrs(andere_student)

    return False


def try_swap_last_or_first_block(student, attr):
    """
    Probeer het laatste of eerste blok op deze attractie te wisselen.
    Als een run langer is dan 3, pak dan de laatste 3 of laatste 2 uur eruit.
    Alleen relevant als student >4 uur op deze attractie staat.
    """
    uren_op_attr = get_hours_on_attr(student, attr)
    if len(uren_op_attr) <= 4:
        return False

    runs = get_runs_on_attr(student, attr)
    if not runs:
        return False

    laatste_run = runs[-1]
    eerste_run  = runs[0]

    def kandidaat_blokken(run):
        """Geef blokken van 2 of 3 uur terug vanuit deze run (einde eerst)."""
        blokken = []
        if len(run) >= 3:
            blokken.append(run[-3:])  # laatste 3
        if len(run) >= 2:
            blokken.append(run[-2:])  # laatste 2
        if len(run) >= 3:
            blokken.append(run[:3])   # eerste 3
        if len(run) >= 2:
            blokken.append(run[:2])   # eerste 2
        return blokken

    # Probeer eerst blokken uit de laatste run
    for blok in kandidaat_blokken(laatste_run):
        if try_swap_specific_block(student, attr, blok):
            return True

    # Dan blokken uit de eerste run (als die anders is)
    if eerste_run != laatste_run:
        for blok in kandidaat_blokken(eerste_run):
            if try_swap_specific_block(student, attr, blok):
                return True

    return False


# Iteratief toepassen tot er niets meer verandert
max_block_swap_passes = 15
for _ in range(max_block_swap_passes):
    wijziging = False

    for student in studenten_workend:
        probleem_attracties = [
            a for a in list(student["assigned_attracties"])
            if len(get_hours_on_attr(student, a)) > 4
        ]

        # Eerst de zwaarste problemen proberen
        probleem_attracties.sort(
            key=lambda a: (
                -len(get_hours_on_attr(student, a)),
                -max(get_hours_on_attr(student, a))
            )
        )

        for attr in probleem_attracties:
            if try_swap_last_or_first_block(student, attr):
                wijziging = True
                break

    if not wijziging:
        break


# -----------------------------
# Volgorde attracties uit Input!BL16:BL33
# -----------------------------
input_volgorde = []
for rij in range(3, 21):  # C3:C20 in Aanpassingen
    waarde = ws_aanpassingen.cell(rij, 3).value  # kolom C
    if waarde:
        input_volgorde.append(str(waarde).strip())

# -----------------------------
# Alle attracties die minstens één keer actief zijn (voor output)
# -----------------------------
alle_actieve_attracties = set()
for uur in open_uren:
    alle_actieve_attracties |= actieve_attracties_per_uur.get(uur, set())

# Eerst de gewone attracties in de volgorde van BL16:BL33
geordende_attracties = [a for a in input_volgorde if a in alle_actieve_attracties]

# Samengevoegde attracties slim invoegen:
# bv. "A + B" direct na de laatste van A/B in de inputvolgorde
samengestelde_attracties = [a for a in alle_actieve_attracties if " + " in str(a)]
overige_attracties = [
    a for a in alle_actieve_attracties
    if a not in geordende_attracties and a not in samengestelde_attracties
]

for sameng in samengestelde_attracties:
    onderdelen = [x.strip() for x in str(sameng).split("+")]

    # Zoek de positie van het laatste onderdeel in de huidige lijst
    laatst_gevonden_index = -1
    for onderdeel in onderdelen:
        if onderdeel in geordende_attracties:
            idx = geordende_attracties.index(onderdeel)
            laatst_gevonden_index = max(laatst_gevonden_index, idx)

    if laatst_gevonden_index >= 0:
        geordende_attracties.insert(laatst_gevonden_index + 1, sameng)
    else:
        # Als geen enkel onderdeel in de inputvolgorde staat,
        # zet hem voorlopig bij de rest
        overige_attracties.append(sameng)

# Voeg tenslotte nog attracties toe die niet in BL16:BL33 stonden
alle_actieve_attracties = geordende_attracties + overige_attracties


# -----------------------------
# Output-fix: houd studenten zo veel mogelijk
# op dezelfde plek (1 of 2) per attractie over opeenvolgende uren
# -----------------------------
def stabiliseer_assigned_map_voor_output():
    """
    Deze functie verandert niets aan WIE waar staat,
    maar alleen in welke volgorde namen in assigned_map[(uur, attr)] staan.

    Doel:
    - Studenten zo veel mogelijk op dezelfde plek (1 of 2) houden over opeenvolgende uren.
    - Extra slim omgaan met uren waarop plek 2 later verdwijnt:
      als iemand doorloopt naar een volgend uur met slechts 1 plek,
      dan zetten we die student liefst al op plek 1 in het uur ervoor.
    """

    def get_namen_op_uur(attr, uur):
        namen = list(assigned_map.get((uur, attr), []))
        unieke_namen = []
        for naam in namen:
            if naam and naam not in unieke_namen:
                unieke_namen.append(naam)
        return unieke_namen

    def get_max_pos(attr, uur):
        max_pos = aantallen[uur].get(attr, 1)
        if attr in second_spot_blocked.get(uur, set()):
            max_pos = 1
        return max_pos

    def naam_staat_op_attr_in_volgend_uur(attr, huidig_uur, naam):
        volgende_uren = [u for u in sorted(open_uren) if u > huidig_uur]
        if not volgende_uren:
            return False
        volgend_uur = volgende_uren[0]
        return naam in get_namen_op_uur(attr, volgend_uur)

    def naam_moet_liefst_naar_plek1(attr, huidig_uur, naam):
        """
        True als deze naam in het volgende uur nog op dezelfde attractie staat
        én het volgende uur maar 1 plek heeft.
        Dan is het logisch om deze student nu al op plek 1 te zetten.
        """
        volgende_uren = [u for u in sorted(open_uren) if u > huidig_uur]
        if not volgende_uren:
            return False

        volgend_uur = volgende_uren[0]
        if get_max_pos(attr, volgend_uur) != 1:
            return False

        return naam in get_namen_op_uur(attr, volgend_uur)

    for attr in alle_actieve_attracties:
        vorige_slots = {1: None, 2: None}

        for uur in sorted(open_uren):
            namen = get_namen_op_uur(attr, uur)
            max_pos = get_max_pos(attr, uur)

            if not namen:
                assigned_map[(uur, attr)] = []
                vorige_slots = {1: None, 2: None}
                continue

            if max_pos <= 1:
                assigned_map[(uur, attr)] = [namen[0]]
                vorige_slots = {1: namen[0], 2: None}
                continue

            # Vanaf hier: 2 plekken beschikbaar
            slots = {1: None, 2: None}
            resterend = namen[:]

            # 1) Eerst vooruitkijken:
            # als een student in het volgende uur doorloopt terwijl daar nog maar 1 plek is,
            # dan krijgt die student nu voorrang op plek 1.
            voorkeursnaam_plek1 = None
            kandidaten_plek1 = [n for n in resterend if naam_moet_liefst_naar_plek1(attr, uur, n)]
            if len(kandidaten_plek1) == 1:
                voorkeursnaam_plek1 = kandidaten_plek1[0]
            elif len(kandidaten_plek1) > 1:
                # Als er meerdere kandidaten zijn:
                # geef voorkeur aan wie vorige uur al op plek 1 stond,
                # anders gewoon de eerste in de huidige lijst.
                if vorige_slots.get(1) in kandidaten_plek1:
                    voorkeursnaam_plek1 = vorige_slots.get(1)
                else:
                    voorkeursnaam_plek1 = kandidaten_plek1[0]

            if voorkeursnaam_plek1 in resterend:
                slots[1] = voorkeursnaam_plek1
                resterend.remove(voorkeursnaam_plek1)

            # 2) Daarna achterwaartse stabiliteit:
            # probeer dezelfde student op dezelfde plek te houden
            for pos in [1, 2]:
                if slots[pos] is not None:
                    continue
                vorige_naam = vorige_slots.get(pos)
                if vorige_naam in resterend:
                    slots[pos] = vorige_naam
                    resterend.remove(vorige_naam)

            # 3) Als plek 1 nog leeg is, geef lichte voorkeur aan iemand
            # die ook in het volgende uur op deze attractie blijft staan
            if slots[1] is None:
                doorlopers = [n for n in resterend if naam_staat_op_attr_in_volgend_uur(attr, uur, n)]
                if len(doorlopers) == 1:
                    slots[1] = doorlopers[0]
                    resterend.remove(doorlopers[0])
                elif len(doorlopers) > 1:
                    # behoud indien mogelijk de oude plek-1 volgorde
                    if vorige_slots.get(1) in doorlopers:
                        slots[1] = vorige_slots.get(1)
                        resterend.remove(vorige_slots.get(1))
                    else:
                        slots[1] = doorlopers[0]
                        resterend.remove(doorlopers[0])

            # 4) Vul de rest gewoon op
            for pos in [1, 2]:
                if slots[pos] is None and resterend:
                    slots[pos] = resterend.pop(0)

            nieuwe_volgorde = []
            if slots[1]:
                nieuwe_volgorde.append(slots[1])
            if slots[2]:
                nieuwe_volgorde.append(slots[2])

            assigned_map[(uur, attr)] = nieuwe_volgorde
            vorige_slots = {1: slots[1], 2: slots[2]}

stabiliseer_assigned_map_voor_output()


# -----------------------------

# Excel output
# -----------------------------
wb_out = Workbook()
ws_out = wb_out.active
ws_out.title = "Planning"

gray_fill = PatternFill(start_color="808080", fill_type="solid")

# Witte fill voor headers en attracties
white_fill = PatternFill(start_color="FFFFFF", fill_type="solid")
pv_fill = PatternFill(start_color="FFF2CC", fill_type="solid")
extra_fill = PatternFill(start_color="FCE4D6", fill_type="solid")
center_align = Alignment(horizontal="center", vertical="center")
thin_border = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin")
)

# Felle, maar lichte pastelkleuren (gelijkmatige felheid, veel variatie)
studenten_namen = sorted({s["naam"] for s in studenten})
# Pauzevlinders krijgen ook een kleur uit het schema
alle_namen = studenten_namen + [pv for pv in pauzevlinder_namen if pv not in studenten_namen]
# Unieke kleuren genereren: als er te weinig kleuren zijn, maak er meer met lichte variatie
base_colors = [
    "FFB3BA", "FFDFBA", "FFFFBA", "BAFFC9", "BAE1FF", "E0BBE4", "957DAD", "D291BC", "FEC8D8", "FFDFD3",
    "B5EAD7", "C7CEEA", "FFDAC1", "E2F0CB", "F6DFEB", "F9E2AE", "B6E2D3", "B6D0E2", "F6E2B3", "F7C5CC",
    "F7E6C5", "C5F7D6", "C5E6F7", "F7F6C5", "F7C5F7", "C5C5F7", "C5F7F7", "F7C5C5", "C5F7C5", "F7E2C5",
    "E2F7C5", "C5F7E2", "E2C5F7", "C5E2F7", "F7C5E2", "F7F7C5", "C5F7F7", "F7C5F7", "C5C5F7", "F7C5C5",
    "C5F7C5", "F7E2C5", "E2F7C5", "C5F7E2", "E2C5F7", "C5E2F7", "F7C5E2", "E2C5F7", "C5F7E2", "E2F7C5"
]
import colorsys
def pastel_variant(hex_color, variant):
    # hex_color: 'RRGGBB', variant: int
    r = int(hex_color[0:2], 16) / 255.0
    g = int(hex_color[2:4], 16) / 255.0
    b = int(hex_color[4:6], 16) / 255.0
    h, l, s = colorsys.rgb_to_hls(r, g, b)
    # kleine variatie in lichtheid en saturatie
    l = min(1, l + 0.03 * (variant % 3))
    s = max(0.3, s - 0.04 * (variant % 5))
    r2, g2, b2 = colorsys.hls_to_rgb(h, l, s)
    return f"{int(r2*255):02X}{int(g2*255):02X}{int(b2*255):02X}"

unique_colors = []
needed = len(alle_namen)
variant = 0
while len(unique_colors) < needed:
    for base in base_colors:
        if len(unique_colors) >= needed:
            break
        # voeg lichte variatie toe als nodig
        color = pastel_variant(base, variant) if variant > 0 else base
        if color not in unique_colors:
            unique_colors.append(color)
    variant += 1

student_kleuren = dict(zip(alle_namen, unique_colors))

ws_out.cell(1, 1, vandaag).font = Font(bold=True)
ws_out.cell(1, 1).fill = white_fill


for col_idx, uur in enumerate(sorted(open_uren), start=2):
    label = uur_labels.get(uur)
    header_text = f"{formatteer_uur(uur)} ({label})" if label else formatteer_uur(uur)
    ws_out.cell(1, col_idx, header_text).font = Font(bold=True)
    ws_out.cell(1, col_idx).fill = white_fill
    ws_out.cell(1, col_idx).alignment = center_align
    ws_out.cell(1, col_idx).border = thin_border

    

rij_out = 2
for attr in alle_actieve_attracties:
    # 1. Bepaal hoeveel rijen deze attractie nodig heeft (1 of 2 plekken)
    max_pos = max(
        max(aantallen[uur].get(attr, 1) for uur in open_uren),
        max(per_hour_assigned_counts[uur].get(attr, 0) for uur in open_uren)
    )

    for pos_idx in range(1, max_pos + 1):
        # --- LAYOUT: Naam gevolgd door spatie en nummer (zonder haakjes) ---
        display_name = f"{attr} {pos_idx}" if max_pos > 1 else attr
        ws_out.cell(rij_out, 1, display_name).font = Font(bold=True)
        ws_out.cell(rij_out, 1).fill = white_fill
        ws_out.cell(rij_out, 1).border = thin_border

        for col_idx, uur in enumerate(sorted(open_uren), start=2):
            cell = ws_out.cell(rij_out, col_idx)

            # Haal de studentnaam op voor dit uur en deze positie
            namen = assigned_map.get((uur, attr), [])
            naam = namen[pos_idx-1] if pos_idx-1 < len(namen) else ""

            # --- LOGICA VOOR GRIJS KLEUREN ---
            current_attr_norm = normalize_attr(attr)
            is_samengesteld = " + " in attr
            groepen_dit_uur = uur_samenvoegingen.get(uur, [])
            
            moet_grijs = False

            # A. Check of de attractie dit uur gesloten is
            if uur in dichte_uren_per_attr.get(current_attr_norm, set()):
                moet_grijs = True

            # B. Check voor samengestelde attracties (bv. 'A + B')
            elif is_samengesteld:
                # De samengevoegde rij is grijs als deze specifieke groep dit uur NIET actief is
                onderdelen_set = {normalize_attr(x.strip()) for x in attr.split("+")}
                actief_als_groep = any({normalize_attr(g) for g in groep} == onderdelen_set for groep in groepen_dit_uur)
                if not actief_als_groep:
                    moet_grijs = True

            # C. Check voor individuele attracties (bv. 'A')
            else:
                # De individuele rij wordt grijs als de attractie opgaat in een samenvoeging
                is_onderdeel_van_samenvoeging = any(current_attr_norm in [normalize_attr(g) for g in groep] for groep in groepen_dit_uur)
                if is_onderdeel_van_samenvoeging:
                    moet_grijs = True

            # D. Check of de tweede plek geblokkeerd is (red spots)
            if pos_idx == 2 and attr in second_spot_blocked.get(uur, set()):
                moet_grijs = True

            # --- Cel invullen en opmaken ---
            cell.value = naam
            cell.alignment = center_align
            cell.border = thin_border

            if moet_grijs:
                cell.fill = gray_fill  # Grijs uit je bronnen
            elif naam and naam in student_kleuren:
                cell.fill = PatternFill(start_color=student_kleuren[naam], fill_type="solid")
            else:
                cell.fill = white_fill

        rij_out += 1
        
# Pauzevlinders
rij_out += 1
pauzevlinder_namen_sorted = [pv["naam"] for pv in selected]

for pv_idx, pvnaam in enumerate(pauzevlinder_namen_sorted, start=1):
    ws_out.cell(rij_out, 1, f"Pauzevlinder {pv_idx}").font = Font(bold=True)
    ws_out.cell(rij_out, 1).fill = white_fill
    ws_out.cell(rij_out, 1).border = thin_border
    for col_idx, uur in enumerate(sorted(open_uren), start=2):
        # Laatste PV: toon niet bij de afgekapte uren
        if pvnaam == selected[-1]["naam"] and uur in afgekapte_pv_uren:
            naam = ""
        else:
            naam = pvnaam if uur in required_pauze_hours else ""
        ws_out.cell(rij_out, col_idx, naam).alignment = center_align
        ws_out.cell(rij_out, col_idx).border = thin_border
        if naam and naam in student_kleuren:
            ws_out.cell(rij_out, col_idx).fill = PatternFill(start_color=student_kleuren[naam], fill_type="solid")
    rij_out += 1

# Extra's per rij
rij_out += 1
extras_flat = []
for uur in sorted(open_uren):
    for naam in extra_assignments[uur]:
        if naam not in extras_flat:
            extras_flat.append(naam)
for extra_idx, naam in enumerate(extras_flat, start=1):
    ws_out.cell(rij_out, 1, f"Extra {extra_idx}").font = Font(bold=True)
    ws_out.cell(rij_out, 1).fill = white_fill
    ws_out.cell(rij_out, 1).border = thin_border
    for col_idx, uur in enumerate(sorted(open_uren), start=2):
        # Toon naam alleen als deze extra op dit uur is ingepland
        cell_naam = naam if naam in extra_assignments[uur] else ""
        ws_out.cell(rij_out, col_idx, cell_naam).alignment = center_align
        ws_out.cell(rij_out, col_idx).border = thin_border
        if cell_naam and cell_naam in student_kleuren:
            ws_out.cell(rij_out, col_idx).fill = PatternFill(start_color=student_kleuren[cell_naam], fill_type="solid")
    rij_out += 1

# Kolombreedte
for col in range(1, len(open_uren) + 2):
    ws_out.column_dimensions[get_column_letter(col)].width = 18

# ---- student_totalen beschikbaar maken voor volgende delen ----
from collections import defaultdict
student_totalen = defaultdict(int)
for row in ws_out.iter_rows(min_row=2, values_only=True):
    for naam in row[1:]:
        if naam and str(naam).strip() != "":
            student_totalen[naam] += 1


# -----------------------------
# Analyse-sheet maken indien nodig
# Alleen als er nog extra's zijn terwijl er elders echte lege plekken zijn
# -----------------------------

def heeft_echte_lege_plek():
    """
    True als er minstens 1 echte lege plek bestaat op de planning:
    - attractie is actief op dat uur
    - niet gesloten / niet red spot
    - geen geblokkeerde 2e plek
    - plaats is binnen de capaciteit
    - er staat nog niemand op die plek
    """
    for uur in open_uren:
        for attr in actieve_attracties_per_uur.get(uur, set()):
            if attr in red_spots.get(uur, set()):
                continue

            max_pos = aantallen[uur].get(attr, 1)
            if attr in second_spot_blocked.get(uur, set()):
                max_pos = 1

            namen = assigned_map.get((uur, attr), [])
            for pos_idx in range(1, max_pos + 1):
                naam = namen[pos_idx - 1] if pos_idx - 1 < len(namen) else ""
                if not naam:
                    return True
    return False


def heeft_extra_studenten():
    return any(len(namen) > 0 for namen in extra_assignments.values())


def student_is_aanwezig_op_uur_zonder_pauzevlinder(student, uur):
    """
    Student telt mee in analyse voor dit uur als:
    - student effectief aanwezig is op dit uur
      (ingepland of extra)
    - en NIET als pauzevlinder bezig is op dit uur
    """
    naam = student["naam"]

    # Pauzevlinder tijdens pauzevlinderuur telt niet mee
    if student.get("is_pauzevlinder") and uur in required_pauze_hours:
        return False

    if uur in set(student.get("assigned_hours", [])):
        return True

    if naam in extra_assignments.get(uur, []):
        return True

    return False


def student_kan_attr_in_analyse(student, attr):
    """
    Voor analyse:
    - respecteer blacklist
    - samengevoegde attractie mag enkel als student alle onderdelen kan
    """
    naam = student["naam"]

    if " + " not in attr:
        return attr.lower() not in student_blacklist.get(naam, set()) and attr in student.get("attracties", [])

    onderdelen = [a.strip() for a in attr.split("+")]
    for onderdeel in onderdelen:
        if onderdeel.lower() in student_blacklist.get(naam, set()):
            return False

    return all(onderdeel in student.get("attracties", []) for onderdeel in onderdelen)


def actieve_analyse_attracties_op_uur(uur, actieve_set=None):
    """
    Geeft attracties terug in de volgorde van Input!BL16:BL33,
    maar aangepast aan het specifieke uur:
    - losse attracties als ze actief zijn
    - samengevoegde attracties enkel als ze dat uur actief samengevoegd zijn
    Optioneel: geef actieve_set mee om de globale actieve_attracties_per_uur te overschrijven.
    """
    if actieve_set is None:
        actieve_set = actieve_attracties_per_uur.get(uur, set())

    input_volgorde_lokaal = []
    for rij_bl in range(16, 34):  # BL16 t.e.m. BL33
        attr = ws[f"BL{rij_bl}"].value
        if attr:
            input_volgorde_lokaal.append(str(attr).strip())

    resultaat = []
    gebruikte = set()

    # Eerst gewone attracties in inputvolgorde
    for attr in input_volgorde_lokaal:
        if attr in actieve_set and attr not in gebruikte:
            resultaat.append(attr)
            gebruikte.add(attr)

        # Kijk of een samengestelde attractie met dit onderdeel actief is op dit uur
        for actief_attr in actieve_set:
            if " + " not in str(actief_attr):
                continue
            onderdelen = [x.strip() for x in str(actief_attr).split("+")]
            if attr in onderdelen and actief_attr not in gebruikte:
                if all(o in input_volgorde_lokaal for o in onderdelen):
                    laatst_idx = max(input_volgorde_lokaal.index(o) for o in onderdelen)
                    huidig_idx = input_volgorde_lokaal.index(attr)
                    if huidig_idx == laatst_idx:
                        resultaat.append(actief_attr)
                        gebruikte.add(actief_attr)

    # Daarna nog eventuele actieve attracties die niet in BL-lijst zaten
    for attr in actieve_set:
        if attr not in gebruikte:
            resultaat.append(attr)
            gebruikte.add(attr)

    return resultaat


def maak_analyse_sheet(wb_arg, am_arg, ea_arg, st_arg, actieve_attracties_override=None):
    # Verwijder oud sheet
    if "Analyse" in wb_arg.sheetnames:
        del wb_arg["Analyse"]

    ws_analyse = wb_arg.create_sheet(title="Analyse")
    analyse_header_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    witte_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

    # Gebruik override als die meegegeven is (last-minute), anders globale data
    def actieve_set_voor_uur(uur):
        if actieve_attracties_override is not None:
            return actieve_attracties_override.get(uur, set())
        return actieve_attracties_per_uur.get(uur, set())

    # Herdefinieer de hulpfunctie lokaal zodat ze de juiste data gebruikt
    def is_aanwezig(student, uur):
        naam = student["naam"]
        if student.get("is_pauzevlinder") and uur in required_pauze_hours:
            return False
        if uur in set(student.get("assigned_hours", [])):
            return True
        if naam in ea_arg.get(uur, []):
            return True
        return False

    titel = "Hier zie je per uur welke studenten aanwezig zijn en welke attracties ze kunnen:"
    ws_analyse.merge_cells(start_row=1, start_column=1, end_row=1, end_column=20)
    titel_cel = ws_analyse.cell(1, 1, titel)
    titel_cel.font = Font(bold=True, size=12)
    titel_cel.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    titel_cel.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    titel_cel.border = thin_border

    start_rij = 3
    for uur in sorted(open_uren):
        actieve_set = actieve_set_voor_uur(uur)
        analyse_attracties_uur = actieve_analyse_attracties_op_uur(uur, actieve_set)

        analyse_studenten_uur = sorted(
            [s for s in st_arg if is_aanwezig(s, uur)],
            key=lambda s: (
                sum(1 for attr in analyse_attracties_uur if student_kan_attr_in_analyse(s, attr)),
                naam_tie_break_key(s["naam"])
            )
        )

        if not analyse_studenten_uur or not analyse_attracties_uur:
            continue

        ws_analyse.cell(start_rij, 1, formatteer_uur(uur)).font = Font(bold=True)
        ws_analyse.cell(start_rij, 1).fill = analyse_header_fill
        ws_analyse.cell(start_rij, 1).alignment = center_align
        ws_analyse.cell(start_rij, 1).border = thin_border
        ws_analyse.cell(start_rij, 2, "Student").font = Font(bold=True)
        ws_analyse.cell(start_rij, 2).fill = analyse_header_fill
        ws_analyse.cell(start_rij, 2).alignment = center_align
        ws_analyse.cell(start_rij, 2).border = thin_border

        start_col_attr = 3
        for idx, attr in enumerate(analyse_attracties_uur, start=start_col_attr):
            cel = ws_analyse.cell(start_rij, idx, attr)
            cel.font = Font(bold=True)
            cel.fill = analyse_header_fill
            cel.alignment = center_align
            cel.border = thin_border

        rij = start_rij + 1
        for s in analyse_studenten_uur:
            naam = s["naam"]
            ws_analyse.cell(rij, 1, rij - start_rij).alignment = center_align
            ws_analyse.cell(rij, 1).border = thin_border
            ws_analyse.cell(rij, 1).fill = witte_fill
            naam_cel = ws_analyse.cell(rij, 2, naam)
            naam_cel.alignment = center_align
            naam_cel.border = thin_border
            student_fill = PatternFill(start_color=student_kleuren[naam], fill_type="solid") if naam in student_kleuren else witte_fill
            naam_cel.fill = student_fill

            for idx, attr in enumerate(analyse_attracties_uur, start=start_col_attr):
                cel = ws_analyse.cell(rij, idx)
                cel.alignment = center_align
                cel.border = thin_border
                cel.font = Font(color="000000")
                if student_kan_attr_in_analyse(s, attr):
                    cel.value = attr
                    cel.fill = student_fill
                else:
                    cel.value = ""
                    cel.fill = witte_fill
            rij += 1

        ws_analyse.column_dimensions["A"].width = 8
        ws_analyse.column_dimensions["B"].width = 24
        for idx in range(start_col_attr, start_col_attr + len(analyse_attracties_uur)):
            ws_analyse.column_dimensions[get_column_letter(idx)].width = 13.5
        start_rij = rij + 3


maak_analyse_sheet(wb_out, assigned_map, extra_assignments, studenten)

#DEEL 2
#oooooooooooooooooooo
#oooooooooooooooooooo

# -----------------------------
# DEEL 2: Pauzevlinder overzicht
# -----------------------------
ws_pauze = wb_out.create_sheet(title="Pauzevlinders")

light_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
center_align = Alignment(horizontal="center", vertical="center")
thin_border = Border(left=Side(style="thin"), right=Side(style="thin"),
                     top=Side(style="thin"), bottom=Side(style="thin"))

# -----------------------------
# Rij 1: Uren
# -----------------------------
# Gebruik compute_pauze_hours/open_uren als basis voor de pauzeplanning-urenrij
uren_rij1 = []
from datetime import datetime, timedelta
if required_pauze_hours:
    start_uur = min(required_pauze_hours)
    eind_uur = max(required_pauze_hours)
    tijd = datetime(2020,1,1,start_uur,0)
    # Laatste pauze mag een kwartier vóór het einde starten
    laatste_pauze = datetime(2020,1,1,eind_uur,30)
    while tijd <= laatste_pauze:
        uren_rij1.append(f"{tijd.hour}u" if tijd.minute==0 else f"{tijd.hour}u{tijd.minute:02d}")
        tijd += timedelta(minutes=15)
else:
    # fallback: gebruik open_uren
    for uur in sorted(open_uren):
        uren_rij1.append(formatteer_uur(uur))

# Schrijf uren in rij 1, start in kolom B
for col_idx, uur in enumerate(uren_rij1, start=2):
    c = ws_pauze.cell(1, col_idx, uur)
    c.fill = light_fill
    c.alignment = center_align
    c.border = thin_border

### Zet de datum van vandaag in cel A1 van de pauzeplanning
a1 = ws_pauze.cell(1, 1, vandaag)
a1.font = Font(bold=True)
a1.fill = light_fill
a1.alignment = center_align
a1.border = thin_border

# -----------------------------
# Pauzevlinders en namen
# -----------------------------
rij_out = 2
for pv_idx, pv in enumerate(selected, start=1):
    # Titel: Pauzevlinder X
    title_cell = ws_pauze.cell(rij_out, 1, f"Pauzevlinder {pv_idx}")
    title_cell.font = Font(bold=True)
    title_cell.fill = light_fill
    title_cell.alignment = center_align
    title_cell.border = thin_border

    # Naam eronder (zelfde stijl en kleur)
    naam_cel = ws_pauze.cell(rij_out + 1, 1, pv["naam"])
    naam_cel.fill = light_fill
    naam_cel.alignment = center_align
    naam_cel.border = thin_border

    rij_out += 3  # lege rij ertussen

# -----------------------------
# Kolombreedte voor overzicht
# -----------------------------

# Automatisch de breedte van kolom A instellen op basis van de langste tekst
max_len_colA = 0
for row in range(1, ws_pauze.max_row + 1):
    val = ws_pauze.cell(row, 1).value
    if val:
        max_len_colA = max(max_len_colA, len(str(val)))
# Voeg wat extra ruimte toe
ws_pauze.column_dimensions['A'].width = max(12, max_len_colA + 2)

for col in range(2, len(uren_rij1) + 2):
    ws_pauze.column_dimensions[get_column_letter(col)].width = 10

# Gebruik exact dezelfde open_uren en headers als in deel 1 voor de pauzeplanning
uren_rij1 = []
for uur in sorted(open_uren):
    # Zoek de originele header uit ws_out (de hoofdplanning)
    for col in range(2, ws_out.max_column + 1):
        header = ws_out.cell(1, col).value
        if header and str(header).startswith(str(uur)):
            uren_rij1.append(header)
            break

# Opslaan met dezelfde unieke naam

# Maak in-memory bestand
output = BytesIO()





#oooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooo


#DEEL 3
#oooooooooooooooooooo
#oooooooooooooooooooo

import random
from collections import defaultdict
from openpyxl.styles import Alignment, Border, Side, PatternFill
import datetime

# -----------------------------
# DEEL 3: Extra naam voor pauzevlinders die langer dan 6 uur werken
# -----------------------------

# Sheet referenties
ws_planning = wb_out["Planning"]
ws_pauze = wb_out["Pauzevlinders"]

# Pauzekolommen (B–G in Pauzevlinders sheet)
# Dynamisch: alle kolommen waar in rij 1 een uur staat (bv. '13u45', '14u', ...)
pauze_cols = []
for col in range(2, ws_pauze.max_column + 1):
    header = ws_pauze.cell(1, col).value
    if header and ("u" in str(header)):
        pauze_cols.append(col)

# Bouw lijst met pauzevlinder-rijen
pv_rows = []
for pv in selected:
    row_found = None
    for r in range(2, ws_pauze.max_row + 1):
        if str(ws_pauze.cell(r, 1).value).strip() == str(pv["naam"]).strip():
            row_found = r
            break
    if row_found is not None:
        pv_rows.append((pv, row_found))

# Bereken totaal uren per student in Werkblad "Planning"
student_totalen = defaultdict(int)
for row in ws_planning.iter_rows(min_row=2, values_only=True):
    for naam in row[1:]:
        if naam and str(naam).strip() != "":
            student_totalen[naam] += 1

# Loop door pauzevlinders in Werkblad "Pauzevlinders"


# ---- OPTIMALISATIE: Meerdere planningen genereren en de beste kiezen ----
import copy
best_score = None
best_state = None
num_runs = 5
for _run in range(num_runs):
    # Maak een deep copy van de relevante werkbladen en variabelen
    ws_pauze_tmp = wb_out.copy_worksheet(ws_pauze)
    ws_pauze_tmp.title = f"Pauzevlinders_tmp_{_run}"
    # Reset alle naamcellen
    for pv, pv_row in pv_rows:
        for col in pauze_cols:
            ws_pauze_tmp.cell(pv_row, col).value = None
    # Herhaal de bestaande logica voor pauzeplanning, maar werk op ws_pauze_tmp
    # ...existing code for pauzeplanning, but use ws_pauze_tmp instead of ws_pauze...
    # (Voor deze patch: laat de bestaande logica staan, dit is een structuurvoorzet. Zie opmerking hieronder)
    # ---- Evalueer deze planning ----
    # 1. Iedereen een pauze?
    korte_pauze_ontvangers = set()
    for pv, pv_row in pv_rows:
        for col in pauze_cols:
            cel = ws_pauze_tmp.cell(pv_row, col)
            if cel.value and str(cel.value).strip() != "":
                # Check of het een korte pauze is (enkel blok, niet dubbel)
                idx = pauze_cols.index(col)
                is_lange = False
                if idx+1 < len(pauze_cols):
                    next_col = pauze_cols[idx+1]
                    cel_next = ws_pauze_tmp.cell(pv_row, next_col)
                    if cel_next.value == cel.value:
                        is_lange = True
                if idx > 0:
                    prev_col = pauze_cols[idx-1]
                    prev_cel = ws_pauze_tmp.cell(pv_row, prev_col)
                    if prev_cel.value == cel.value:
                        is_lange = True
                if not is_lange:
                    korte_pauze_ontvangers.add(str(cel.value).strip())
    alle_studenten = [s["naam"] for s in studenten if student_totalen.get(s["naam"], 0) >= 4]
    iedereen_pauze = all(naam in korte_pauze_ontvangers for naam in alle_studenten)
    # 2. Eerlijkheid: verschil max-min korte pauzes per pauzevlinder
    from collections import Counter
    pv_korte_pauze_count = Counter()
    for pv, pv_row in pv_rows:
        for col in pauze_cols:
            cel = ws_pauze_tmp.cell(pv_row, col)
            if cel.value and str(cel.value).strip() != "":
                idx = pauze_cols.index(col)
                is_lange = False
                if idx+1 < len(pauze_cols):
                    next_col = pauze_cols[idx+1]
                    cel_next = ws_pauze_tmp.cell(pv_row, next_col)
                    if cel_next.value == cel.value:
                        is_lange = True
                if idx > 0:
                    prev_col = pauze_cols[idx-1]
                    prev_cel = ws_pauze_tmp.cell(pv_row, prev_col)
                    if prev_cel.value == cel.value:
                        is_lange = True
                if not is_lange:
                    pv_korte_pauze_count[pv["naam"]] += 1
    if pv_korte_pauze_count:
        eerlijkheid = max(pv_korte_pauze_count.values()) - min(pv_korte_pauze_count.values())
    else:
        eerlijkheid = 999
    # Score: eerst iedereen_pauze, dan eerlijkheid
    score = (iedereen_pauze, -eerlijkheid)
    if (best_score is None) or (score > best_score):
        best_score = score
        best_state = copy.deepcopy(ws_pauze_tmp)

# Na alle runs: kopieer best_state naar ws_pauze
if best_state is not None:

    for pv, pv_row in pv_rows:
        for col in pauze_cols:
            ws_pauze.cell(pv_row, col).value = best_state.cell(pv_row, col).value

# ---- Verwijder tijdelijke werkbladen ----
tmp_sheets = [ws for ws in wb_out.worksheets if ws.title.startswith("Pauzevlinders_tmp")]
for ws in tmp_sheets:
    wb_out.remove(ws)

# ---- Lege naamcellen inkleuren ----
naam_leeg_fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
naam_leeg_fill_pp2 = naam_leeg_fill
thin_border = Border(left=Side(style="thin"), right=Side(style="thin"),
                     top=Side(style="thin"), bottom=Side(style="thin"))
center_align = Alignment(horizontal="center", vertical="center")

for pv, pv_row in pv_rows:
    for col in pauze_cols:
        if ws_pauze.cell(pv_row, col).value in [None, ""]:
            ws_pauze.cell(pv_row, col).fill = naam_leeg_fill






#ooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooo

# DEEL 4: Lange werkers (>6 uur) pauze inplannen – gegarandeerd
# -----------------------------

from openpyxl.styles import Alignment, Border, Side, PatternFill
import random  # <— toegevoegd voor willekeurige verdeling

thin_border = Border(left=Side(style="thin"), right=Side(style="thin"),
                     top=Side(style="thin"), bottom=Side(style="thin"))
center_align = Alignment(horizontal="center", vertical="center")
# Zachtblauw, anders dan je titelkleuren; alleen voor naamcellen
naam_leeg_fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")

# Alleen kolommen B..G
# Dynamisch: alle kolommen waar in rij 1 een uur staat (bv. '13u45', '14u', ...)
pauze_cols = []
for col in range(2, ws_pauze.max_column + 1):
    header = ws_pauze.cell(1, col).value
    if header and ("u" in str(header)):
        pauze_cols.append(col)


def is_student_extra(naam):
    """Check of student in Planning bij 'extra' staat (kolom kan 'Extra' zijn of specifieke marker)."""
    for row in range(2, ws_planning.max_row + 1):
        if ws_planning.cell(row, 1).value:  # rij met attractienaam
            for col in range(2, ws_planning.max_column + 1):
                if str(ws_planning.cell(row, col).value).strip().lower() == str(naam).strip().lower():
                    header = str(ws_planning.cell(1, col).value).strip().lower()
                    if "extra" in header:  # of een andere logica afhankelijk van hoe 'extra' wordt aangeduid
                        return True
    return False


def is_pauzevlinder(naam):
    """Is deze naam een pauzevlinder?"""
    return any(pv["naam"] == naam for pv in selected)



# ---- Helpers ----
def parse_header_uur(header):
    """Map headertekst (bv. '14u', '14:00', '14:30') naar het hele uur (14)."""
    if not header:
        return None
    s = str(header).strip()
    try:
        if "u" in s:
            # '14u' of '14u30' -> 14
            return int(s.split("u")[0])
        if ":" in s:
            # '14:00' of '14:30' -> 14 (halfuur koppelen aan het hele uur)
            uur, _min = s.split(":")
            return int(uur)
        return int(s)  # fallback
    except:
        return None

# ---- Pauze-restrictie: geen korte pauze in eerste 12 kwartieren van de pauzeplanning (tenzij <=6u open) ----
def get_verboden_korte_pauze_kolommen():
    """Geeft de kolomnummers van de eerste 12 kwartieren in ws_pauze (B t/m M)."""
    return list(range(2, 12))  # kolommen 2 t/m 11 (B t/m M)

def is_korte_pauze_toegestaan_col(col, student_naam=None):
    """
    Controleert of een korte pauze in deze kolom mag.
    Uitzondering: als een student vóór 15:00 stopt, mag de pauze ALTIJD.
    """
    if len(open_uren) <= 6:
        return True
    
    # Check of de student een vroege stopper is (stopt vóór 15u)
    if student_naam:
        werk_uren = get_student_work_hours(student_naam)
        # De check 'if werk_uren' voorkomt dat de code crasht bij een lege lijst
        if werk_uren and max(werk_uren) < 15:
            return True
            
    return col not in get_verboden_korte_pauze_kolommen()

def normalize_attr(name):
    """Normaliseer attractienaam zodat 'X 2' telt als 'X'; trim & lower-case voor vergelijking."""
    if not name:
        return ""
    s = str(name).strip()
    parts = s.rsplit(" ", 1)
    if len(parts) == 2 and parts[1].isdigit():
        s = parts[0]
    return s.strip().lower()

# Build: pauzevlinder-rijen en capability-sets
pv_rows = []  # lijst: (pv_dict, naam_rij_index)
pv_cap_set = {}  # pv-naam -> set genormaliseerde attracties
for pv in selected:
    # zoek de rij waar de pv-naam in kolom A staat
    row_found = None
    for r in range(2, ws_pauze.max_row + 1):
        if str(ws_pauze.cell(r, 1).value).strip() == str(pv["naam"]).strip():
            row_found = r
            break
    if row_found is not None:
        pv_rows.append((pv, row_found))
        pv_cap_set[pv["naam"]] = {normalize_attr(a) for a in pv.get("attracties", [])}


# -----------------------------
# DEEL 1.5: Samengevoegde attracties correct registreren
# -----------------------------
# Plaats dit nadat je de PV-capabilities hebt opgebouwd, bv. na:
# pv_cap_set[pv["naam"]] = {normalize_attr(a) for a in pv.get("attracties", [])}

for pv in selected:
    nieuwe_caps = set()
    for attr in pv.get("attracties", []):
        attr_norm = normalize_attr(attr)
        # Check: bevat '+' → samengevoegde attractie
        if "+" in attr_norm:
            # split en normaliseer elk onderdeel
            onderdelen = [normalize_attr(x) for x in attr_norm.split("+")]
            # als PV elk onderdeel kan, voeg samengevoegde attractie toe
            if all(x in pv_cap_set[pv["naam"]] for x in onderdelen):
                nieuwe_caps.add(attr_norm)  # hele samengestelde attractie toevoegen
        else:
            nieuwe_caps.add(attr_norm)
    # overschrijf set met de uitgebreide mogelijkheden
    pv_cap_set[pv["naam"]] = nieuwe_caps


# Lange werkers: namen-set voor snelle checks


lange_werkers = [
    s for s in studenten
    if student_totalen.get(s["naam"], 0) > 6
    and s["naam"] not in [pv["naam"] for pv in selected]
]
lange_werkers_names = {s["naam"] for s in lange_werkers}

def get_student_work_hours(naam):
    """Welke uren werkt deze student echt (zoals te zien in werkblad 'Planning')?"""
    uren = set()
    for col in range(2, ws_planning.max_column + 1):
        header = ws_planning.cell(1, col).value
        uur = parse_header_uur(header)
        if uur is None:
            continue
        # check of student in deze kolom ergens staat
        for row in range(2, ws_planning.max_row + 1):
            if ws_planning.cell(row, col).value == naam:
                uren.add(uur)
                break
    return sorted(uren)

def vind_attractie_op_uur(naam, uur):
    """Geef attractienaam (exact zoals in Planning-kolom A) waar student staat op dit uur; None als niet gevonden."""
    for col in range(2, ws_planning.max_column + 1):
        header = ws_planning.cell(1, col).value
        col_uur = parse_header_uur(header)
        if col_uur != uur:
            continue
        for row in range(2, ws_planning.max_row + 1):
            if ws_planning.cell(row, col).value == naam:
                return ws_planning.cell(row, 1).value
    return None

def pv_kan_attr(pv, attr):
    """Check of pv attr kan (met normalisatie, zodat 'X 2' telt als 'X')."""
    base = normalize_attr(attr)
    if base == "extra":
        return True
    return base in pv_cap_set.get(pv["naam"], set())

# Willekeurige volgorde van pauzeplekken (pv-rij x kolom) om lege cellen random te spreiden
slot_order = [(pv, pv_row, col) for (pv, pv_row) in pv_rows for col in pauze_cols]
random.shuffle(slot_order)  # <— kern om lege plekken later willekeurig te verspreiden

def plaats_student(student, harde_mode=False):
    """
    Plaats student bij een geschikte pauzevlinder in B..G op een uur waar student effectief werkt.
    - Overschrijven alleen in harde_mode én alleen als de huidige inhoud géén lange werker is.
    - Volgorde van slots is willekeurig (slot_order) zodat lege plekken random verdeeld blijven.
    - Korte pauze blijft altijd in dezelfde rij als de lange pauze.
    - Pauzevlinders krijgen hun korte pauze uitsluitend in hun eigen rij.
    """
    naam = student["naam"]
    werk_uren = get_student_work_hours(naam)
    werk_uren_set = set(werk_uren)
    if len(werk_uren) > 2:
        verboden_uren = {werk_uren[0], werk_uren[-1]}
    else:
        verboden_uren = set(werk_uren)

    pauze_cols_sorted = sorted(pauze_cols)
    uur_col_pairs = []
    for col in pauze_cols_sorted:
        col_header = ws_pauze.cell(1, col).value
        col_uur = parse_header_uur(col_header)
        if col_uur is not None and col_uur in werk_uren_set and col_uur not in verboden_uren:
            uur_col_pairs.append((col_uur, col))

    if not hasattr(plaats_student, "pauze_registry"):
        plaats_student.pauze_registry = {}
    reg = plaats_student.pauze_registry.setdefault(naam, {"lange": False, "korte": False})

    lange_pauze_opties = []
    for i in range(len(uur_col_pairs)-1):
        uur1, col1 = uur_col_pairs[i]
        uur2, col2 = uur_col_pairs[i+1]
        if col2 == col1 + 1:
            lange_pauze_opties.append((i, uur1, col1, uur2, col2))

    if not reg["lange"] and not heeft_al_lange_pauze(naam):
        for optie in lange_pauze_opties:
            i, uur1, col1, uur2, col2 = optie
            attr1 = vind_attractie_op_uur(naam, uur1)
            attr2 = vind_attractie_op_uur(naam, uur2)
            if not attr1 or not attr2:
                continue
            for (pv, pv_row, _) in slot_order:
                if not pv_kan_attr(pv, attr1) and not is_student_extra(naam):
                    continue
                cel1 = ws_pauze.cell(pv_row, col1)
                cel2 = ws_pauze.cell(pv_row, col2)
                boven_cel1 = ws_pauze.cell(pv_row-1, col1)
                boven_cel2 = ws_pauze.cell(pv_row-1, col2)
                if cel1.value in [None, ""] and cel2.value in [None, ""] and not heeft_al_lange_pauze(naam):
                    boven_cel1.value = attr1
                    boven_cel1.alignment = center_align
                    boven_cel1.border = thin_border
                    boven_cel2.value = attr2
                    boven_cel2.alignment = center_align
                    boven_cel2.border = thin_border
                    cel1.value = naam
                    cel1.alignment = center_align
                    cel1.border = thin_border
                    cel2.value = naam
                    cel2.alignment = center_align
                    cel2.border = thin_border
                    reg["lange"] = True
                    if not reg["korte"]:
                        found = False
                        for min_blokjes in list(range(10, len(uur_col_pairs)-i)) + list(range(9, 0, -1)):
                            for j in range(i+min_blokjes, len(uur_col_pairs)):
                                uur_kort, col_kort = uur_col_pairs[j]
                                if not is_korte_pauze_toegestaan_col(col_kort):
                                    continue
                                attr_kort = vind_attractie_op_uur(naam, uur_kort)
                                if not attr_kort:
                                    continue
                                if (not pv_kan_attr(pv, attr_kort)) and (not is_student_extra(naam)):
                                    continue
                                # Altijd dezelfde rij als de lange pauze
                                cel_kort = ws_pauze.cell(pv_row, col_kort)
                                boven_cel_kort = ws_pauze.cell(pv_row-1, col_kort)
                                if cel_kort.value in [None, ""]:
                                    boven_cel_kort.value = attr_kort
                                    boven_cel_kort.alignment = center_align
                                    boven_cel_kort.border = thin_border
                                    cel_kort.value = naam
                                    cel_kort.alignment = center_align
                                    cel_kort.border = thin_border
                                    reg["korte"] = True
                                    found = True
                                    return True
                                elif harde_mode:
                                    occupant = str(cel_kort.value).strip() if cel_kort.value else ""
                                    if occupant not in lange_werkers_names:
                                        boven_cel_kort.value = attr_kort
                                        boven_cel_kort.alignment = center_align
                                        boven_cel_kort.border = thin_border
                                        cel_kort.value = naam
                                        cel_kort.alignment = center_align
                                        cel_kort.border = thin_border
                                        reg["korte"] = True
                                        found = True
                                        return True
                            if found:
                                break
                    return True

                # -------------------------------------------------------
                # FIX 1: harde_mode lange pauze – korte pauze vastpinnen
                # aan pv_row, niet itereren over alle PV-rijen via slot_order
                # -------------------------------------------------------
                elif harde_mode:
                    occupant1 = str(cel1.value).strip() if cel1.value else ""
                    occupant2 = str(cel2.value).strip() if cel2.value else ""
                    if (occupant1 not in lange_werkers_names) and (occupant2 not in lange_werkers_names) and not heeft_al_lange_pauze(naam):
                        boven_cel1.value = attr1
                        boven_cel1.alignment = center_align
                        boven_cel1.border = thin_border
                        boven_cel2.value = attr2
                        boven_cel2.alignment = center_align
                        boven_cel2.border = thin_border
                        cel1.value = naam
                        cel1.alignment = center_align
                        cel1.border = thin_border
                        cel2.value = naam
                        cel2.alignment = center_align
                        cel2.border = thin_border
                        reg["lange"] = True
                        if not reg["korte"]:
                            for j in range(i+6, len(uur_col_pairs)):
                                uur_kort, col_kort = uur_col_pairs[j]
                                attr_kort = vind_attractie_op_uur(naam, uur_kort)
                                if not attr_kort:
                                    continue
                                # Was: for (pv2, pv_row2, _) in slot_order → vrije rij
                                # Nu:  altijd pv_row (zelfde rij als lange pauze)
                                if not pv_kan_attr(pv, attr_kort) and not is_student_extra(naam):
                                    continue
                                cel_kort = ws_pauze.cell(pv_row, col_kort)
                                boven_cel_kort = ws_pauze.cell(pv_row-1, col_kort)
                                if cel_kort.value in [None, ""]:
                                    boven_cel_kort.value = attr_kort
                                    boven_cel_kort.alignment = center_align
                                    boven_cel_kort.border = thin_border
                                    cel_kort.value = naam
                                    cel_kort.alignment = center_align
                                    cel_kort.border = thin_border
                                    reg["korte"] = True
                                    return True
                                elif harde_mode:
                                    occupant = str(cel_kort.value).strip() if cel_kort.value else ""
                                    if occupant not in lange_werkers_names:
                                        boven_cel_kort.value = attr_kort
                                        boven_cel_kort.alignment = center_align
                                        boven_cel_kort.border = thin_border
                                        cel_kort.value = naam
                                        cel_kort.alignment = center_align
                                        cel_kort.border = thin_border
                                        reg["korte"] = True
                                        return True
                        return True

    # -------------------------------------------------------
    # FIX 2: fallback – pauzevlinders krijgen uitsluitend hun
    # eigen rij; gewone studenten mogen elke vrije rij pakken
    # -------------------------------------------------------
    eigen_pv_row = next((row for (pv, row) in pv_rows if pv == naam), None)

    for uur in random.sample(werk_uren, len(werk_uren)):
        if uur in verboden_uren:
            continue
        attr = vind_attractie_op_uur(naam, uur)
        if not attr:
            continue
        for (pv, pv_row, col) in slot_order:
            col_header = ws_pauze.cell(1, col).value
            col_uur = parse_header_uur(col_header)
            if col_uur != uur:
                continue
            if not is_korte_pauze_toegestaan_col(col, naam):
                continue
            if not pv_kan_attr(pv, attr) and not is_student_extra(naam):
                continue
            # Pauzevlinder mag enkel in zijn eigen rij terechtkomen
            if eigen_pv_row is not None and pv_row != eigen_pv_row:
                continue
            cel = ws_pauze.cell(pv_row, col)
            boven_cel = ws_pauze.cell(pv_row - 1, col)
            current_val = cel.value
            if current_val in [None, ""]:
                if not reg["korte"]:
                    boven_cel.value = attr
                    boven_cel.alignment = center_align
                    boven_cel.border = thin_border
                    cel.value = naam
                    cel.alignment = center_align
                    cel.border = thin_border
                    reg["korte"] = True
                    return True
            else:
                if harde_mode:
                    occupant = str(current_val).strip()
                    if occupant not in lange_werkers_names:
                        if not reg["korte"]:
                            boven_cel.value = attr
                            boven_cel.alignment = center_align
                            boven_cel.border = thin_border
                            cel.value = naam
                            cel.alignment = center_align
                            cel.border = thin_border
                            reg["korte"] = True
                            return True
    return False
    
# ---- Fase 1: zachte toewijzing (niet overschrijven) ----
def heeft_al_lange_pauze(naam):
    # Check of student al een dubbele blok (lange pauze) heeft in het pauzeoverzicht
    for pv, pv_row in pv_rows:
        for idx, col in enumerate(pauze_cols):
            cel = ws_pauze.cell(pv_row, col)
            if cel.value == naam:
                # Check of volgende cel ook deze naam heeft (dubbele blok)
                if idx+1 < len(pauze_cols):
                    next_col = pauze_cols[idx+1]
                    cel_next = ws_pauze.cell(pv_row, next_col)
                    if cel_next.value == naam:
                        return True
    return False

lichtgroen_fill = PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid")  # lange pauze
lichtpaars_fill = PatternFill(start_color="E6DAF7", end_color="E6DAF7", fill_type="solid")  # kwartierpauze
naam_leeg_fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
lange_pauze_ontvangers = set()
# --- Verspreid lange pauzes van lange werkers net als bij pauzevlinders ---
niet_geplaatst = []
for s in random.sample(lange_werkers, len(lange_werkers)):
    naam = s["naam"]
    if naam in lange_pauze_ontvangers or heeft_al_lange_pauze(naam):
        if not plaats_student(s, harde_mode=False):
            niet_geplaatst.append(s)
        continue
    werk_uren = get_student_work_hours(naam)
    if len(werk_uren) <= 6:
        if not plaats_student(s, harde_mode=False):
            niet_geplaatst.append(s)
        continue
    # Alleen de eerste 11 kwartieren (indices 0 t/m 10) zijn toegestaan voor lange pauzes
    halve_uren = []  # lijst van (col1, col2, uur1, uur2, pv, pv_row)
    werk_uren_set = set(werk_uren)
    verboden_uren = {werk_uren[0], werk_uren[-1]} if len(werk_uren) > 2 else set(werk_uren)
    max_start_idx = min(8, len(pauze_cols)-2)  # idx 0 t/m 10 zijn halve uren binnen eerste 11 kwartieren
    for pv, pv_row in pv_rows:
        for idx in range(max_start_idx+1):
            col1 = pauze_cols[idx]
            col2 = pauze_cols[idx+1]
            col1_header = ws_pauze.cell(1, col1).value
            col2_header = ws_pauze.cell(1, col2).value
            # Alleen starten op heel of half uur
            try:
                min1 = int(str(col1_header).split('u')[1]) if 'u' in str(col1_header) and len(str(col1_header).split('u')) > 1 else 0
            except:
                min1 = 0
            if min1 not in (0, 30):
                continue
            uur1 = parse_header_uur(col1_header)
            uur2 = parse_header_uur(col2_header)
            if uur1 is None or uur2 is None:
                continue
            if uur1 not in werk_uren_set or uur2 not in werk_uren_set:
                continue
            if uur1 in verboden_uren or uur2 in verboden_uren:
                continue
            cel1 = ws_pauze.cell(pv_row, col1)
            cel2 = ws_pauze.cell(pv_row, col2)
            # Attractie moet kloppen
            attr1 = vind_attractie_op_uur(naam, uur1)
            attr2 = vind_attractie_op_uur(naam, uur2)
            if not attr1 or not attr2:
                continue
            if not pv_kan_attr(pv, attr1) and not is_student_extra(naam):
                continue
            if cel1.value in [None, ""] and cel2.value in [None, ""]:
                halve_uren.append((col1, col2, uur1, uur2, pv, pv_row))
    random.shuffle(halve_uren)
    # Fairness: keep a live counter of long breaks per pauzevlinder
    from collections import Counter
    if not hasattr(plaats_student, "pv_lange_pauze_count"):
        plaats_student.pv_lange_pauze_count = Counter()
    pv_lange_pauze_count = plaats_student.pv_lange_pauze_count
    for pv, _ in pv_rows:
        if pv["naam"] not in pv_lange_pauze_count:
            pv_lange_pauze_count[pv["naam"]] = 0
    geplaatst = False
    # Sorteer bij elke poging op actuele telling
    halve_uren_sorted = sorted(halve_uren, key=lambda x: pv_lange_pauze_count[x[4]["naam"]])
    for col1, col2, uur1, uur2, pv, pv_row in halve_uren_sorted:
        cel1 = ws_pauze.cell(pv_row, col1)
        cel2 = ws_pauze.cell(pv_row, col2)
        boven_cel1 = ws_pauze.cell(pv_row-1, col1)
        boven_cel2 = ws_pauze.cell(pv_row-1, col2)
        attr1 = vind_attractie_op_uur(naam, uur1)
        attr2 = vind_attractie_op_uur(naam, uur2)
        if cel1.value in [None, ""] and cel2.value in [None, ""] and not heeft_al_lange_pauze(naam):
            boven_cel1.value = attr1
            boven_cel1.alignment = center_align
            boven_cel1.border = thin_border
            boven_cel2.value = attr2
            boven_cel2.alignment = center_align
            boven_cel2.border = thin_border
            cel1.value = naam
            cel1.alignment = center_align
            cel1.border = thin_border
            cel1.fill = lichtgroen_fill
            cel2.value = naam
            cel2.alignment = center_align
            cel2.border = thin_border
            cel2.fill = lichtgroen_fill
            lange_pauze_ontvangers.add(naam)
            geplaatst = True
            # Fairness: niet meetellen als deze lange pauze (een van de twee blokken) een 'Extra' overname is
            if (normalize_attr(attr1) != 'extra') and (normalize_attr(attr2) != 'extra'):
                pv_lange_pauze_count[pv["naam"]] += 1
            break
    if not geplaatst:
        if not plaats_student(s, harde_mode=False):
            niet_geplaatst.append(s)

# ---- Fase 2: iteratief, met gecontroleerd overschrijven van niet-lange-werkers ----
# we herhalen meerdere passes om iedereen >6u te kunnen plaatsen
max_passes = 12
for _ in range(max_passes):
    if not niet_geplaatst:
        break
    rest = []
    # Ook hier willekeurige volgorde voor extra spreiding
    for s in random.sample(niet_geplaatst, len(niet_geplaatst)):
        if not plaats_student(s, harde_mode=True):
            rest.append(s)
    # Als niets veranderde in een hele pass, stoppen we
    if len(rest) == len(niet_geplaatst):
        break
    niet_geplaatst = rest

# ---- Lege naamcellen inkleuren (alleen de NAAM-rij; bovenliggende attractie-rij NIET kleuren) ----

# ---- Pauze-kleuren: lichtgroen voor lange pauze (dubbele blok), lichtpaars voor kwartierpauze (enkel blok) ----

lichtgroen_fill = PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid")  # lange pauze
lichtpaars_fill = PatternFill(start_color="E6DAF7", end_color="E6DAF7", fill_type="solid")  # kwartierpauze
roze_fill = PatternFill(start_color="FFD6E7", end_color="FFD6E7", fill_type="solid")

# Pauze kleuren invullen (lange en korte pauzes)
for pv, pv_row in pv_rows:
    for idx, col in enumerate(pauze_cols):
        cel = ws_pauze.cell(pv_row, col)
        if cel.value in [None, ""]:
            cel.fill = naam_leeg_fill
        else:
            # Check of dit een lange pauze is (dubbele blok: zelfde naam in 2 naast elkaar liggende cellen)
            is_langepauze = False
            # Kijk vooruit: als deze en de volgende cel dezelfde naam hebben, kleur beide groen
            if idx+1 < len(pauze_cols):
                next_col = pauze_cols[idx+1]
                cel_next = ws_pauze.cell(pv_row, next_col)
                if cel_next.value == cel.value and cel.value not in [None, ""]:
                    is_langepauze = True
                    cel.fill = lichtgroen_fill
                    cel_next.fill = lichtgroen_fill
                    continue  # sla de volgende cel over, die is al gekleurd
            # Kijk achteruit: als vorige cel al groen is door lange pauze, deze niet opnieuw kleuren
            if idx > 0:
                prev_col = pauze_cols[idx-1]
                prev_cel = ws_pauze.cell(pv_row, prev_col)
                if prev_cel.value == cel.value and cel.value not in [None, ""]:
                    # Deze cel is al als tweede helft van lange pauze gekleurd
                    continue
            # Anders: kwartierpauze
            cel.fill = lichtpaars_fill

# ---- Korte pauze voor pauzevlinders zelf toevoegen (eerst, met afstandscriterium) ----
def _pv_has_short_pause(naam, pv_row):
    for idx, col in enumerate(pauze_cols):
        if ws_pauze.cell(pv_row, col).value == naam:
            left_same = idx > 0 and ws_pauze.cell(pv_row, pauze_cols[idx-1]).value == naam
            right_same = idx+1 < len(pauze_cols) and ws_pauze.cell(pv_row, pauze_cols[idx+1]).value == naam
            if not left_same and not right_same:
                return True
    return False

def _pv_place_best_short(pv, pv_row, target_gap=10):
    """Plaats korte pauze op eigen rij met voorkeur: exact +10 blokken na lange pauze-einde,
    anders +11, +12, ...; als dat niet past, probeer +9, +8, ...; valt terug op globale laatste lange-pauze-einde als geen eigen lange pauze."""
    naam = pv["naam"]
    # Als er al een korte pauze staat, niets doen
    if _pv_has_short_pause(naam, pv_row):
        return False

    # Hulpfuncties
    def is_toegestaan_pv_col(col):
        if len(open_uren) <= 6:
            return True
        return col not in get_verboden_korte_pauze_kolommen()

    # Zoek eigen lange pauze-einde op deze rij
    lange_blok_einde = None
    i = 0
    while i < len(pauze_cols)-1:
        c1 = pauze_cols[i]
        c2 = pauze_cols[i+1]
        if ws_pauze.cell(pv_row, c1).value == naam and ws_pauze.cell(pv_row, c2).value == naam:
            lange_blok_einde = i+1
            # ga door; kies de laatste indien meerdere
            i += 2
        else:
            i += 1

    # Geen eigen lange pauze: kies een geldige plek op eigen rij (liefst vroegste index >= target_gap)
    if lange_blok_einde is None:
        werk_uren = get_student_work_hours(naam)
        if len(werk_uren) > 2:
            verboden_uren = {werk_uren[0], werk_uren[-1]}
        else:
            verboden_uren = set(werk_uren)
        candidates = []  # (prefer, idx, col, uur)
        for i, col in enumerate(pauze_cols):
            header = ws_pauze.cell(1, col).value
            uur = parse_header_uur(header)
            if uur not in werk_uren or uur in verboden_uren:
                continue
            if not is_toegestaan_pv_col(col):
                continue
            if ws_pauze.cell(pv_row, col).value not in [None, ""]:
                continue
            prefer = 1 if i >= target_gap else 0
            candidates.append((prefer, i, col, uur))
        if not candidates:
            return False
        # Kies met voorkeur voor index >= target_gap; binnen die groep de laatste (grootste index) om niet te vroeg te vallen
        candidates.sort(key=lambda x: (x[0], x[1]), reverse=True)
        _pref, i, col, uur = candidates[0]
        attr = vind_attractie_op_uur(naam, uur)
        if not attr:
            return False
        # Voor PV-korte pauzes: laat het vakje erboven leeg
        ws_pauze.cell(pv_row-1, col).value = None
        ws_pauze.cell(pv_row-1, col).alignment = center_align
        ws_pauze.cell(pv_row-1, col).border = thin_border
        cel = ws_pauze.cell(pv_row, col)
        cel.value = naam
        cel.fill = lichtpaars_fill
        cel.alignment = center_align
        cel.border = thin_border
        return True
    else:
        anchor_idx = lange_blok_einde

    if anchor_idx is None or anchor_idx < 0:
        # Geen anchor beschikbaar: kies de eerste toegestane lege cel op eigen rij (zeldzaam)
        werk_uren = get_student_work_hours(naam)
        if len(werk_uren) > 2:
            verboden_uren = {werk_uren[0], werk_uren[-1]}
        else:
            verboden_uren = set(werk_uren)
        for i, col in enumerate(pauze_cols):
            header = ws_pauze.cell(1, col).value
            uur = parse_header_uur(header)
            if uur not in werk_uren or uur in verboden_uren:
                continue
            if is_toegestaan_pv_col(col) and ws_pauze.cell(pv_row, col).value in [None, ""]:
                # schrijf bovenliggende attr
                attr = vind_attractie_op_uur(naam, uur)
                if not attr:
                    continue
                # Voor PV-korte pauzes: laat het vakje erboven leeg
                ws_pauze.cell(pv_row-1, col).value = None
                ws_pauze.cell(pv_row-1, col).alignment = center_align
                ws_pauze.cell(pv_row-1, col).border = thin_border
                cel = ws_pauze.cell(pv_row, col)
                cel.value = naam
                cel.fill = lichtpaars_fill
                cel.alignment = center_align
                cel.border = thin_border
                return True
        return False

    # Eerst exact +10 blokken, dan +11, +12, ...
    werk_uren = get_student_work_hours(naam)
    if len(werk_uren) > 2:
        verboden_uren = {werk_uren[0], werk_uren[-1]}
    else:
        verboden_uren = set(werk_uren)
    for d in range(target_gap, len(pauze_cols)-anchor_idx):
        idx = anchor_idx + d
        if idx >= len(pauze_cols):
            break
        col = pauze_cols[idx]
        if not is_toegestaan_pv_col(col):
            continue
        header = ws_pauze.cell(1, col).value
        uur = parse_header_uur(header)
        if uur not in werk_uren or uur in verboden_uren:
            continue
        if ws_pauze.cell(pv_row, col).value in [None, ""]:
            # schrijf bovenliggende attr
            attr = vind_attractie_op_uur(naam, uur)
            if not attr:
                continue
            # Voor PV-korte pauzes: laat het vakje erboven leeg
            ws_pauze.cell(pv_row-1, col).value = None
            ws_pauze.cell(pv_row-1, col).alignment = center_align
            ws_pauze.cell(pv_row-1, col).border = thin_border
            cel = ws_pauze.cell(pv_row, col)
            cel.value = naam
            cel.fill = lichtpaars_fill
            cel.alignment = center_align
            cel.border = thin_border
            return True

    # Dan lagere alternatieven: +9, +8, ... +1
    for d in range(target_gap-1, 0, -1):
        idx = anchor_idx + d
        if 0 <= idx < len(pauze_cols):
            col = pauze_cols[idx]
            if not is_toegestaan_pv_col(col):
                continue
            header = ws_pauze.cell(1, col).value
            uur = parse_header_uur(header)
            if uur not in werk_uren or uur in verboden_uren:
                continue
            if ws_pauze.cell(pv_row, col).value in [None, ""]:
                attr = vind_attractie_op_uur(naam, uur)
                if not attr:
                    continue
                # Voor PV-korte pauzes: laat het vakje erboven leeg
                ws_pauze.cell(pv_row-1, col).value = None
                ws_pauze.cell(pv_row-1, col).alignment = center_align
                ws_pauze.cell(pv_row-1, col).border = thin_border
                cel = ws_pauze.cell(pv_row, col)
                cel.value = naam
                cel.fill = lichtpaars_fill
                cel.alignment = center_align
                cel.border = thin_border
                return True

    return False

for pv, pv_row in pv_rows:
    _pv_place_best_short(pv, pv_row, target_gap=8)


# ---- Korte pauze voor ALLE studenten (ook <=6u, behalve pauzevlinders en lange werkers) ----
# --- Houd bij wie al een korte pauze heeft gekregen ---
korte_pauze_ontvangers = set()
# Zoek alle namen die al een korte pauze hebben in het pauzeoverzicht
for pv, pv_row in pv_rows:
    for col in pauze_cols:
        cel = ws_pauze.cell(pv_row, col)
        if cel.value and str(cel.value).strip() != "":
            # Check of het een korte pauze is (enkel blok, niet dubbel)
            idx = pauze_cols.index(col)
            is_lange = False
            if idx+1 < len(pauze_cols):
                next_col = pauze_cols[idx+1]
                cel_next = ws_pauze.cell(pv_row, next_col)
                if cel_next.value == cel.value:
                    is_lange = True
            if idx > 0:
                prev_col = pauze_cols[idx-1]
                prev_cel = ws_pauze.cell(pv_row, prev_col)
                if prev_cel.value == cel.value:
                    is_lange = True
            if not is_lange:
                korte_pauze_ontvangers.add(str(cel.value).strip())



# ---- Korte pauze voor ALLE studenten (ook <=6u, behalve pauzevlinders en lange werkers) ----
# --- Houd bij wie al een korte pauze heeft gekregen ---
korte_pauze_ontvangers = set()
# Zoek alle namen die al een korte pauze hebben in het pauzeoverzicht
for pv, pv_row in pv_rows:
    for col in pauze_cols:
        cel = ws_pauze.cell(pv_row, col)
        if cel.value and str(cel.value).strip() != "":
            # Check of het een korte pauze is (enkel blok, niet dubbel)
            idx = pauze_cols.index(col)
            is_lange = False
            if idx+1 < len(pauze_cols):
                next_col = pauze_cols[idx+1]
                cel_next = ws_pauze.cell(pv_row, next_col)
                if cel_next.value == cel.value:
                    is_lange = True
            if idx > 0:
                prev_col = pauze_cols[idx-1]
                prev_cel = ws_pauze.cell(pv_row, prev_col)
                if prev_cel.value == cel.value:
                    is_lange = True
            if not is_lange:
                korte_pauze_ontvangers.add(str(cel.value).strip())


# Nieuwe logica: eerlijke verdeling van korte pauzes over pauzevlinders
from collections import Counter

# Tel per pauzevlinder het aantal korte pauzes dat al is toegekend (EXTRA niet meetellen)
pv_korte_pauze_count = Counter()
for pv, pv_row in pv_rows:
    for col in pauze_cols:
        cel = ws_pauze.cell(pv_row, col)
        if cel.value and str(cel.value).strip() != "":
            # Check of het een korte pauze is (enkel blok, niet dubbel)
            idx = pauze_cols.index(col)
            is_lange = False
            if idx+1 < len(pauze_cols):
                next_col = pauze_cols[idx+1]
                cel_next = ws_pauze.cell(pv_row, next_col)
                if cel_next.value == cel.value:
                    is_lange = True
            if idx > 0:
                prev_col = pauze_cols[idx-1]
                prev_cel = ws_pauze.cell(pv_row, prev_col)
                if prev_cel.value == cel.value:
                    is_lange = True
            if not is_lange:
                # Kijk naar bovenliggende attractie; tel niet als dit 'Extra' is of leeg (zoals bij PV zelf)
                attr_above = ws_pauze.cell(pv_row-1, col).value
                if attr_above and normalize_attr(attr_above) != 'extra':
                    pv_korte_pauze_count[pv["naam"]] += 1

niet_geplaatste_korte_pauze = []

# --- NIEUW: Eerst korte pauzes toewijzen aan iedereen met een LANGE pauze,
# in volgorde van wie het LAATST z'n lange pauze had ---

def _has_long_pause(naam):
    for _pv, pv_row in pv_rows:
        for idx, col in enumerate(pauze_cols[:-1]):
            if ws_pauze.cell(pv_row, col).value == naam and ws_pauze.cell(pv_row, pauze_cols[idx+1]).value == naam:
                return True
    return False

def _last_long_pause_end_index(naam):
    """Geef de hoogste kolomindex (in pauze_cols) die een tweede helft is van een dubbele blok voor deze naam; -1 indien geen lange pauze."""
    last_idx = -1
    for _pv, pv_row in pv_rows:
        for idx in range(len(pauze_cols)-1):
            c1 = pauze_cols[idx]
            c2 = pauze_cols[idx+1]
            if ws_pauze.cell(pv_row, c1).value == naam and ws_pauze.cell(pv_row, c2).value == naam:
                last_idx = max(last_idx, idx+1)
    return last_idx

def _has_short_pause(naam):
    for _pv, pv_row in pv_rows:
        for idx, col in enumerate(pauze_cols):
            if ws_pauze.cell(pv_row, col).value == naam:
                # korte pauze: geen buur met dezelfde naam
                left_same = (idx > 0 and ws_pauze.cell(pv_row, pauze_cols[idx-1]).value == naam)
                right_same = (idx+1 < len(pauze_cols) and ws_pauze.cell(pv_row, pauze_cols[idx+1]).value == naam)
                if not left_same and not right_same:
                    return True
    return False

def _place_short_pause_for(naam):
    if _has_short_pause(naam):
        return True
    werk_uren = get_student_work_hours(naam)
    if not werk_uren:
        return False
    verboden_uren = {werk_uren[0], werk_uren[-1]} if len(werk_uren) > 2 else set(werk_uren)
    # Zoek anker = einde van eigen lange pauze (kolomindex in pauze_cols)
    def _last_long_end_index_for(naam):
        best = -1
        for _pv, _row in pv_rows:
            for idx in range(len(pauze_cols)-1):
                if ws_pauze.cell(_row, pauze_cols[idx]).value == naam and ws_pauze.cell(_row, pauze_cols[idx+1]).value == naam:
                    best = max(best, idx+1)
        return best
    anchor = _last_long_end_index_for(naam)

    # Helper om te checken en plaatsen op gewenste col
    def _try_place_at_col(col):
        header = ws_pauze.cell(1, col).value
        uur = parse_header_uur(header)
        if uur not in werk_uren or uur in verboden_uren:
            return False
        if not is_korte_pauze_toegestaan_col(col, naam):
            return False
        attr = vind_attractie_op_uur(naam, uur)
        if not attr:
            return False
        # verzamel geldige (pv_row) volgens regels
        rows = []
        for pv, pv_row in pv_rows:
            if is_pauzevlinder(naam) and pv["naam"] != naam:
                continue
            if ws_pauze.cell(pv_row, col).value not in [None, ""]:
                continue
            if not pv_kan_attr(pv, attr) and not is_student_extra(naam):
                continue
            rows.append((pv, pv_row))
        if not rows:
            return False
        # fairness op pv-rijen
        rows.sort(key=lambda r: pv_korte_pauze_count[r[0]["naam"]])
        pv, pv_row = rows[0]
        # Voor PV-korte pauzes: laat het vakje erboven leeg
        if is_pauzevlinder(naam):
            ws_pauze.cell(pv_row-1, col).value = None
        else:
            ws_pauze.cell(pv_row-1, col).value = attr
        ws_pauze.cell(pv_row-1, col).alignment = center_align
        ws_pauze.cell(pv_row-1, col).border = thin_border
        cel = ws_pauze.cell(pv_row, col)
        cel.value = naam
        cel.alignment = center_align
        cel.border = thin_border
        cel.fill = lichtpaars_fill
        # Niet meetellen als dit een EXTRA overname is of als de pauze voor een PV zelf is
        if (not is_pauzevlinder(naam)) and (normalize_attr(attr) != 'extra'):
            pv_korte_pauze_count[pv["naam"]] += 1
        return True

    # Als anchor bestaat, probeer exact +10, dan groter, anders lagere alternatieven
    if anchor >= 0:
        # +10 en verder
        for d in range(10, len(pauze_cols)-anchor):
            if _try_place_at_col(pauze_cols[anchor + d]):
                return True
        # lager
        for d in range(9, 0, -1):
            idx = anchor + d
            if 0 <= idx < len(pauze_cols) and _try_place_at_col(pauze_cols[idx]):
                return True

    # Als geen anchor of niets gevonden: val terug op fairness, maar zonder links-bias
    # Kies uit alle geldige (pv_row, col) paren, sorteer op pv fairness en dan op kolomindex die het verst ligt van begin (rechts-bias)
    pairs = []
    for col in pauze_cols:
        if not is_korte_pauze_toegestaan_col(col, naam):
            continue
        header = ws_pauze.cell(1, col).value
        uur = parse_header_uur(header)
        if uur not in werk_uren or uur in verboden_uren:
            continue
        attr = vind_attractie_op_uur(naam, uur)
        if not attr:
            continue
        for pv, pv_row in pv_rows:
            if is_pauzevlinder(naam) and pv["naam"] != naam:
                continue
            if ws_pauze.cell(pv_row, col).value not in [None, ""]:
                continue
            if not pv_kan_attr(pv, attr) and not is_student_extra(naam):
                continue
            pairs.append((pv, pv_row, col))
    if not pairs:
        return False
    pairs.sort(key=lambda x: (pv_korte_pauze_count[x[0]["naam"]], -pauze_cols.index(x[2])))
    pv, pv_row, col = pairs[0]
    attr = vind_attractie_op_uur(naam, parse_header_uur(ws_pauze.cell(1, col).value))
    ws_pauze.cell(pv_row-1, col).value = attr
    ws_pauze.cell(pv_row-1, col).alignment = center_align
    ws_pauze.cell(pv_row-1, col).border = thin_border
    cel = ws_pauze.cell(pv_row, col)
    cel.value = naam
    cel.alignment = center_align
    cel.border = thin_border
    cel.fill = lichtpaars_fill
    # Niet meetellen als dit een EXTRA overname is of als de pauze voor een PV zelf is
    if (not is_pauzevlinder(naam)) and (normalize_attr(attr) != 'extra'):
        pv_korte_pauze_count[pv["naam"]] += 1
    return True

# verzamel alle namen met een lange pauze en sorteer op laatste einde (desc)
names_with_long = []
alle_studenten_namen = {s["naam"] for s in studenten if student_totalen.get(s["naam"], 0) >= 4}
for naam in alle_studenten_namen:
    if _has_long_pause(naam):
        end_idx = _last_long_pause_end_index(naam)
        names_with_long.append((end_idx, naam))
names_with_long.sort(reverse=True)  # laatste eerst

for _end, naam in names_with_long:
    _place_short_pause_for(naam)

# Zorg dat latere rondes deze personen overslaan: recompute korte_pauze_ontvangers nu
korte_pauze_ontvangers = set()
for pv, pv_row in pv_rows:
    for idx, col in enumerate(pauze_cols):
        cel = ws_pauze.cell(pv_row, col)
        if cel.value and str(cel.value).strip() != "":
            # korte pauze = enkel blok
            is_lange = False
            if idx+1 < len(pauze_cols):
                next_col = pauze_cols[idx+1]
                cel_next = ws_pauze.cell(pv_row, next_col)
                if cel_next.value == cel.value:
                    is_lange = True
            if idx > 0:
                prev_col = pauze_cols[idx-1]
                prev_cel = ws_pauze.cell(pv_row, prev_col)
                if prev_cel.value == cel.value:
                    is_lange = True
            if not is_lange:
                korte_pauze_ontvangers.add(str(cel.value).strip())

# Bepaal wie geen lange pauze heeft gekregen
studenten_zonder_lange_pauze = []
for s in studenten:
    naam = s["naam"]
    heeft_lange = False
    for pv, pv_row in pv_rows:
        for idx, col in enumerate(pauze_cols):
            cel = ws_pauze.cell(pv_row, col)
            if cel.value == naam:
                # Check of volgende cel ook deze naam heeft (dubbele blok)
                if idx+1 < len(pauze_cols):
                    next_col = pauze_cols[idx+1]
                    cel_next = ws_pauze.cell(pv_row, next_col)
                    if cel_next.value == naam:
                        heeft_lange = True
                        break
        if heeft_lange:
            break
    if not heeft_lange:
        studenten_zonder_lange_pauze.append(s)

# Eerst: korte pauze toewijzen aan studenten zonder lange pauze
def korte_pauze_toewijzen(studenten_lijst):
    for s in studenten_lijst:
        if s["naam"] in korte_pauze_ontvangers or _has_short_pause(s["naam"]):
            continue
        naam = s["naam"]
        werk_uren = get_student_work_hours(naam)
        if len(werk_uren) > 2:
            verboden_uren = {werk_uren[0], werk_uren[-1]}
        else:
            verboden_uren = set(werk_uren)
        pauze_cols_sorted = sorted(pauze_cols)
        geplaatst = False
        for uur in random.sample(werk_uren, len(werk_uren)):
            if uur in verboden_uren:
                continue
            attr = vind_attractie_op_uur(naam, uur)
            if not attr:
                continue
            geldige_slots = []
            for (pv, pv_row) in pv_rows:
                # Pauzevlinders: enkel op eigen rij
                if is_pauzevlinder(naam) and pv["naam"] != naam:
                    continue
                for col in pauze_cols:
                    col_header = ws_pauze.cell(1, col).value
                    col_uur = parse_header_uur(col_header)
                    if col_uur != uur:
                        continue
                    if not is_korte_pauze_toegestaan_col(col, naam):
                        continue
                    if not pv_kan_attr(pv, attr) and not is_student_extra(naam):
                        continue
                    cel = ws_pauze.cell(pv_row, col)
                    if cel.value in [None, ""]:
                        geldige_slots.append((pv, pv_row, col))
            geldige_slots.sort(key=lambda slot: pv_korte_pauze_count[slot[0]["naam"]])
            for (pv, pv_row, col) in geldige_slots:
                boven_cel = ws_pauze.cell(pv_row - 1, col)
                # PV korte pauze: laat boven leeg
                boven_cel.value = None if is_pauzevlinder(naam) else attr
                boven_cel.alignment = center_align
                boven_cel.border = thin_border
                cel = ws_pauze.cell(pv_row, col)
                cel.value = naam
                cel.alignment = center_align
                cel.border = thin_border
                cel.fill = lichtpaars_fill
                korte_pauze_ontvangers.add(naam)
                # Niet meetellen als dit een EXTRA overname is of als de pauze voor een PV zelf is
                if (not is_pauzevlinder(naam)) and (normalize_attr(attr) != 'extra'):
                    pv_korte_pauze_count[pv["naam"]] += 1
                geplaatst = True
                break
            if geplaatst:
                break
        if not geplaatst:
            niet_geplaatste_korte_pauze.append(naam)

korte_pauze_toewijzen(studenten_zonder_lange_pauze)
# Daarna: de rest
korte_pauze_toewijzen([s for s in studenten if s not in studenten_zonder_lange_pauze])
korte_pauze_toewijzen([s for s in studenten if s not in studenten_zonder_lange_pauze])

# --- Iteratief wisselen: studenten zonder korte pauze proberen te ruilen met anderen (geen pauzevlinders) ---

def vind_korte_pauze_cell(naam):
    """Vind (pv_row, col) van de korte pauze van deze student, of None."""
    for pv, pv_row in pv_rows:
        for idx, col in enumerate(pauze_cols):
            cel = ws_pauze.cell(pv_row, col)
            if cel.value == naam:
                # Check of het een korte pauze is (enkel blok, niet dubbel)
                is_lange = False
                if idx+1 < len(pauze_cols):
                    next_col = pauze_cols[idx+1]
                    cel_next = ws_pauze.cell(pv_row, next_col)
                    if cel_next.value == naam:
                        is_lange = True
                if idx > 0:
                    prev_col = pauze_cols[idx-1]
                    prev_cel = ws_pauze.cell(pv_row, prev_col)
                    if prev_cel.value == naam:
                        is_lange = True
                if not is_lange:
                    return (pv_row, col)
    return None

def kan_student_korte_pauze_op_plek(naam, pv_row, col):
    """Check of student naam op deze plek een korte pauze mag hebben."""
    # Mag niet op pauzevlinder-rij
    if is_pauzevlinder(naam):
        return False
    # Moet werken op dit uur
    col_header = ws_pauze.cell(1, col).value
    col_uur = parse_header_uur(col_header)
    werk_uren = get_student_work_hours(naam)
    if col_uur not in werk_uren:
        return False
    # Niet in eerste/laatste werkuur
    if len(werk_uren) > 2:
        if col_uur == werk_uren[0] or col_uur == werk_uren[-1]:
            return False
    # Attractie moet kloppen
    attr = vind_attractie_op_uur(naam, col_uur)
    if not attr:
        return False
    # Pauzevlinder moet deze attractie kunnen
    pv = None
    for pv_obj, row in pv_rows:
        if row == pv_row:
            pv = pv_obj
            break
    if not pv:
        return False
    if not pv_kan_attr(pv, attr) and not is_student_extra(naam):
        return False
    # Kolom moet korte pauze toestaan
    if not is_korte_pauze_toegestaan_col(col, naam):
        return False
    return True

# Verzamel actuele lijst van studenten zonder korte pauze
werkende_studenten = [s for s in studenten if student_totalen.get(s["naam"], 0) >= 4 and not is_pauzevlinder(s["naam"])]
studenten_zonder_korte_pauze = []
for s in werkende_studenten:
    naam = s["naam"]
    heeft_korte = False
    for pv, pv_row in pv_rows:
        for idx, col in enumerate(pauze_cols):
            cel = ws_pauze.cell(pv_row, col)
            if cel.value == naam:
                # Check of GEEN dubbele blok (dus geen lange pauze)
                is_lange = False
                if idx+1 < len(pauze_cols):
                    next_col = pauze_cols[idx+1]
                    cel_next = ws_pauze.cell(pv_row, next_col)
                    if cel_next.value == naam:
                        is_lange = True
                if idx > 0:
                    prev_col = pauze_cols[idx-1]
                    prev_cel = ws_pauze.cell(pv_row, prev_col)
                    if prev_cel.value == naam:
                        is_lange = True
                if not is_lange:
                    heeft_korte = True
                    break
        if heeft_korte:
            break
    if not heeft_korte:
        studenten_zonder_korte_pauze.append(naam)

max_wissel_passes = 10
for _ in range(max_wissel_passes):
    if not studenten_zonder_korte_pauze:
        break
    verbeterd = False
    for naam_zonder in studenten_zonder_korte_pauze:
        # Probeer te ruilen met een student die wél een korte pauze heeft (geen pauzevlinder)
        for s in werkende_studenten:
            naam_met = s["naam"]
            if naam_met == naam_zonder:
                continue
            if naam_met in studenten_zonder_korte_pauze:
                continue
            # Vind de korte pauze van deze student
            plek = vind_korte_pauze_cell(naam_met)
            if not plek:
                continue
            pv_row, col = plek
            # Mag naam_zonder op deze plek een korte pauze hebben?
            if not kan_student_korte_pauze_op_plek(naam_zonder, pv_row, col):
                continue
            # Bepaal attractie voor naam_zonder op deze plek
            col_header = ws_pauze.cell(1, col).value
            col_uur = parse_header_uur(col_header)
            attr_zonder = vind_attractie_op_uur(naam_zonder, col_uur)
            if not attr_zonder:
                continue
            # Mag naam_met elders een korte pauze krijgen?
            # Zoek alternatieve plek voor naam_met
            gevonden = False
            for pv2, pv_row2 in pv_rows:
                if is_pauzevlinder(naam_met):
                    continue
                for col2 in pauze_cols:
                    if (pv_row2, col2) == (pv_row, col):
                        continue
                    cel2 = ws_pauze.cell(pv_row2, col2)
                    if cel2.value not in [None, ""]:
                        continue
                    if not kan_student_korte_pauze_op_plek(naam_met, pv_row2, col2):
                        continue
                    # Bepaal attractie voor naam_met op nieuwe plek
                    col2_header = ws_pauze.cell(1, col2).value
                    col2_uur = parse_header_uur(col2_header)
                    attr_met = vind_attractie_op_uur(naam_met, col2_uur)
                    if not attr_met:
                        continue
                    # Wissel uitvoeren
                    # 1. naam_met uit oude plek halen
                    ws_pauze.cell(pv_row, col).value = None
                    ws_pauze.cell(pv_row, col).fill = naam_leeg_fill
                    ws_pauze.cell(pv_row-1, col).value = None
                    # 2. naam_zonder op deze plek zetten
                    ws_pauze.cell(pv_row, col).value = naam_zonder
                    ws_pauze.cell(pv_row, col).fill = lichtpaars_fill
                    ws_pauze.cell(pv_row, col).alignment = center_align
                    ws_pauze.cell(pv_row, col).border = thin_border
                    ws_pauze.cell(pv_row-1, col).value = attr_zonder
                    ws_pauze.cell(pv_row-1, col).alignment = center_align
                    ws_pauze.cell(pv_row-1, col).border = thin_border
                    # 3. naam_met op nieuwe plek zetten
                    ws_pauze.cell(pv_row2, col2).value = naam_met
                    ws_pauze.cell(pv_row2, col2).fill = lichtpaars_fill
                    ws_pauze.cell(pv_row2, col2).alignment = center_align
                    ws_pauze.cell(pv_row2, col2).border = thin_border
                    ws_pauze.cell(pv_row2-1, col2).value = attr_met
                    ws_pauze.cell(pv_row2-1, col2).alignment = center_align
                    ws_pauze.cell(pv_row2-1, col2).border = thin_border
                    verbeterd = True
                    gevonden = True
                    break
                if gevonden:
                    break
            if verbeterd:
                break
        if verbeterd:
            break
    # Update lijst van studenten zonder korte pauze
    studenten_zonder_korte_pauze = []
    for s in werkende_studenten:
        naam = s["naam"]
        heeft_korte = False
        for pv, pv_row in pv_rows:
            for idx, col in enumerate(pauze_cols):
                cel = ws_pauze.cell(pv_row, col)
                if cel.value == naam:
                    # Check of GEEN dubbele blok (dus geen lange pauze)
                    is_lange = False
                    if idx+1 < len(pauze_cols):
                        next_col = pauze_cols[idx+1]
                        cel_next = ws_pauze.cell(pv_row, next_col)
                        if cel_next.value == naam:
                            is_lange = True
                    if idx > 0:
                        prev_col = pauze_cols[idx-1]
                        prev_cel = ws_pauze.cell(pv_row, prev_col)
                        if prev_cel.value == naam:
                            is_lange = True
                    if not is_lange:
                        heeft_korte = True
                        break
            if heeft_korte:
                break
        if not heeft_korte:
            studenten_zonder_korte_pauze.append(naam)
    if not verbeterd:
        break  # geen verbetering meer mogelijk

# Iteratieve optimalisatie: verschuif korte pauzes van "rijke" naar "arme" pauzevlinders
max_opt_passes = 10
for _ in range(max_opt_passes):
    # Zoek max en min aantal korte pauzes
    if not pv_korte_pauze_count:
        break
    max_pv = max(pv_korte_pauze_count, key=lambda k: pv_korte_pauze_count[k])
    min_pv = min(pv_korte_pauze_count, key=lambda k: pv_korte_pauze_count[k])
    if pv_korte_pauze_count[max_pv] - pv_korte_pauze_count[min_pv] <= 1:
        break  # verdeling is al redelijk
    # Zoek een korte pauze van max_pv die overgezet kan worden naar min_pv
    found = False
    for col in pauze_cols:
        pv_row_max = next((row for pv, row in pv_rows if pv["naam"] == max_pv), None)
        pv_row_min = next((row for pv, row in pv_rows if pv["naam"] == min_pv), None)
        if pv_row_max is None or pv_row_min is None:
            continue
        cel_max = ws_pauze.cell(pv_row_max, col)
        naam = cel_max.value
        if not naam or str(naam).strip() == "":
            continue
        # Check of het een korte pauze is (enkel blok, niet dubbel)
        idx = pauze_cols.index(col)
        is_lange = False
        if idx+1 < len(pauze_cols):
            next_col = pauze_cols[idx+1]
            cel_next = ws_pauze.cell(pv_row_max, next_col)
            if cel_next.value == cel_max.value:
                is_lange = True
        if idx > 0:
            prev_col = pauze_cols[idx-1]
            prev_cel = ws_pauze.cell(pv_row_max, prev_col)
            if prev_cel.value == cel_max.value:
                is_lange = True
        if is_lange:
            continue
        # Mag min_pv deze attractie overnemen?
        attr = ws_pauze.cell(pv_row_max-1, col).value
        if not pv_kan_attr(next(pv for pv, _ in pv_rows if pv["naam"] == min_pv), attr):
            continue
        # Is de cel bij min_pv vrij?
        cel_min = ws_pauze.cell(pv_row_min, col)
        if cel_min.value not in [None, ""]:
            continue
        # Wissel de korte pauze van max_pv naar min_pv
        cel_min.value = naam
        cel_min.alignment = center_align
        cel_min.border = thin_border
        cel_min.fill = lichtpaars_fill
        ws_pauze.cell(pv_row_min-1, col).value = attr
        ws_pauze.cell(pv_row_min-1, col).alignment = center_align
        ws_pauze.cell(pv_row_min-1, col).border = thin_border
        cel_max.value = None
        ws_pauze.cell(pv_row_max-1, col).value = None
        # Pas telling aan enkel als dit geen EXTRA-overname is
        if attr and normalize_attr(attr) != 'extra':
            pv_korte_pauze_count[max_pv] -= 1
            pv_korte_pauze_count[min_pv] += 1
        found = True
        break
    if not found:
        break  # geen verschuiving meer mogelijk



# --- Iteratieve optimalisatie: verdeel lange pauzes zo eerlijk mogelijk over pauzevlinders ---

max_opt_passes_lange = 10
from collections import Counter
for _ in range(max_opt_passes_lange):
    pass  # (oude optimalisatie-code is verwijderd, want niet meer nodig)

# --- Pauzevlinders met >6u: altijd lange pauze in eigen rij ---
import random
# --- Pauzevlinders met >6u: altijd lange pauze in eigen rij, gespreid over eerste drie pauzeuren ---
for pv, pv_row in pv_rows:
    naam = pv["naam"]
    werk_uren = get_student_work_hours(naam)
    if len(werk_uren) > 6:
        # Alleen de eerste 11 kwartieren (indices 0 t/m 10) zijn toegestaan voor lange pauzes
        if heeft_al_lange_pauze(naam):
            continue
        halve_uren = []  # lijst van (idx, col1, col2)
        max_start_idx = min(8, len(pauze_cols)-2)  # idx 0 t/m 10 zijn halve uren binnen eerste 11 kwartieren
        for idx in range(max_start_idx+1):
            col1 = pauze_cols[idx]
            col2 = pauze_cols[idx+1]
            col1_header = ws_pauze.cell(1, col1).value
            # Alleen starten op heel of half uur
            try:
                min1 = int(str(col1_header).split('u')[1]) if 'u' in str(col1_header) and len(str(col1_header).split('u')) > 1 else 0
            except:
                min1 = 0
            if min1 not in (0, 30):
                continue
            cel1 = ws_pauze.cell(pv_row, col1)
            cel2 = ws_pauze.cell(pv_row, col2)
            if cel1.value in [None, ""] and cel2.value in [None, ""]:
                halve_uren.append((idx, col1, col2))
        # Shuffle de halve uren
        random.shuffle(halve_uren)
        # Probeer in geshuffelde volgorde een lange pauze te plaatsen
        geplaatst = False
        for idx, col1, col2 in halve_uren:
            cel1 = ws_pauze.cell(pv_row, col1)
            cel2 = ws_pauze.cell(pv_row, col2)
            if cel1.value in [None, ""] and cel2.value in [None, ""] and not heeft_al_lange_pauze(naam):
                cel1.value = naam
                cel2.value = naam
                cel1.alignment = center_align
                cel2.alignment = center_align
                cel1.border = thin_border
                cel2.border = thin_border
                cel1.fill = lichtgroen_fill
                cel2.fill = lichtgroen_fill
                geplaatst = True
                break
        # Indien geen plek gevonden, doe niets (komt zelden voor)



output = BytesIO()

# -----------------------------
# ENFORCE: Korte pauzes minimaal 10 blokken uit elkaar (maximaliseer anders)
# -----------------------------

# We beschouwen alle pauzecellen (korte én beide helften van lange pauzes).
# We verplaatsen alleen korte pauzes (enkelblok) naar lege geschikte slots.

def _get_student_pause_cols(naam):
    cols = []
    for _pv, pv_row in pv_rows:
        for col in pauze_cols:
            if ws_pauze.cell(pv_row, col).value == naam:
                cols.append(col)
    return sorted(cols)

def _get_student_short_pause_positions(naam):
    pos = []  # lijst van (pv_row, col)
    for _pv, pv_row in pv_rows:
        for idx, col in enumerate(pauze_cols):
            cel = ws_pauze.cell(pv_row, col)
            if cel.value != naam:
                continue
            # controleer of GEEN deel van een dubbele blok (lange pauze)
            is_dubbel = False
            if idx+1 < len(pauze_cols):
                if ws_pauze.cell(pv_row, pauze_cols[idx+1]).value == naam:
                    is_dubbel = True
            if idx > 0:
                if ws_pauze.cell(pv_row, pauze_cols[idx-1]).value == naam:
                    is_dubbel = True
            if not is_dubbel:
                pos.append((pv_row, col))
    return pos

def _min_gap(cols):
    if len(cols) < 2:
        return 10**9
    cols = sorted(cols)
    mg = min(cols[i+1]-cols[i] for i in range(len(cols)-1))
    return mg

def _can_place_short_pause(naam, pv_row, col):
    # cel moet leeg zijn
    if ws_pauze.cell(pv_row, col).value not in [None, ""]:
        return False
    # kolom moet korte pauze toelaten
    if not is_korte_pauze_toegestaan_col(col, naam):
        return False
    # student moet werken op dit uur en niet in eerste/laatste werkuur
    header = ws_pauze.cell(1, col).value
    uur = parse_header_uur(header)
    if uur is None:
        return False
    werk_uren = get_student_work_hours(naam)
    if uur not in werk_uren:
        return False
    if len(werk_uren) > 2:
        if uur == werk_uren[0] or uur == werk_uren[-1]:
            return False
    else:
        # bij 1-2 uren: geen korte pauze plannen
        return False
    # pauzevlinder-capability: pauzevlinder op pv_row moet attractie kunnen
    attr = vind_attractie_op_uur(naam, uur)
    if not attr:
        return False
    pv_obj = None
    for pv, row in pv_rows:
        if row == pv_row:
            pv_obj = pv
            break
    if pv_obj is None:
        return False
    if not pv_kan_attr(pv_obj, attr) and not is_student_extra(naam):
        return False
    return True

def _move_short_pause(naam, from_row, from_col, to_row, to_col):
    # leegmaken bron
    ws_pauze.cell(from_row, from_col).value = None
    ws_pauze.cell(from_row-1, from_col).value = None
    # invullen doel
    header = ws_pauze.cell(1, to_col).value
    uur = parse_header_uur(header)
    attr = vind_attractie_op_uur(naam, uur)
    # Voor PV korte pauze: laat boven leeg
    ws_pauze.cell(to_row-1, to_col).value = None if is_pauzevlinder(naam) else attr
    ws_pauze.cell(to_row-1, to_col).alignment = center_align
    ws_pauze.cell(to_row-1, to_col).border = thin_border
    ws_pauze.cell(to_row, to_col).value = naam
    ws_pauze.cell(to_row, to_col).alignment = center_align
    ws_pauze.cell(to_row, to_col).border = thin_border

def _recolor_pauze_sheet():
    # Kleur korte pauze paars, lange (dubbel) groen, leeg lichtblauw
    for _pv, pv_row in pv_rows:
        for idx, col in enumerate(pauze_cols):
            cel = ws_pauze.cell(pv_row, col)
            val = cel.value
            if val in [None, ""]:
                cel.fill = naam_leeg_fill
                continue
            # lange pauze als naastliggende cel dezelfde naam heeft
            is_dubbel = False
            if idx+1 < len(pauze_cols):
                c2 = ws_pauze.cell(pv_row, pauze_cols[idx+1])
                if c2.value == val:
                    cel.fill = lichtgroen_fill
                    c2.fill = lichtgroen_fill
                    is_dubbel = True
                    continue
            if idx > 0:
                c1 = ws_pauze.cell(pv_row, pauze_cols[idx-1])
                if c1.value == val:
                    # onderdeel van reeds gekleurde lange pauze
                    continue
            # anders korte pauze
            cel.fill = lichtpaars_fill

def _enforce_min_gap_for_short_pauses(desired_gap=10, max_passes=5):
    changed = False
    for _ in range(max_passes):
        improved = False
        # itereren over alle studenten met minstens 1 korte pauze
        alle_namen = {s["naam"] for s in studenten if student_totalen.get(s["naam"], 0) > 0}
        for naam in alle_namen:
            short_pos = _get_student_short_pause_positions(naam)
            if not short_pos:
                continue
            all_cols = _get_student_pause_cols(naam)
            # bekijk elke korte pauze afzonderlijk
            for (from_row, from_col) in short_pos:
                # huidige minimale gap voor deze korte pauze
                other_cols = [c for c in all_cols if c != from_col]
                cur_gap = min((abs(from_col - c) for c in other_cols), default=10**9)
                if cur_gap >= desired_gap:
                    continue  # al goed
                # zoek beste lege slot
                best = None  # (best_gap, to_row, to_col)
                for _pv, pv_row in pv_rows:
                    # Pauzevlinders: korte pauze enkel op eigen rij verplaatsen/plaatsen
                    if is_pauzevlinder(naam) and _pv["naam"] != naam:
                        continue
                    for col in pauze_cols:
                        if not _can_place_short_pause(naam, pv_row, col):
                            continue
                        # gap als we hierheen verplaatsen
                        new_gap = min((abs(col - c) for c in other_cols), default=10**9)
                        if (best is None) or (new_gap > best[0]) or (new_gap == best[0] and col < best[2]):
                            best = (new_gap, pv_row, col)
                            # early exit als we desired halen
                            if new_gap >= desired_gap:
                                break
                    if best and best[0] >= desired_gap:
                        break
                if best and best[0] > cur_gap:
                    _move_short_pause(naam, from_row, from_col, best[1], best[2])
                    improved = True
                    changed = True
                    # update caches voor volgende iteraties
                    all_cols = _get_student_pause_cols(naam)
        if not improved:
            break
    # herkleuren na eventuele wijzigingen
    if changed:
        _recolor_pauze_sheet()
    return changed

# Voer de enforce stap uit met doelafstand 10 blokken
_enforce_min_gap_for_short_pauses(desired_gap=10, max_passes=6)

# Optionele samenvatting in Streamlit
# Debug samenvatting (globale minimale pauze-afstand) verwijderd om UI schoon te houden.
# Indien opnieuw nodig: functie _global_min_gap_summary() herstellen.

# --- FEEDBACK SHEET ---
ws_feedback = wb_out.create_sheet("Feedback")
row_fb = 1

# 1. Lange werkers (>6u) zonder lange pauze
lange_werkers_zonder_lange_pauze = set()

def _heeft_lange_pauze_naam(naam: str) -> bool:
    """Zoek in ws_pauze of deze persoon een dubbele blok (lange pauze) heeft."""
    for pv, pv_row in pv_rows:
        for idx, col in enumerate(pauze_cols):
            cel = ws_pauze.cell(pv_row, col)
            if cel.value == naam:
                # Check of volgende cel ook deze naam heeft (dubbele blok)
                if idx+1 < len(pauze_cols):
                    next_col = pauze_cols[idx+1]
                    cel_next = ws_pauze.cell(pv_row, next_col)
                    if cel_next.value == naam:
                        return True
    return False

# a) reguliere lange werkers
for s in lange_werkers:
    naam = s["naam"]
    if not _heeft_lange_pauze_naam(naam):
        lange_werkers_zonder_lange_pauze.add(naam)

# b) pauzevlinders die >6u werken meenemen
for pv, _pv_row in pv_rows:
    naam = pv["naam"]
    werk_uren = get_student_work_hours(naam) or []
    if len(werk_uren) > 6:
        if not _heeft_lange_pauze_naam(naam):
            lange_werkers_zonder_lange_pauze.add(naam)

ws_feedback.cell(row_fb, 1, "Lange werkers (>6u) zonder lange pauze:")
row_fb += 1
if lange_werkers_zonder_lange_pauze:
    for naam in sorted(lange_werkers_zonder_lange_pauze):
        ws_feedback.cell(row_fb, 1, naam)
        row_fb += 1
else:
    vinkje_cel = ws_feedback.cell(row_fb, 1, "✓")
    ws_feedback.cell(row_fb, 2, "Iedereen heeft een lange pauze gekregen.")
    from openpyxl.styles import PatternFill, Font
    vinkje_cel.fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")  # opvallend groen
    vinkje_cel.font = Font(bold=True, color="006100")  # donkergroen
    row_fb += 1

# 2. Werkende studenten zonder korte pauze
werkende_studenten = [s for s in studenten if student_totalen.get(s["naam"], 0) >= 4]
studenten_zonder_korte_pauze = []
for s in werkende_studenten:
    naam = s["naam"]
    # Zoek in ws_pauze of deze student een korte pauze heeft (enkel blok, niet dubbel)
    heeft_korte = False
    for pv, pv_row in pv_rows:
        for idx, col in enumerate(pauze_cols):
            cel = ws_pauze.cell(pv_row, col)
            if cel.value == naam:
                # Check of GEEN dubbele blok (dus geen lange pauze)
                is_lange = False
                if idx+1 < len(pauze_cols):
                    next_col = pauze_cols[idx+1]
                    cel_next = ws_pauze.cell(pv_row, next_col)
                    if cel_next.value == naam:
                        is_lange = True
                if idx > 0:
                    prev_col = pauze_cols[idx-1]
                    prev_cel = ws_pauze.cell(pv_row, prev_col)
                    if prev_cel.value == naam:
                        is_lange = True
                if not is_lange:
                    heeft_korte = True
                    break
        if heeft_korte:
            break
    if not heeft_korte:
        studenten_zonder_korte_pauze.append(naam)

ws_feedback.cell(row_fb, 1, "Werkende studenten zonder korte pauze:")
row_fb += 1
if studenten_zonder_korte_pauze:
    for naam in studenten_zonder_korte_pauze:
        ws_feedback.cell(row_fb, 1, naam)
        row_fb += 1
else:
    vinkje_cel = ws_feedback.cell(row_fb, 1, "✓")
    ws_feedback.cell(row_fb, 2, "Iedereen heeft een korte pauze gekregen.")
    vinkje_cel.fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
    vinkje_cel.font = Font(bold=True, color="006100")
    row_fb += 1


##### EXTRA INFO TOEVOEGEN AAN PAUZEPLANNING (A12 e.v.)
##### -------------------------------------------------------------
ws_pauze_sheet = wb_out["Pauzevlinders"]
witte_fill = PatternFill(start_color="FFFFFF", fill_type="solid")

# --- VINKJE V3 + TEKST W3 in Aanpassingen ---
vinkje_aanpassingen = ws_aanpassingen.cell(row=3, column=22).value  # V3

if vinkje_aanpassingen in [1, True, "WAAR", "X"]:
    waarde = ws_aanpassingen.cell(row=3, column=23).value  # W3 (samengevoegde cel)
    if waarde:
        regels = str(waarde).split("\n")
        for i, regel in enumerate(regels):
            regel = regel.strip()
            if regel:
                cel = ws_pauze_sheet.cell(row=14 + i, column=1, value=regel)
                cel.fill = witte_fill
                cel.border = thin_border
                cel.alignment = Alignment(horizontal="left", vertical="center")
# -------------------------------------



wb_out.save(output)
output.seek(0)  # Zorg dat lezen vanaf begin kan


#NIEUWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWW
#NIEUWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWW

# DEEL 5: PP optie 2 + Feedback optie 2
# ──────────────────────────────────────────────────────────────────────
def maak_pp2_sheets(wb_arg, am_arg):
    global ws_planning, student_totalen

    # Globals tijdelijk omwisselen
    _ws_planning_bak     = ws_planning
    _student_totalen_bak = student_totalen

    ws_planning = wb_arg["Planning"]

    # Bouw student_totalen vanuit de Planning-sheet van wb_arg,
    # net zoals de globale versie dat doet vanuit ws_out.
    # Dit is belangrijk voor pauzevlinders: hun pauzevlinderuren
    # staan NIET in assigned_map, maar wél in de Planning-sheet.
    student_totalen = defaultdict(int)
    _ws_plan_tmp = wb_arg["Planning"]
    for row in _ws_plan_tmp.iter_rows(min_row=2, values_only=True):
        for naam in row[1:]:
            if naam and str(naam).strip() != "":
                student_totalen[naam] += 1

    for sheet_name in ["Pauzeplanning", "Feedback PP"]:
        if sheet_name in wb_arg.sheetnames:
            wb_arg.remove(wb_arg[sheet_name])

    ws_pauze_basis = wb_arg["Pauzevlinders"]
    ws_pp2 = wb_arg.copy_worksheet(ws_pauze_basis)
    ws_pp2.title = "Pauzeplanning"

    # ── hierna de rest van DEEL 5 geïndenteerd ──

    # -----------------------------
    # Helpers
    # -----------------------------
    
    def pp2_is_minderjarig(naam):
        return "-18" in str(naam)
    
    
    def pp2_parse_kwartier_header(header):
        """
        Zet '12u', '12u15', '12u30', '12u45' om naar minuten sinds 00:00.
        """
        if not header:
            return None
        s = str(header).strip().lower()
        if "u" not in s:
            return None
        parts = s.split("u", 1)
        try:
            uur = int(parts[0])
            mins = int(parts[1]) if parts[1] != "" else 0
            return uur * 60 + mins
        except:
            return None
    
    def pp2_get_pauze_cols(ws_sheet):
        cols = []
        for col in range(2, ws_sheet.max_column + 1):
            header = ws_sheet.cell(1, col).value
            if header and "u" in str(header):
                cols.append(col)
        return cols
    
    def pp2_get_pv_rows(ws_sheet, selected):
        """
        Geeft lijst van tuples: (pv_dict, naam_rij)
        waarbij naam_rij de rij is waar de naam van de pauzevlinder staat.
        """
        rows = []
        for pv in selected:
            found = None
            for r in range(2, ws_sheet.max_row + 1):
                val = ws_sheet.cell(r, 1).value
                if val and str(val).strip() == str(pv["naam"]).strip():
                    found = r
                    break
            if found is not None:
                rows.append((pv, found))
        return rows
    
    def pp2_get_student_work_hours(naam):
        """
        Leest echte werkuren uit het werkblad Planning.
        """
        uren = set()
        for col in range(2, ws_planning.max_column + 1):
            header = ws_planning.cell(1, col).value
            uur = parse_header_uur(header)
            if uur is None:
                continue
            for row in range(2, ws_planning.max_row + 1):
                if ws_planning.cell(row, col).value == naam:
                    uren.add(uur)
                    break
        return sorted(uren)
    
    def pp2_is_first_or_last_work_hour(naam, kwartier_col, ws_sheet):
        """
        Checkt of dit kwartier in het eerste of laatste werkuur valt.
        """
        werk_uren = pp2_get_student_work_hours(naam)
        if not werk_uren:
            return True
    
        header = ws_sheet.cell(1, kwartier_col).value
        pauze_uur = parse_header_uur(header)
        if pauze_uur is None:
            return True
    
        return pauze_uur == werk_uren[0] or pauze_uur == werk_uren[-1]
    
    def pp2_candidate_cols_for_student(naam, ws_sheet, pauze_cols):
        """
        Alle geldige kwartierkolommen voor korte pauze:
        - student werkt dat uur
        - niet in eerste of laatste werkuur
        """
        werk_uren = pp2_get_student_work_hours(naam)
        if len(werk_uren) < 4:
            return []
    
        first_hour = werk_uren[0]
        last_hour = werk_uren[-1]
    
        candidates = []
        for col in pauze_cols:
            header = ws_sheet.cell(1, col).value
            uur = parse_header_uur(header)
            if uur is None:
                continue
            if uur in werk_uren and uur != first_hour and uur != last_hour:
                candidates.append(col)
    
        return candidates
    
    def pp2_choose_middle_col(naam, ws_sheet, pauze_cols):
        """
        Kies een kwartier zo goed mogelijk in het midden van de shift,
        rekening houdend met de toegelaten kwartieren.
        """
        candidates = pp2_candidate_cols_for_student(naam, ws_sheet, pauze_cols)
        if not candidates:
            return None
    
        werk_uren = pp2_get_student_work_hours(naam)
        shift_start = min(werk_uren) * 60
        shift_end = (max(werk_uren) + 1) * 60
        midpoint = (shift_start + shift_end) / 2
    
        best_col = None
        best_score = None
    
        for col in candidates:
            mins = pp2_parse_kwartier_header(ws_sheet.cell(1, col).value)
            if mins is None:
                continue
            score = abs(mins - midpoint)
            if best_score is None or score < best_score:
                best_score = score
                best_col = col
    
        return best_col
    
    def pp2_is_valid_short_break_for_student(naam, col, ws_sheet):
        """
        Een korte pauze mag alleen als:
        - student werkt in dat kwartier
        - niet in eerste of laatste werkuur
        - student op dat kwartier nog nergens anders in het pauzerooster staat
        """
        header = ws_sheet.cell(1, col).value
        uur = parse_header_uur(header)
    
        if uur is None:
            return False
    
        werk_uren = pp2_get_student_work_hours(naam)
        if not werk_uren:
            return False
    
        if uur not in werk_uren:
            return False
    
        eerste_uur = werk_uren[0]
        laatste_uur = werk_uren[-1]
    
        if uur == eerste_uur or uur == laatste_uur:
            return False
    
        if pp2_student_heeft_al_pauze_op_kolom(
            naam=naam,
            col=col,
            ws_sheet=ws_sheet,
            pv_rows=pv_rows_pp2
        ):
            return False
    
        return True
    
    def pp2_choose_middle_double_col_for_minor(naam, ws_sheet, pauze_cols):
        """
        Zoek startkolom voor 2 opeenvolgende kwartieren voor minderjarigen:
        - student werkt op beide kwartieren
        - student stopt om of voor 16u (dus laatste werkblok <= 15)
        - student werkt >4u en <=6u
        - start enkel op een half uur (:00 of :30)
        - zo vroeg mogelijk in de shift
        - beide cellen moeten geldig zijn volgens de gewone korte-pauze-regels
        """
        werk_uren = pp2_get_student_work_hours(naam)
        if not werk_uren:
            return None
    
        if len(werk_uren) <= 4 or len(werk_uren) > 6:
            return None
    
        if max(werk_uren) > 15:
            return None
    
        for idx in range(len(pauze_cols) - 1):
            col1 = pauze_cols[idx]
            col2 = pauze_cols[idx + 1]
    
            # moeten opeenvolgende kwartieren zijn
            if col2 != col1 + 1:
                continue
    
            header1 = ws_sheet.cell(1, col1).value
            uur1 = parse_header_uur(header1)
            if uur1 is None:
                continue
    
            # start enkel op heel uur of half uur
            header_text = str(header1).strip().lower()
            if not (header_text.endswith("u") or header_text.endswith("u30")):
                continue
    
            # beide kwartieren moeten geldig zijn volgens gewone korte-pauze-regels
            if not pp2_is_valid_short_break_for_student(naam, col1, ws_sheet):
                continue
            if not pp2_is_valid_short_break_for_student(naam, col2, ws_sheet):
                continue
    
            # beide kwartieren moeten effectief tijdens werkuren vallen
            uur2 = parse_header_uur(ws_sheet.cell(1, col2).value)
            if uur2 is None:
                continue
    
            if uur1 not in werk_uren or uur2 not in werk_uren:
                continue
    
            # eerste geldige optie meteen nemen
            return col1
    
        return None
    
    
    
    def pp2_same_halfhour(col_a, col_b, ws_sheet):
        mins_a = pp2_parse_kwartier_header(ws_sheet.cell(1, col_a).value)
        mins_b = pp2_parse_kwartier_header(ws_sheet.cell(1, col_b).value)
        if mins_a is None or mins_b is None:
            return False
        return (mins_a // 30) == (mins_b // 30)
    
    def pp2_choose_adjacent_same_halfhour(base_col, student_name, ws_sheet, pauze_cols, pv_name_row):
        """
        Tweede student van het duo moet naast de eerste zitten
        in hetzelfde halfuur, indien dat volgens de regels kan.
        """
        if base_col not in pauze_cols:
            return None
    
        idx = pauze_cols.index(base_col)
        opties = []
    
        if idx - 1 >= 0:
            opties.append(pauze_cols[idx - 1])
        if idx + 1 < len(pauze_cols):
            opties.append(pauze_cols[idx + 1])
    
        # Eerst alleen opties in hetzelfde halfuur
        opties = [c for c in opties if pp2_same_halfhour(base_col, c, ws_sheet)]
    
        for col in opties:
            # vak moet leeg zijn
            if ws_sheet.cell(pv_name_row, col).value not in [None, ""]:
                continue
            # niet in eerste/laatste werkuur van deze student
            if pp2_is_first_or_last_work_hour(student_name, col, ws_sheet):
                continue
            # student moet effectief dat uur werken
            uur = parse_header_uur(ws_sheet.cell(1, col).value)
            werk_uren = pp2_get_student_work_hours(student_name)
            if uur not in werk_uren:
                continue
            return col
    
        return None
    
    def pp2_write_name(ws_sheet, row_name, col, naam):
        """
        Schrijf in PP optie 2:
        - bovenste vak: attractie waarop student dat moment staat
        - onderste vak: naam van student
        - korte pauze = paars
        - lange pauze = groen (voor later bruikbaar)
        """
        lichtgroen_fill = PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid")
        lichtpaars_fill = PatternFill(start_color="E6DAF7", end_color="E6DAF7", fill_type="solid")
    
        # bepaal uur van deze kolom
        header = ws_sheet.cell(1, col).value
        uur = parse_header_uur(header)
    
        # attractie erboven invullen
        info_cel = ws_sheet.cell(row_name - 1, col)
        attr = vind_attractie_op_uur(naam, uur) if uur is not None else None
        info_cel.value = attr if attr else ""
        info_cel.alignment = center_align
        info_cel.border = thin_border
    
        # naam invullen
        cel = ws_sheet.cell(row_name, col)
        cel.value = naam
        cel.alignment = center_align
        cel.border = thin_border
    
        # check of dit een lange of korte pauze is
        is_lange_pauze = False
        if col - 1 >= 2 and ws_sheet.cell(row_name, col - 1).value == naam:
            is_lange_pauze = True
        if col + 1 <= ws_sheet.max_column and ws_sheet.cell(row_name, col + 1).value == naam:
            is_lange_pauze = True
    
        cel.fill = lichtgroen_fill if is_lange_pauze else lichtpaars_fill
    
    def pp2_clear_pauze_grid(ws_sheet, pv_rows, pauze_cols):
        """
        Wis enkel de effectieve pauzevakken:
        - rij erboven: attractie/info
        - naamrij: naam
        Kolom A en extra info lager op het blad blijven behouden.
        """
        leeg_fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
    
        for pv, naam_rij in pv_rows:
            info_rij = naam_rij - 1
            for col in pauze_cols:
                # bovenste rij leegmaken
                ws_sheet.cell(info_rij, col).value = None
                ws_sheet.cell(info_rij, col).alignment = center_align
                ws_sheet.cell(info_rij, col).border = thin_border
    
                # naamrij leegmaken
                ws_sheet.cell(naam_rij, col).value = None
                ws_sheet.cell(naam_rij, col).alignment = center_align
                ws_sheet.cell(naam_rij, col).border = thin_border
                ws_sheet.cell(naam_rij, col).fill = leeg_fill
    
    
    def pp2_student_heeft_al_pauze_op_kolom(naam, col, ws_sheet, pv_rows):
        """
        True als deze student op deze kwartierkolom al ergens in het pauzerooster staat,
        ongeacht op welke pauzevlinder-rij.
        """
        for _pv, pv_row in pv_rows:
            if ws_sheet.cell(pv_row, col).value == naam:
                return True
        return False
    
    
    def pp2_student_heeft_al_lange_pauze_op_blok(naam, col1, col2, ws_sheet, pv_rows):
        """
        True als deze student deze 2 kwartieren al ergens als lange pauze heeft staan.
        """
        for _pv, pv_row in pv_rows:
            if (
                ws_sheet.cell(pv_row, col1).value == naam and
                ws_sheet.cell(pv_row, col2).value == naam
            ):
                return True
        return False
    
    
    # -----------------------------
    # Vind de pauzevlinder-rijen in PP optie 2
    # -----------------------------
    pauze_cols_pp2 = pp2_get_pauze_cols(ws_pp2)
    pv_rows_pp2 = pp2_get_pv_rows(ws_pp2, selected)
    
    # Maak de grid leeg, maar behoud layout
    pp2_clear_pauze_grid(ws_pp2, pv_rows_pp2, pauze_cols_pp2)
    
    # -----------------------------
    # Closed spots PP optie 2
    # Afgekapte uren van laatste PV: worden hier gemarkeerd
    # zodat géén enkele pauzelogica er iets in plaatst
    # -----------------------------
    afgeknipt_fill_pp2 = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    pp2_closed_spots = set()  # set van (naam_rij, col)
    
    if afgekapte_pv_uren and selected:
        laatste_pv = selected[-1]
        for pv, naam_rij in pv_rows_pp2:
            if pv["naam"] == laatste_pv["naam"]:
                for col in pauze_cols_pp2:
                    col_uur = parse_header_uur(ws_pp2.cell(1, col).value)
                    if col_uur in afgekapte_pv_uren:
                        pp2_closed_spots.add((naam_rij, col))
                        ws_pp2.cell(naam_rij, col).value = "X"
                        ws_pp2.cell(naam_rij, col).fill = afgeknipt_fill_pp2
                        ws_pp2.cell(naam_rij, col).alignment = center_align
                        ws_pp2.cell(naam_rij, col).border = thin_border
                        ws_pp2.cell(naam_rij - 1, col).value = None
                        ws_pp2.cell(naam_rij - 1, col).fill = afgeknipt_fill_pp2
                        ws_pp2.cell(naam_rij - 1, col).alignment = center_align
                        ws_pp2.cell(naam_rij - 1, col).border = thin_border
                break
    
    def pp2_is_beschikbaar(ws_sheet, rij, col):
        """Cel is beschikbaar als ze leeg is EN geen closed spot."""
        if (rij, col) in pp2_closed_spots:
            return False
        return ws_sheet.cell(rij, col).value in [None, ""]
    
    
    
    # -----------------------------
    # STAP 1:
    # Vroege stoppers (minstens 4u gewerkt en laatste werkblok <= 15)
    # - minderjarige vroege stoppers: eerst halfuur zo vroeg mogelijk,
    #   dan kwartier zo laat mogelijk, zo mogelijk op dezelfde PV-rij
    # - gewone vroege stoppers: duo-logica zoals voorheen
    # Excl. pauzevlinders zelf
    # -----------------------------
    pauzevlinder_namen_set = {pv["naam"] for pv in selected}
    
    vroege_stoppers_gewoon = []
    vroege_stoppers_minderjarig = []
    
    for s in studenten:
        naam = s["naam"]
    
        if naam in pauzevlinder_namen_set:
            continue
    
        werk_uren = pp2_get_student_work_hours(naam)
        if len(werk_uren) < 4:
            continue
    
        laatste_werkblok = max(werk_uren)
        startuur = min(werk_uren)
        aantal_uren = len(werk_uren)
    
        if laatste_werkblok > 15:
            continue
    
        item = {
            "naam": naam,
            "werk_uren": werk_uren,
            "einduur": laatste_werkblok,
            "startuur": startuur,
            "aantal_uren": aantal_uren
        }
    
        if pp2_is_minderjarig(naam):
            vroege_stoppers_minderjarig.append(item)
        else:
            vroege_stoppers_gewoon.append(item)
    
    # Sorteervolgorde: vroegst stoppend, vroegst beginnend, alfabetisch
    vroege_stoppers_minderjarig.sort(key=lambda x: (x["einduur"], x["startuur"], x["naam"]))
    vroege_stoppers_gewoon.sort(key=lambda x: (x["einduur"], x["startuur"], x["naam"]))
    
    pp2_geplaatste_pauzes = []
    pp2_niet_geplaatst = []
    
    # -----------------------------
    # STAP 1a: minderjarige vroege stoppers
    # Pauze 1: halfuur (2 opeenvolgende kwartieren) zo vroeg mogelijk
    # Pauze 2: kwartier zo laat mogelijk
    # Beide pauzes: niet in eerste of laatste werkuur
    # Pauze 2 zo mogelijk op dezelfde PV-rij als pauze 1
    # -----------------------------
    pp2_minderjarige_vroege_stopper_rij = {}
    
    if pv_rows_pp2:
        for idx, item in enumerate(vroege_stoppers_minderjarig):
            naam = item["naam"]
            werk_uren = item["werk_uren"]
            eerste_uur = werk_uren[0]
            laatste_uur = werk_uren[-1]
    
            pv_index = idx % len(pv_rows_pp2)
            pv, pv_name_row = pv_rows_pp2[pv_index]
            pv_label = pv["naam"]
    
            # -- Pauze 1: halfuur zo vroeg mogelijk --
            col1_gekozen = None
            for i in range(len(pauze_cols_pp2) - 1):
                col1 = pauze_cols_pp2[i]
                col2 = pauze_cols_pp2[i + 1]
    
                # opeenvolgende kwartieren
                if col2 != col1 + 1:
                    continue
    
                uur1 = parse_header_uur(ws_pp2.cell(1, col1).value)
                uur2 = parse_header_uur(ws_pp2.cell(1, col2).value)
    
                if uur1 is None or uur2 is None:
                    continue
    
                # niet in eerste of laatste werkuur
                if uur1 == eerste_uur or uur1 == laatste_uur:
                    continue
                if uur2 == eerste_uur or uur2 == laatste_uur:
                    continue
    
                # student moet beide uren werken
                if uur1 not in werk_uren or uur2 not in werk_uren:
                    continue
    
                # cellen moeten leeg zijn op deze PV-rij
                if not pp2_is_beschikbaar(ws_pp2, pv_name_row, col1):
                    continue
                if not pp2_is_beschikbaar(ws_pp2, pv_name_row, col2):
                    continue
    
                # eerste geldige optie nemen
                col1_gekozen = col1
                break
    
            if col1_gekozen is not None:
                col2_gekozen = col1_gekozen + 1
                pp2_write_name(ws_pp2, pv_name_row, col1_gekozen, naam)
                pp2_write_name(ws_pp2, pv_name_row, col2_gekozen, naam)
                pp2_minderjarige_vroege_stopper_rij[naam] = pv_name_row
    
                pp2_geplaatste_pauzes.append({
                    "naam": naam,
                    "pauzevlinder": pv_label,
                    "tijd": f"{ws_pp2.cell(1, col1_gekozen).value}-{ws_pp2.cell(1, col2_gekozen).value}",
                    "type": "minderjarig vroege stopper - halfuur"
                })
            else:
                pp2_niet_geplaatst.append({
                    "naam": naam,
                    "reden": "geen geldig halfuur gevonden voor minderjarige vroege stopper (pauze 1)"
                })
    
            # -- Pauze 2: kwartier zo laat mogelijk, bij voorkeur zelfde PV-rij --
            vaste_rij = pp2_minderjarige_vroege_stopper_rij.get(naam)
    
            # Kandidaat-kolommen van achter naar voor
            kandidaten = list(reversed(pauze_cols_pp2))
    
            kort_geplaatst = False
            for gebruik_rij in ([vaste_rij] if vaste_rij else []) + [r for (_pv2, r) in pv_rows_pp2 if r != vaste_rij]:
                for col in kandidaten:
                    uur = parse_header_uur(ws_pp2.cell(1, col).value)
                    if uur is None:
                        continue
    
                    # niet in eerste of laatste werkuur
                    if uur == eerste_uur or uur == laatste_uur:
                        continue
    
                    # student moet dat uur werken
                    if uur not in werk_uren:
                        continue
    
                    # cel moet leeg zijn
                    if not pp2_is_beschikbaar(ws_pp2, gebruik_rij, col):
                        continue
    
                    # student mag op dit kwartier nog nergens staan
                    if pp2_student_heeft_al_pauze_op_kolom(naam, col, ws_pp2, pv_rows_pp2):
                        continue
    
                    pp2_write_name(ws_pp2, gebruik_rij, col, naam)
    
                    pp2_geplaatste_pauzes.append({
                        "naam": naam,
                        "pauzevlinder": ws_pp2.cell(gebruik_rij, 1).value or f"rij {gebruik_rij}",
                        "tijd": ws_pp2.cell(1, col).value,
                        "type": "minderjarig vroege stopper - kort kwartier"
                    })
    
                    kort_geplaatst = True
                    break
    
                if kort_geplaatst:
                    break
    
            if not kort_geplaatst:
                pp2_niet_geplaatst.append({
                    "naam": naam,
                    "reden": "geen geldig kwartier gevonden voor minderjarige vroege stopper (pauze 2)"
                })
    
    # -----------------------------
    # STAP 1b: gewone vroege stoppers
    # Inplannen per duo:
    # 1-2 bij PV1, 3-4 bij PV2, 5-6 bij PV3, ...
    # als er meer duo's zijn dan pauzevlinders, dan cyclisch verder
    # Als de voorkeurs-PV-rij al bezet is op de gekozen kolom,
    # worden andere PV-rijen geprobeerd.
    # -----------------------------
    duo_basis_col = {}
    duo_basis_pv_row = {}
    
    if pv_rows_pp2:
        for idx, item in enumerate(vroege_stoppers_gewoon):
            naam = item["naam"]
    
            duo_nummer = idx // 2
            pv_index_voorkeur = duo_nummer % len(pv_rows_pp2)
    
            # Eerste van het duo
            if idx % 2 == 0:
                gekozen_col = pp2_choose_middle_col(naam, ws_pp2, pauze_cols_pp2)
    
                if gekozen_col is None:
                    pp2_niet_geplaatst.append({
                        "naam": naam,
                        "reden": "geen geldige middenplek gevonden voor eerste van duo"
                    })
                    continue
    
                # Probeer eerst voorkeurs-PV-rij, daarna de rest
                pv_volgorde = (
                    [pv_rows_pp2[pv_index_voorkeur]]
                    + [r for i, r in enumerate(pv_rows_pp2) if i != pv_index_voorkeur]
                )
    
                geplaatst_eerste = False
                for pv, pv_name_row in pv_volgorde:
                    if not pp2_is_beschikbaar(ws_pp2, pv_name_row, gekozen_col):
                        continue
    
                    pp2_write_name(ws_pp2, pv_name_row, gekozen_col, naam)
                    duo_basis_col[duo_nummer] = gekozen_col
                    duo_basis_pv_row[duo_nummer] = pv_name_row
    
                    pp2_geplaatste_pauzes.append({
                        "naam": naam,
                        "pauzevlinder": pv["naam"],
                        "tijd": ws_pp2.cell(1, gekozen_col).value,
                        "type": "eerste van duo"
                    })
                    geplaatst_eerste = True
                    break
    
                if not geplaatst_eerste:
                    pp2_niet_geplaatst.append({
                        "naam": naam,
                        "reden": "geen geldige middenplek gevonden voor eerste van duo (alle rijen bezet)"
                    })
    
            # Tweede van het duo
            else:
                basis_col = duo_basis_col.get(duo_nummer)
                pv_name_row = duo_basis_pv_row.get(duo_nummer)
    
                if basis_col is None or pv_name_row is None:
                    pp2_niet_geplaatst.append({
                        "naam": naam,
                        "reden": "geen basisplek beschikbaar van eerste duo-genoot"
                    })
                    continue
    
                pv_label = next(
                    (pv["naam"] for pv, r in pv_rows_pp2 if r == pv_name_row),
                    f"rij {pv_name_row}"
                )
    
                buur_cols = []
                if basis_col - 1 in pauze_cols_pp2:
                    buur_cols.append(basis_col - 1)
                if basis_col + 1 in pauze_cols_pp2:
                    buur_cols.append(basis_col + 1)
    
                geplaatste_tweede = False
    
                for buur_col in buur_cols:
                    if not pp2_is_beschikbaar(ws_pp2, pv_name_row, buur_col):
                        continue
    
                    if not pp2_is_valid_short_break_for_student(naam, buur_col, ws_pp2):
                        continue
    
                    pp2_write_name(ws_pp2, pv_name_row, buur_col, naam)
    
                    pp2_geplaatste_pauzes.append({
                        "naam": naam,
                        "pauzevlinder": pv_label,
                        "tijd": ws_pp2.cell(1, buur_col).value,
                        "type": "tweede van duo"
                    })
    
                    geplaatste_tweede = True
                    break
    
                if not geplaatste_tweede:
                    pp2_niet_geplaatst.append({
                        "naam": naam,
                        "reden": "geen geldige buurplek gevonden voor tweede van duo"
                    })
    
    
    #STAP 2 2222222222222222222222222222222222222222222222222222222222222222222222222222222222222222222222222222222222222
    
    # -----------------------------
    # STAP 2 PP optie 2:
    # lange pauzes invullen van links naar rechts,
    # per halfuurblok en per pauzevlinder
    # -----------------------------
    
    lichtgroen_fill = PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid")
    
    def pp2_heeft_al_lange_pauze(naam, ws_sheet, pv_rows, pauze_cols):
        """
        Check of naam al ergens een dubbele blok heeft in PP optie 2.
        """
        for _pv, pv_row in pv_rows:
            for idx in range(len(pauze_cols) - 1):
                col1 = pauze_cols[idx]
                col2 = pauze_cols[idx + 1]
                if (
                    ws_sheet.cell(pv_row, col1).value == naam and
                    ws_sheet.cell(pv_row, col2).value == naam
                ):
                    return True
        return False
    
    
    def pp2_lange_werkers_lijst():
        """
        Studenten die in stap 2 recht hebben op een halfuur pauze:
        - alle minderjarigen met minstens 4 uur werk
        - alle overige studenten met meer dan 6 uur werk
        - inclusief pauzevlinders indien ze eraan voldoen
        """
        result = []
        al_toegevoegd = set()
    
        for s in studenten:
            naam = s["naam"]
            gewerkte_uren = student_totalen.get(naam, 0)
            is_minderjarig = "-18" in str(naam)
    
            if is_minderjarig and gewerkte_uren >= 4:
                if naam not in al_toegevoegd:
                    result.append(naam)
                    al_toegevoegd.add(naam)
            elif gewerkte_uren > 6:
                if naam not in al_toegevoegd:
                    result.append(naam)
                    al_toegevoegd.add(naam)
    
        return result
    
    
    
    def pp2_aantal_lange_pauzes_nodig_in_stap2(naam):
        """
        Hoeveel halfuren moet deze student in stap 2 krijgen?
        - minderjarige met < 4u werk => 0
        - minderjarige met >= 4u en <= 6u werk => 1
        - minderjarige met > 6u werk => 2
        - niet-minderjarige met > 6u werk => 1
        - anders => 0
        """
        gewerkte_uren = student_totalen.get(naam, 0)
        is_minderjarig = "-18" in str(naam)
    
        if is_minderjarig:
            if gewerkte_uren < 4:
                return 0
            if gewerkte_uren > 6:
                return 2
            return 1
    
        if gewerkte_uren > 6:
            return 1
    
        return 0
    
    
    def pp2_sort_step2_namen(namenlijst):
        """
        Sorteer voor stap 2:
        - eerst wie vroeger stopt
        - bij gelijke eindtijd: random volgorde
        """
        per_einduur = defaultdict(list)
    
        for naam in namenlijst:
            werk_uren = pp2_get_student_work_hours(naam)
            if werk_uren:
                einduur = max(werk_uren)
                per_einduur[einduur].append(naam)
    
        resultaat = []
        for einduur in sorted(per_einduur.keys()):
            groep = per_einduur[einduur][:]
            random.shuffle(groep)
            resultaat.extend(groep)
    
        return resultaat
    
    def pp2_get_pv_row_for_name(naam, pv_rows):
        """
        Geef de naamrij terug van de pauzevlinder met deze naam.
        """
        for pv, pv_row in pv_rows:
            if pv["naam"] == naam:
                return pv_row
        return None
    
    
    def pp2_find_first_valid_long_block_any_row(naam, ws_sheet, pv_rows, pauze_cols):
        """
        Zoek het vroegst mogelijke geldige halfuur voor deze student
        over alle pauzevlinder-rijen heen, van links naar rechts.
        Retourneert (pv_row, col1, col2) of None.
        """
        blokken = pp2_halfuur_blokken(pauze_cols, ws_sheet)
    
        for col1, col2 in blokken:
            for _pv, pv_row in pv_rows:
                if ws_sheet.cell(pv_row, col1).value not in [None, ""]:
                    continue
                if ws_sheet.cell(pv_row, col2).value not in [None, ""]:
                    continue
    
                if not pp2_is_valid_long_break_for_student(naam, col1, col2, ws_sheet):
                    continue
    
                return (pv_row, col1, col2)
    
        return None
    
    
    def pp2_find_first_valid_long_block_on_fixed_row(naam, ws_sheet, pv_row, pauze_cols):
        """
        Zoek het vroegst mogelijke geldige halfuur voor deze student
        op één vaste pauzevlinder-rij, van links naar rechts.
        Retourneert (col1, col2) of None.
        """
        blokken = pp2_halfuur_blokken(pauze_cols, ws_sheet)
    
        for col1, col2 in blokken:
            if ws_sheet.cell(pv_row, col1).value not in [None, ""]:
                continue
            if ws_sheet.cell(pv_row, col2).value not in [None, ""]:
                continue
    
            if not pp2_is_valid_long_break_for_student(naam, col1, col2, ws_sheet):
                continue
    
            return (col1, col2)
    
        return None
        
    
    
    def pp2_is_valid_long_break_for_student(naam, col1, col2, ws_sheet):
        """
        Een lange pauze mag alleen als:
        - beide kwartieren samen exact 30 min vormen
        - student werkt in beide kwartieren
        - niet in eerste of laatste werkuur
        - student op geen van beide kwartieren al elders in het pauzerooster staat
        """
        header1 = ws_sheet.cell(1, col1).value
        header2 = ws_sheet.cell(1, col2).value
    
        mins1 = pp2_parse_kwartier_header(header1)
        mins2 = pp2_parse_kwartier_header(header2)
    
        if mins1 is None or mins2 is None:
            return False
    
        if mins2 - mins1 != 15:
            return False
    
        werk_uren = pp2_get_student_work_hours(naam)
        if not werk_uren:
            return False
    
        uur1 = parse_header_uur(header1)
        uur2 = parse_header_uur(header2)
    
        if uur1 is None or uur2 is None:
            return False
    
        if uur1 not in werk_uren or uur2 not in werk_uren:
            return False
    
        eerste_uur = werk_uren[0]
        laatste_uur = werk_uren[-1]
    
        if uur1 == eerste_uur or uur1 == laatste_uur:
            return False
        if uur2 == eerste_uur or uur2 == laatste_uur:
            return False
    
        if pp2_student_heeft_al_pauze_op_kolom(
            naam=naam,
            col=col1,
            ws_sheet=ws_sheet,
            pv_rows=pv_rows_pp2
        ):
            return False
    
        if pp2_student_heeft_al_pauze_op_kolom(
            naam=naam,
            col=col2,
            ws_sheet=ws_sheet,
            pv_rows=pv_rows_pp2
        ):
            return False
    
        return True
        
    
    def pp2_write_long_break(ws_sheet, pv_row, col1, col2, naam, leave_top_blank=False):
        """
        Schrijf een lange pauze van 2 kwartieren:
        - normaal: attractie erboven
        - voor pauzevlinder op eigen rij: bovenste cel leeg laten
        - naam in beide vakjes
        - groen kleuren
        """
        for col in [col1, col2]:
            info_cel = ws_sheet.cell(pv_row - 1, col)
            info_cel.alignment = center_align
            info_cel.border = thin_border
    
            if leave_top_blank:
                info_cel.value = ""
            else:
                header = ws_sheet.cell(1, col).value
                uur = parse_header_uur(header)
                attr = vind_attractie_op_uur(naam, uur) if uur is not None else None
                info_cel.value = attr if attr else ""
    
            naam_cel = ws_sheet.cell(pv_row, col)
            naam_cel.value = naam
            naam_cel.alignment = center_align
            naam_cel.border = thin_border
            naam_cel.fill = lichtgroen_fill
    
    
    def pp2_halfuur_blokken(pauze_cols, ws_sheet):
        """
        Geeft alle mogelijke halfuurblokken terug, van links naar rechts.
        Flexibel:
        - mag starten op heel uur
        - mag ook starten op :15
        Dus bv.:
        (12u00, 12u15), (12u15, 12u30), (12u30, 12u45), ...
        zolang de cellen exact 15 minuten uit elkaar liggen.
        """
        blokken = []
    
        for idx in range(len(pauze_cols) - 1):
            col1 = pauze_cols[idx]
            col2 = pauze_cols[idx + 1]
    
            mins1 = pp2_parse_kwartier_header(ws_sheet.cell(1, col1).value)
            mins2 = pp2_parse_kwartier_header(ws_sheet.cell(1, col2).value)
    
            if mins1 is None or mins2 is None:
                continue
    
            if mins2 - mins1 == 15:
                blokken.append((col1, col2))
    
        return blokken
    
    
    def pp2_place_long_break_for_pv_in_own_row(pv, pv_name_row, ws_sheet, pauze_cols, lange_pauze_ontvangers, lange_werkers_random):
        """
        Geef een langwerkende pauzevlinder verplicht een lange pauze in de eigen rij.
        We proberen de blokken strikt van links naar rechts.
        De cellen erboven blijven leeg.
        """
        naam = pv["naam"]
    
        if naam not in lange_werkers_random:
            return False
    
        if naam in lange_pauze_ontvangers:
            return False
    
        blokken = pp2_halfuur_blokken(pauze_cols, ws_sheet)
    
        for col1, col2 in blokken:
            # beide kwartieren moeten leeg zijn op eigen rij
            if ws_sheet.cell(pv_name_row, col1).value not in [None, ""]:
                continue
            if ws_sheet.cell(pv_name_row, col2).value not in [None, ""]:
                continue
    
            if not pp2_is_valid_long_break_for_student(naam, col1, col2, ws_sheet):
                continue
    
            pp2_write_long_break(
                ws_sheet=ws_sheet,
                pv_row=pv_name_row,
                col1=col1,
                col2=col2,
                naam=naam,
                leave_top_blank=True
            )
            lange_pauze_ontvangers.add(naam)
            return True
    
        return False
    
    
    # 1) Bouw de kandidatenlijsten voor stap 2
    pp2_step2_basis = pp2_lange_werkers_lijst()
    
    pp2_step2_minderjarigen = []
    pp2_step2_overige_lange_werkers = []
    
    for naam in pp2_step2_basis:
        if "-18" in str(naam):
            pp2_step2_minderjarigen.append(naam)
        else:
            pp2_step2_overige_lange_werkers.append(naam)
    
    pp2_step2_minderjarigen = pp2_sort_step2_namen(pp2_step2_minderjarigen)
    pp2_step2_overige_lange_werkers = pp2_sort_step2_namen(pp2_step2_overige_lange_werkers)
    
    # Deze lijst houden we voor de bestaande latere logica aan
    pp2_lange_werkers_random = pp2_step2_minderjarigen + pp2_step2_overige_lange_werkers
    
    # 2) Houd bij wie al minstens één lange pauze kreeg
    pp2_lange_pauze_ontvangers = set()
    for naam in pp2_lange_werkers_random:
        if pp2_heeft_al_lange_pauze(naam, ws_pp2, pv_rows_pp2, pauze_cols_pp2):
            pp2_lange_pauze_ontvangers.add(naam)
    
    # Voor minderjarigen willen we onthouden op welke rij hun EERSTE halfuur kwam
    pp2_minderjarige_eerste_halfuur_rij = {}
    
    # 3) Eerst: alle minderjarigen die in stap 2 recht hebben op een halfuur
    #    krijgen hun EERSTE halfuur zo vroeg mogelijk
    for naam in pp2_step2_minderjarigen:
        nodig = pp2_aantal_lange_pauzes_nodig_in_stap2(naam)
        if nodig <= 0:
            continue
    
        # Heeft al ergens een lang halfuur? Dan niet nog eens als "eerste" plaatsen
        if naam in pp2_lange_pauze_ontvangers:
            continue
    
        gevonden = pp2_find_first_valid_long_block_any_row(
            naam=naam,
            ws_sheet=ws_pp2,
            pv_rows=pv_rows_pp2,
            pauze_cols=pauze_cols_pp2
        )
    
        if gevonden is None:
            continue
    
        pv_row, col1, col2 = gevonden
    
        eigen_pv_row = pp2_get_pv_row_for_name(naam, pv_rows_pp2)
        leave_top_blank = eigen_pv_row == pv_row
    
        pp2_write_long_break(
            ws_sheet=ws_pp2,
            pv_row=pv_row,
            col1=col1,
            col2=col2,
            naam=naam,
            leave_top_blank=leave_top_blank
        )
    
        pp2_lange_pauze_ontvangers.add(naam)
        pp2_minderjarige_eerste_halfuur_rij[naam] = pv_row
    
    # 4) Daarna: bestaande logica voor overige lange pauzevlinders op eigen rij
    for pv, pv_name_row in pv_rows_pp2:
        pp2_place_long_break_for_pv_in_own_row(
            pv=pv,
            pv_name_row=pv_name_row,
            ws_sheet=ws_pp2,
            pauze_cols=pauze_cols_pp2,
            lange_pauze_ontvangers=pp2_lange_pauze_ontvangers,
            lange_werkers_random=pp2_lange_werkers_random
        )
    
    # 5) Daarna: algemene verdeling van andere lange werkers
    pp2_blokken = pp2_halfuur_blokken(pauze_cols_pp2, ws_pp2)
    
    for col1, col2 in pp2_blokken:
        for pv, pv_name_row in pv_rows_pp2:
            if not pp2_is_beschikbaar(ws_pp2, pv_name_row, col1):
                continue
            if not pp2_is_beschikbaar(ws_pp2, pv_name_row, col2):
                continue
    
            toegewezen_naam = None
    
            for kandidaat in pp2_step2_overige_lange_werkers:
                if kandidaat in pp2_lange_pauze_ontvangers:
                    continue
                if pp2_is_valid_long_break_for_student(kandidaat, col1, col2, ws_pp2):
                    toegewezen_naam = kandidaat
                    break
    
            if toegewezen_naam:
                pp2_write_long_break(
                    ws_sheet=ws_pp2,
                    pv_row=pv_name_row,
                    col1=col1,
                    col2=col2,
                    naam=toegewezen_naam,
                    leave_top_blank=False
                )
                pp2_lange_pauze_ontvangers.add(toegewezen_naam)
    
    # 6) Helemaal als laatste:
    #    minderjarigen met > 6u krijgen nog een TWEEDE halfuur
    #    op exact dezelfde rij als hun eerste halfuur
    for naam in pp2_step2_minderjarigen:
        if pp2_aantal_lange_pauzes_nodig_in_stap2(naam) < 2:
            continue
    
        vaste_rij = pp2_minderjarige_eerste_halfuur_rij.get(naam)
        if vaste_rij is None:
            continue
    
        gevonden = pp2_find_first_valid_long_block_on_fixed_row(
            naam=naam,
            ws_sheet=ws_pp2,
            pv_row=vaste_rij,
            pauze_cols=pauze_cols_pp2
        )
    
        if gevonden is None:
            continue
    
        col1, col2 = gevonden
    
        eigen_pv_row = pp2_get_pv_row_for_name(naam, pv_rows_pp2)
        leave_top_blank = eigen_pv_row == vaste_rij
    
        pp2_write_long_break(
            ws_sheet=ws_pp2,
            pv_row=vaste_rij,
            col1=col1,
            col2=col2,
            naam=naam,
            leave_top_blank=leave_top_blank
        )
    
    
    
    #STAP 3 333333333333333333333333333333333333333333333333333333333333333333333333333333333333333333333333333333333
    
    # -----------------------------
    # STAP 3 PP optie 2:
    # open spots berekenen en verdelen
    # + korte pauzes van pauzevlinders zelf invullen
    # -----------------------------
    
    lichtpaars_fill = PatternFill(start_color="E6DAF7", end_color="E6DAF7", fill_type="solid")
    naam_leeg_fill_pp2 = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
    
    
    
    def pp2_benodigde_korte_kwartieren(naam):
        """
        Nieuwe regel voor PP optie 2:
        - < 4 uur gewerkt => 0 korte kwartieren
        - >= 4 uur gewerkt => 1 kort kwartier
    
        Dit geldt nu ook voor minderjarigen:
        - minderjarige 4u t.e.m. 6u => 1 kort kwartier
        - minderjarige > 6u => ook 1 kort kwartier
        """
        gewerkte_uren = student_totalen.get(naam, 0)
    
        if gewerkte_uren < 4:
            return 0
    
        return 1
    
    
    def pp2_count_total_assigned_quarters_for_student(naam, ws_sheet, pv_rows, pauze_cols):
        """
        Telt alle kwartiercellen in PP optie 2 waar deze naam al staat.
        Dit zijn dus ALLE reeds toegekende pauzekwartieren samen.
        """
        count = 0
        for _pv, pv_row in pv_rows:
            for col in pauze_cols:
                if ws_sheet.cell(pv_row, col).value == naam:
                    count += 1
        return count
    
    
    def pp2_count_al_toegekende_lange_kwartieren(naam, ws_sheet, pv_rows, pauze_cols):
        """
        Telt hoeveel reeds toegekende kwartieren deel uitmaken van een LANGE pauze
        voor deze student in PP optie 2.
    
        Een lange pauze herkennen we als 2 opeenvolgende kwartieren op dezelfde rij
        met exact dezelfde naam.
    
        Voorbeelden:
        - 1 halfuur lange pauze => 2 kwartieren
        - 2 halve uren lange pauze => 4 kwartieren
        """
        count = 0
        gebruikte_cols_per_row = set()
    
        for _pv, pv_row in pv_rows:
            for idx in range(len(pauze_cols) - 1):
                col1 = pauze_cols[idx]
                col2 = pauze_cols[idx + 1]
    
                if (pv_row, col1) in gebruikte_cols_per_row or (pv_row, col2) in gebruikte_cols_per_row:
                    continue
    
                val1 = ws_sheet.cell(pv_row, col1).value
                val2 = ws_sheet.cell(pv_row, col2).value
    
                if val1 == naam and val2 == naam:
                    count += 2
                    gebruikte_cols_per_row.add((pv_row, col1))
                    gebruikte_cols_per_row.add((pv_row, col2))
    
        return count
    
    
    def pp2_count_al_toegekende_korte_kwartieren(naam, ws_sheet, pv_rows, pauze_cols, lange_pauze_ontvangers):
        """
        Telt hoeveel KORTE kwartieren deze student al heeft.
    
        Nieuwe logica:
        - tel eerst alle reeds ingevulde kwartieren van deze student
        - trek daar alle kwartieren af die deel uitmaken van een lange pauze
        - wat overblijft, zijn korte kwartieren
    
        Hierdoor werkt dit ook correct voor minderjarigen die 2 halve uren kregen.
        """
        totaal = pp2_count_total_assigned_quarters_for_student(
            naam=naam,
            ws_sheet=ws_sheet,
            pv_rows=pv_rows,
            pauze_cols=pauze_cols
        )
    
        lange_kwartieren = pp2_count_al_toegekende_lange_kwartieren(
            naam=naam,
            ws_sheet=ws_sheet,
            pv_rows=pv_rows,
            pauze_cols=pauze_cols
        )
    
        return max(0, totaal - lange_kwartieren)
    
    
    def pp2_resterende_korte_kwartieren(naam, ws_sheet, pv_rows, pauze_cols, lange_pauze_ontvangers):
        """
        Hoeveel korte kwartieren heeft deze student nog nodig?
        """
        nodig = pp2_benodigde_korte_kwartieren(naam)
        al_kort = pp2_count_al_toegekende_korte_kwartieren(
            naam=naam,
            ws_sheet=ws_sheet,
            pv_rows=pv_rows,
            pauze_cols=pauze_cols,
            lange_pauze_ontvangers=lange_pauze_ontvangers
        )
        return max(0, nodig - al_kort)
    
    
    def pp2_heeft_al_voldoende_korte_pauze(naam, ws_sheet, pv_rows, pauze_cols, lange_pauze_ontvangers):
        """
        True als student al genoeg korte kwartieren heeft gekregen
        volgens de nieuwe PP2-regels.
        """
        return pp2_resterende_korte_kwartieren(
            naam=naam,
            ws_sheet=ws_sheet,
            pv_rows=pv_rows,
            pauze_cols=pauze_cols,
            lange_pauze_ontvangers=lange_pauze_ontvangers
        ) == 0
    
    
    def pp2_korte_pauze_nodig_namen():
        """
        Iedereen met minstens 4 uur werk heeft recht op 1 kort kwartier,
        BEHALVE minderjarige vroege stoppers.
    
        Minderjarige vroege stoppers:
        - minderjarig
        - minstens 4u gewerkt
        - laatste werkuur <= 15
        """
        namen = []
    
        for s in studenten:
            naam = s["naam"]
            werk_uren = pp2_get_student_work_hours(naam)
    
            is_minor_early_stopper = (
                pp2_is_minderjarig(naam)
                and len(werk_uren) >= 4
                and werk_uren
                and max(werk_uren) <= 15
            )
    
            if is_minor_early_stopper:
                continue
    
            if pp2_benodigde_korte_kwartieren(naam) > 0:
                namen.append(naam)
    
        return namen
    
    
    
    def pp2_count_remaining_empty_quarters(ws_sheet, pv_rows, pauze_cols):
        """
        Telt alle nog lege kwartiercellen in de naamrijen van PP optie 2.
        """
        count = 0
        for _pv, pv_row in pv_rows:
            for col in pauze_cols:
                if ws_sheet.cell(pv_row, col).value in [None, ""]:
                    count += 1
        return count
    
    
    def pp2_get_empty_cols_for_pv_row(ws_sheet, pv_row, pauze_cols, open_spots_set):
        """
        Geeft alle lege kwartierkolommen terug voor deze pauzevlinder-rij,
        exclusief reeds gemarkeerde open spots.
        """
        cols = []
        for col in pauze_cols:
            if (pv_row, col) in open_spots_set:
                continue
            if ws_sheet.cell(pv_row, col).value in [None, ""]:
                cols.append(col)
        return cols
    
    
    def pp2_mark_open_spot(ws_sheet, pv_row, col):
        """
        Open spot blijft gewoon blauw en leeg.
        """
        top_cel = ws_sheet.cell(pv_row - 1, col)
        top_cel.value = ""
        top_cel.alignment = center_align
        top_cel.border = thin_border
    
        cel = ws_sheet.cell(pv_row, col)
        cel.value = ""
        cel.alignment = center_align
        cel.border = thin_border
        cel.fill = naam_leeg_fill_pp2
    
    
    
    
    def pp2_write_short_break_for_pv(ws_sheet, pv_row, col, naam):
        """
        Schrijf 1 kort kwartier voor een pauzevlinder zelf:
        - bovenliggende cel leeg
        - naam paars
        """
        top_cel = ws_sheet.cell(pv_row - 1, col)
        top_cel.value = ""
        top_cel.alignment = center_align
        top_cel.border = thin_border
    
        cel = ws_sheet.cell(pv_row, col)
        cel.value = naam
        cel.alignment = center_align
        cel.border = thin_border
        cel.fill = lichtpaars_fill
    
    
    def pp2_find_short_break_cols_for_pv(naam, pv_row, ws_sheet, pauze_cols, open_spots_set, needed_quarters):
        """
        Zoek geldige kolom/kolommen voor de korte pauze van een pauzevlinder in de eigen rij.
    
        - needed_quarters == 1:
          neem het eerstvolgende geldige vrije kwartier
    
        - needed_quarters == 2:
          neem de eerste geldige set van 2 opeenvolgende kwartieren
        """
        if needed_quarters <= 0:
            return []
    
        if needed_quarters == 1:
            for col in pauze_cols:
                if (pv_row, col) in open_spots_set:
                    continue
                if ws_sheet.cell(pv_row, col).value not in [None, ""]:
                    continue
                if not pp2_is_valid_short_break_for_student(naam, col, ws_sheet):
                    continue
                return [col]
            return []
    
        if needed_quarters == 2:
            for idx in range(len(pauze_cols) - 1):
                col1 = pauze_cols[idx]
                col2 = pauze_cols[idx + 1]
    
                if col2 != col1 + 1:
                    continue
    
                if (pv_row, col1) in open_spots_set or (pv_row, col2) in open_spots_set:
                    continue
    
                if ws_sheet.cell(pv_row, col1).value not in [None, ""]:
                    continue
                if ws_sheet.cell(pv_row, col2).value not in [None, ""]:
                    continue
    
                if not pp2_is_valid_short_break_for_student(naam, col1, ws_sheet):
                    continue
                if not pp2_is_valid_short_break_for_student(naam, col2, ws_sheet):
                    continue
    
                return [col1, col2]
    
            return []
    
        return []
    
    
    
    # 1) Is dit een korte dag?
    #    Korte dag = openingsuren zijn 6 uur of minder
    pp2_is_korte_dag = len(open_uren) <= 6
    
    pp2_open_spots = set()
    pp2_pv_short_breaks_placed = []
    
    # -----------------------------
    # Hulploop: korte pauzes van pauzevlinders zelf invullen
    #
    # Nieuwe regel:
    # - pauzevlinders die GEEN lange werker zijn: hier al plaatsen
    # - pauzevlinders die WEL lange werker zijn: nog NIET hier plaatsen
    #   -> die komen later in stap 4, na de korte werkers
    # -----------------------------
    for pv, pv_row in pv_rows_pp2:
        naam = pv["naam"]
    
        # Lange pauzevlinders hier nog overslaan:
        # hun korte pauze moet pas later komen
        if naam in pp2_lange_werkers_lijst():
            continue
    
        resterend_nodig = pp2_resterende_korte_kwartieren(
            naam=naam,
            ws_sheet=ws_pp2,
            pv_rows=pv_rows_pp2,
            pauze_cols=pauze_cols_pp2,
            lange_pauze_ontvangers=pp2_lange_pauze_ontvangers
        )
    
        if resterend_nodig <= 0:
            continue
    
        gekozen_cols = pp2_find_short_break_cols_for_pv(
            naam=naam,
            pv_row=pv_row,
            ws_sheet=ws_pp2,
            pauze_cols=pauze_cols_pp2,
            open_spots_set=pp2_open_spots,
            needed_quarters=resterend_nodig
        )
    
        if not gekozen_cols:
            continue
    
        for col in gekozen_cols:
            pp2_write_short_break_for_pv(ws_pp2, pv_row, col, naam)
    
        pp2_pv_short_breaks_placed.append({
            "naam": naam,
            "kolommen": gekozen_cols,
            "tijden": [ws_pp2.cell(1, col).value for col in gekozen_cols]
        })
    
    # -----------------------------
    # 2) Tellen hoeveel kwartierblokjes nog leeg zijn
    # -----------------------------
    pp2_remaining_empty_quarters = pp2_count_remaining_empty_quarters(
        ws_sheet=ws_pp2,
        pv_rows=pv_rows_pp2,
        pauze_cols=pauze_cols_pp2
    )
    
    # -----------------------------
    # 3) Tellen hoeveel KORTE kwartieren nog gegeven moeten worden
    #    Nieuwe telling:
    #    - gewone student meestal 1
    #    - minderjarige >4u = 2
    #    - ook minderjarige >6u met al lange pauze telt hier nog voor 2
    # -----------------------------
    pp2_korte_pauze_gerechtigden = pp2_korte_pauze_nodig_namen()
    
    pp2_remaining_short_quarters_needed = 0
    for naam in pp2_korte_pauze_gerechtigden:
        pp2_remaining_short_quarters_needed += pp2_resterende_korte_kwartieren(
            naam=naam,
            ws_sheet=ws_pp2,
            pv_rows=pv_rows_pp2,
            pauze_cols=pauze_cols_pp2,
            lange_pauze_ontvangers=pp2_lange_pauze_ontvangers
        )
    
    # -----------------------------
    # 4) Open spots berekenen
    # -----------------------------
    pp2_open_spots_count = pp2_remaining_empty_quarters - pp2_remaining_short_quarters_needed
    if pp2_open_spots_count < 0:
        pp2_open_spots_count = 0
    
    # -----------------------------
    # 5) Open spots verdelen
    #
    # KORTE DAG:
    # - eerst korte pauzes geplaatst
    # - dus open spots vallen automatisch NA de korte pauzes
    #
    # LANGE DAG:
    # - script blijft exact hetzelfde gedrag houden als nu
    #   => open spots verdelen zoals nu
    # -----------------------------
    
    if not pp2_is_korte_dag:
        # ---------------------------------------------------
        # LANGE DAG:
        # - Open spots eerst verdelen (zoals in het originele script)
        # - Daarna enkel de korte pauzes van KORTE pauzevlinders plaatsen
        # - Lange pauzevlinders komen pas in stap 4 aan bod
        # ---------------------------------------------------
    
        # Reset eerst eventuele eerder geplaatste korte pauzes van pauzevlinders
        for item in pp2_pv_short_breaks_placed:
            naam = item["naam"]
    
            pv_row = next(
                pv_row for pv, pv_row in pv_rows_pp2
                if pv["naam"] == naam
            )
    
            for col in item["kolommen"]:
                top_cel = ws_pp2.cell(pv_row - 1, col)
                top_cel.value = ""
                top_cel.alignment = center_align
                top_cel.border = thin_border
    
                cel = ws_pp2.cell(pv_row, col)
                cel.value = ""
                cel.alignment = center_align
                cel.border = thin_border
                cel.fill = naam_leeg_fill_pp2
    
        pp2_pv_short_breaks_placed = []
    
        # ---------------------------------------------------
        # 1) Open spots verdelen
        # ---------------------------------------------------
        ronde_nummer = 0
    
        while len(pp2_open_spots) < pp2_open_spots_count:
            iets_geplaatst_deze_ronde = False
            vooraan = (ronde_nummer % 2 == 0)
    
            for _pv, pv_row in pv_rows_pp2:
                if len(pp2_open_spots) >= pp2_open_spots_count:
                    break
    
                lege_cols = pp2_get_empty_cols_for_pv_row(
                    ws_sheet=ws_pp2,
                    pv_row=pv_row,
                    pauze_cols=pauze_cols_pp2,
                    open_spots_set=pp2_open_spots
                )
    
                if not lege_cols:
                    continue
    
                gekozen_col = lege_cols[0] if vooraan else lege_cols[-1]
    
                pp2_open_spots.add((pv_row, gekozen_col))
                pp2_mark_open_spot(ws_pp2, pv_row, gekozen_col)
                iets_geplaatst_deze_ronde = True
    
            if not iets_geplaatst_deze_ronde:
                break
    
            ronde_nummer += 1
    
        # ---------------------------------------------------
        # 2) Enkel korte pauzes van KORTE pauzevlinders plaatsen
        # Lange pauzevlinders worden hier overgeslagen
        # ---------------------------------------------------
        for pv, pv_row in pv_rows_pp2:
            naam = pv["naam"]
    
            # Lange pauzevlinders hier overslaan
            if naam in pp2_lange_werkers_lijst():
                continue
    
            resterend_nodig = pp2_resterende_korte_kwartieren(
                naam=naam,
                ws_sheet=ws_pp2,
                pv_rows=pv_rows_pp2,
                pauze_cols=pauze_cols_pp2,
                lange_pauze_ontvangers=pp2_lange_pauze_ontvangers
            )
    
            if resterend_nodig <= 0:
                continue
    
            gekozen_cols = pp2_find_short_break_cols_for_pv(
                naam=naam,
                pv_row=pv_row,
                ws_sheet=ws_pp2,
                pauze_cols=pauze_cols_pp2,
                open_spots_set=pp2_open_spots,
                needed_quarters=resterend_nodig
            )
    
            if not gekozen_cols:
                continue
    
            for col in gekozen_cols:
                pp2_write_short_break_for_pv(
                    ws_sheet=ws_pp2,
                    pv_row=pv_row,
                    col=col,
                    naam=naam
                )
    
            pp2_pv_short_breaks_placed.append({
                "naam": naam,
                "kolommen": gekozen_cols,
                "tijden": [
                    ws_pp2.cell(1, col).value for col in gekozen_cols
                ]
            })
    
    
    #STAP 4 44444444444444444444444444444444444444444444444444444444444444444444444444444444444444444444444444444444444
    
    
    # -----------------------------
    # STAP 4 PP optie 2:
    # korte pauzes voor:
    # 1) studenten die vroeger stoppen dan het einduur
    # 2) daarna lange pauzevlinders zelf (in eigen rij)
    # met nieuwe minderjarigenlogica
    # -----------------------------
    
    lichtpaars_fill = PatternFill(start_color="E6DAF7", end_color="E6DAF7", fill_type="solid")
    
    
    def pp2_get_day_end_hour():
        """
        Einduur van de dag op basis van open_uren.
        """
        if not open_uren:
            return None
        return max(open_uren)
    
    
    def pp2_get_students_stopping_before_end():
        """
        Studenten die vroeger stoppen dan het einduur van de dag
        en minstens 4 uur werken.
        """
        einduur_dag = pp2_get_day_end_hour()
        result = []
    
        if einduur_dag is None:
            return result
    
        for s in studenten:
            naam = s["naam"]
            werk_uren = pp2_get_student_work_hours(naam)
    
            if len(werk_uren) < 4:
                continue
    
            if max(werk_uren) < einduur_dag:
                result.append(naam)
    
        return result
    
    
    def pp2_write_short_break_regular(ws_sheet, pv_row, col, naam):
        """
        Korte pauze voor gewone student:
        - bovenliggende cel = attractie
        - naamcel:
            * lichtgeel voor minderjarigen die >4u werken
            * lichtpaars voor alle andere korte pauzes
        """
        header = ws_sheet.cell(1, col).value
        uur = parse_header_uur(header)
    
        attr = vind_attractie_op_uur(naam, uur) if uur is not None else None
    
        top_cel = ws_sheet.cell(pv_row - 1, col)
        top_cel.value = attr if attr else ""
        top_cel.alignment = center_align
        top_cel.border = thin_border
    
        cel = ws_sheet.cell(pv_row, col)
        cel.value = naam
        cel.alignment = center_align
        cel.border = thin_border
    
        if pp2_is_minderjarig(naam) and student_totalen.get(naam, 0) > 4:
            cel.fill = roze_fill
        else:
            cel.fill = lichtpaars_fill
    
    
    def pp2_get_long_break_owners_on_row(ws_sheet, pv_row, pauze_cols):
        """
        Geeft alle studenten terug die op deze rij een lange pauze hebben,
        gesorteerd op het ankerpunt voor hun korte pauze:
        - minderjarige lange werkers (>6u): gesorteerd op hun LAATSTE halfuur
        - alle anderen: gesorteerd op hun EERSTE halfuur (= volgorde van links naar rechts)
        """
        # Verzamel per student de eerste én laatste halfuur-eindkolom op deze rij
        eerste_col = {}
        laatste_col = {}
    
        for idx in range(len(pauze_cols) - 1):
            col1 = pauze_cols[idx]
            col2 = pauze_cols[idx + 1]
    
            val1 = ws_sheet.cell(pv_row, col1).value
            val2 = ws_sheet.cell(pv_row, col2).value
    
            if val1 and val1 == val2:
                naam = str(val1).strip()
                if naam not in eerste_col:
                    eerste_col[naam] = col2
                laatste_col[naam] = col2
    
        owners = list(eerste_col.keys())
    
        def sorteersleutel(naam):
            is_minor_long_worker = (
                pp2_is_minderjarig(naam)
                and student_totalen.get(naam, 0) > 6
            )
            if is_minor_long_worker:
                return laatste_col.get(naam, 0)
            else:
                return eerste_col.get(naam, 0)
    
        owners.sort(key=sorteersleutel)
        return owners
    
    def pp2_student_has_long_break_in_row(naam, ws_sheet, pv_row, pauze_cols):
        """
        Check of deze student een lange pauze heeft op precies deze rij.
        """
        for idx in range(len(pauze_cols) - 1):
            col1 = pauze_cols[idx]
            col2 = pauze_cols[idx + 1]
    
            if (
                ws_sheet.cell(pv_row, col1).value == naam and
                ws_sheet.cell(pv_row, col2).value == naam
            ):
                return True
    
        return False
    
    
    def pp2_student_is_long_worker(naam):
        return naam in pp2_lange_werkers_lijst()
    
    
    def pp2_find_two_consecutive_valid_cols_for_student_on_row(naam, pv_row, ws_sheet, pauze_cols, open_spots_set):
        """
        Zoek 2 opeenvolgende geldige kwartieren voor deze student op deze specifieke rij.
        """
        for idx in range(len(pauze_cols) - 1):
            col1 = pauze_cols[idx]
            col2 = pauze_cols[idx + 1]
    
            if col2 != col1 + 1:
                continue
    
            if (pv_row, col1) in open_spots_set or (pv_row, col2) in open_spots_set:
                continue
    
            if ws_sheet.cell(pv_row, col1).value not in [None, ""]:
                continue
            if ws_sheet.cell(pv_row, col2).value not in [None, ""]:
                continue
    
            if not pp2_is_valid_short_break_for_student(naam, col1, ws_sheet):
                continue
            if not pp2_is_valid_short_break_for_student(naam, col2, ws_sheet):
                continue
    
            return [col1, col2]
    
        return []
    
    
    def pp2_find_one_valid_col_for_student_on_row(naam, pv_row, ws_sheet, pauze_cols, open_spots_set):
        """
        Zoek 1 geldig kwartier voor deze student op deze specifieke rij.
        """
        for col in pauze_cols:
            if (pv_row, col) in open_spots_set:
                continue
    
            if ws_sheet.cell(pv_row, col).value not in [None, ""]:
                continue
    
            if not pp2_is_valid_short_break_for_student(naam, col, ws_sheet):
                continue
    
            return [col]
    
        return []
    
    
    def pp2_find_needed_short_cols_for_student_on_row(naam, pv_row, ws_sheet, pauze_cols, open_spots_set, min_col_exclusive=None, zoek_zo_laat_mogelijk=False):
        """
        Zoek het korte kwartier dat deze student nog nodig heeft op deze specifieke rij.
    
        - min_col_exclusive: zoek pas NA deze kolom
        - zoek_zo_laat_mogelijk: zoek van rechts naar links (voor minderjarige lange werkers)
        """
        resterend = pp2_resterende_korte_kwartieren(
            naam=naam,
            ws_sheet=ws_sheet,
            pv_rows=pv_rows_pp2,
            pauze_cols=pauze_cols,
            lange_pauze_ontvangers=pp2_lange_pauze_ontvangers
        )
    
        if resterend <= 0:
            return []
    
        kandidaat_cols = list(reversed(pauze_cols)) if zoek_zo_laat_mogelijk else list(pauze_cols)
    
        for col in kandidaat_cols:
            if min_col_exclusive is not None and col <= min_col_exclusive:
                continue
    
            if (pv_row, col) in open_spots_set:
                continue
    
            if ws_sheet.cell(pv_row, col).value not in [None, ""]:
                continue
    
            if not pp2_is_valid_short_break_for_student(naam, col, ws_sheet):
                continue
    
            return [col]
    
        return []
    
    
    
    def pp2_place_short_break_cols_on_row(naam, pv, pv_row, cols):
        """
        Schrijf 1 of 2 korte kwartieren voor gewone student op een bepaalde rij.
        """
        for col in cols:
            pp2_write_short_break_regular(
                ws_sheet=ws_pp2,
                pv_row=pv_row,
                col=col,
                naam=naam
            )
    
        pp2_regular_short_breaks_placed.append({
            "naam": naam,
            "pauzevlinder": pv["naam"],
            "tijden": [ws_pp2.cell(1, col).value for col in cols],
            "zelfde_rij_als_lange_pauze": pp2_student_has_long_break_in_row(
                naam, ws_pp2, pv_row, pauze_cols_pp2
            )
        })
    
    
    
    
    
    def pp2_student_heeft_nog_lange_pauze_nodig(naam, ws_sheet, pv_rows, pauze_cols):
        """
        Bepaal of deze student volgens de nieuwe regels nog minstens 1 lang halfuur mist.
    
        Regels minderjarigen:
        - < 4u gewerkt => 0 lange pauzes
        - 4u t.e.m. 6u => 1 lange pauze
        - > 6u => 2 lange pauzes
    
        Regels meerderjarigen:
        - > 6u => 1 lange pauze
        - anders => 0
        """
        gewerkte_uren = student_totalen.get(naam, 0)
        is_minor = pp2_is_minderjarig(naam)
    
        if is_minor:
            if gewerkte_uren < 4:
                nodig = 0
            elif gewerkte_uren <= 6:
                nodig = 1
            else:
                nodig = 2
        else:
            nodig = 1 if gewerkte_uren > 6 else 0
    
        # tel hoeveel lange halve uren al effectief ingepland zijn
        al = 0
        for _pv, pv_row in pv_rows:
            for idx in range(len(pauze_cols) - 1):
                col1 = pauze_cols[idx]
                col2 = pauze_cols[idx + 1]
    
                if (
                    ws_sheet.cell(pv_row, col1).value == naam and
                    ws_sheet.cell(pv_row, col2).value == naam
                ):
                    al += 1
    
        return al < nodig
    
    
    def pp2_find_first_valid_long_block_in_step4(naam, ws_sheet, pv_rows, pauze_cols, open_spots_set):
        """
        Zoek in stap 4 een geldig halfuur voor een student.
    
        Voor minderjarige vroege stoppers:
        - kies het EERSTE geldige halfuur (dus zo vroeg mogelijk)
    
        Voor alle anderen:
        - behoud ook het eerste geldige halfuur
        """
        for idx in range(len(pauze_cols) - 1):
            col1 = pauze_cols[idx]
            col2 = pauze_cols[idx + 1]
    
            for _pv, pv_row in pv_rows:
                if (pv_row, col1) in open_spots_set or (pv_row, col2) in open_spots_set:
                    continue
    
                if ws_sheet.cell(pv_row, col1).value not in [None, ""]:
                    continue
                if ws_sheet.cell(pv_row, col2).value not in [None, ""]:
                    continue
    
                if not pp2_is_valid_long_break_for_student(naam, col1, col2, ws_sheet):
                    continue
    
                return pv_row, col1, col2
    
        return None
    
    # ---------------------------------------
    # 0) Eerst: minderjarigen die nog een LANGE pauze missen alsnog proberen plaatsen
    #    Dit vangt het geval op waarin een minderjarige laat start
    #    en stap 2 geen geldig halfuur vond.
    # ---------------------------------------
    pp2_step4_late_long_break_rescue = []
    pp2_regular_short_breaks_placed = []
    
    # ---------------------------------------
    # 0A) Eerst: minderjarige vroege stoppers
    #     die nog een LANGE pauze missen
    #     => zo vroeg mogelijk (links naar rechts)
    # ---------------------------------------
    pp2_minor_early_stoppers = [
        s["naam"] for s in studenten
        if (
            pp2_is_minderjarig(s["naam"])
            and len(pp2_get_student_work_hours(s["naam"])) >= 4
            and pp2_get_student_work_hours(s["naam"])
            and max(pp2_get_student_work_hours(s["naam"])) <= 15
        )
    ]
    
    for naam in pp2_minor_early_stoppers:
        if not pp2_student_heeft_nog_lange_pauze_nodig(
            naam=naam,
            ws_sheet=ws_pp2,
            pv_rows=pv_rows_pp2,
            pauze_cols=pauze_cols_pp2
        ):
            continue
    
        gevonden = None
    
        for idx in range(len(pauze_cols_pp2) - 1):
            col1 = pauze_cols_pp2[idx]
            col2 = pauze_cols_pp2[idx + 1]
    
            for _pv, pv_row in pv_rows_pp2:
                if (pv_row, col1) in pp2_open_spots or (pv_row, col2) in pp2_open_spots:
                    continue
    
                if not pp2_is_beschikbaar(ws_pp2, pv_row, col1):
                    continue
                if not pp2_is_beschikbaar(ws_pp2, pv_row, col2):
                    continue
    
                if not pp2_is_valid_long_break_for_student(naam, col1, col2, ws_pp2):
                    continue
    
                gevonden = (pv_row, col1, col2)
                break
    
            if gevonden is not None:
                break
    
        if gevonden is None:
            continue
    
        pv_row, col1, col2 = gevonden
    
        pp2_write_long_break(
            ws_sheet=ws_pp2,
            pv_row=pv_row,
            col1=col1,
            col2=col2,
            naam=naam,
            leave_top_blank=False
        )
    
        pp2_lange_pauze_ontvangers.add(naam)
    
        pp2_step4_late_long_break_rescue.append({
            "naam": naam,
            "tijden": [ws_pp2.cell(1, col1).value, ws_pp2.cell(1, col2).value]
        })
    
    
    # ---------------------------------------
    # 0B) Daarna: exact diezelfde minderjarige
    #     vroege stoppers hun KORTE pauze
    #     => zo laat mogelijk (rechts naar links)
    #     => bij voorkeur op dezelfde rij als het laatste halfuur
    # ---------------------------------------
    for naam in pp2_minor_early_stoppers:
        resterend = pp2_resterende_korte_kwartieren(
            naam=naam,
            ws_sheet=ws_pp2,
            pv_rows=pv_rows_pp2,
            pauze_cols=pauze_cols_pp2,
            lange_pauze_ontvangers=pp2_lange_pauze_ontvangers
        )
    
        if resterend <= 0:
            continue
    
        # Zoek de rij én eindkolom van het LAATSTE halfuur van deze student
        laatste_lange_eindcol = None
        laatste_lange_rij = None
    
        for _pv, pv_row in pv_rows_pp2:
            for idx in range(len(pauze_cols_pp2) - 1):
                col1 = pauze_cols_pp2[idx]
                col2 = pauze_cols_pp2[idx + 1]
    
                if (
                    ws_pp2.cell(pv_row, col1).value == naam
                    and ws_pp2.cell(pv_row, col2).value == naam
                ):
                    if laatste_lange_eindcol is None or col2 > laatste_lange_eindcol:
                        laatste_lange_eindcol = col2
                        laatste_lange_rij = pv_row
    
        if laatste_lange_eindcol is None:
            continue
    
        gekozen = None
    
        # Bouw PV-rij volgorde: eerst de rij van het laatste halfuur, dan de rest
        pv_volgorde = (
            [(pv, pv_row) for pv, pv_row in pv_rows_pp2 if pv_row == laatste_lange_rij]
            + [(pv, pv_row) for pv, pv_row in pv_rows_pp2 if pv_row != laatste_lange_rij]
        )
    
        for col in reversed(pauze_cols_pp2):
            if col <= laatste_lange_eindcol:
                continue
    
            for pv, pv_row in pv_volgorde:
                if (pv_row, col) in pp2_open_spots:
                    continue
    
                if not pp2_is_beschikbaar(ws_pp2, pv_row, col):
                    continue
    
                if not pp2_is_valid_short_break_for_student(naam, col, ws_pp2):
                    continue
    
                gekozen = (pv, pv_row, col)
                break
    
            if gekozen is not None:
                break
    
        if gekozen is None:
            continue
    
        pv, pv_row, col = gekozen
    
        pp2_place_short_break_cols_on_row(
            naam=naam,
            pv=pv,
            pv_row=pv_row,
            cols=[col]
        )
    
    
    # ---------------------------------------
    # 1) Daarna pas: gewone korte werkers
    #    die vroeger stoppen dan het einde
    #    van de dag, maar GEEN minderjarige
    #    vroege stoppers zijn
    # ---------------------------------------
    pp2_students_before_end_all = pp2_get_students_stopping_before_end()
    
    pp2_students_before_end_pending = [
        naam for naam in pp2_students_before_end_all
        if naam not in pp2_minor_early_stoppers
        and pp2_resterende_korte_kwartieren(
            naam=naam,
            ws_sheet=ws_pp2,
            pv_rows=pv_rows_pp2,
            pauze_cols=pauze_cols_pp2,
            lange_pauze_ontvangers=pp2_lange_pauze_ontvangers
        ) > 0
    ]
    
    def pp2_get_last_long_break_end_col_for_sort(naam):
        """
        Geeft de eindkolom van het LAATSTE halfuur van deze student terug,
        over alle PV-rijen heen. Studenten zonder lange pauze krijgen -1,
        zodat ze vooraan komen in de sortering.
        """
        eindcol = -1
        for _pv, pv_row in pv_rows_pp2:
            for idx in range(len(pauze_cols_pp2) - 1):
                col1 = pauze_cols_pp2[idx]
                col2 = pauze_cols_pp2[idx + 1]
                if (
                    ws_pp2.cell(pv_row, col1).value == naam
                    and ws_pp2.cell(pv_row, col2).value == naam
                ):
                    if col2 > eindcol:
                        eindcol = col2
        return eindcol
    
    pp2_students_before_end_pending.sort(
        key=lambda naam: pp2_get_last_long_break_end_col_for_sort(naam)
    )
    
    pp2_regular_short_breaks_placed = []
    
    
    
    for col in pauze_cols_pp2:
        if not pp2_students_before_end_pending:
            break
    
        for pv, pv_row in pv_rows_pp2:
            if not pp2_students_before_end_pending:
                break
    
            if (pv_row, col) in pp2_open_spots:
                continue
    
            if not pp2_is_beschikbaar(ws_pp2, pv_row, col):
                continue
    
            toegewezen_naam = None
            toegewezen_cols = []
    
            # ---------------------------------------------------
            # PRIORITEIT 1:
            # studenten die op deze rij al een lange pauze hebben
            # Voor minderjarige lange werkers (>6u): zoek na het
            # LAATSTE halfuur op DEZE rij (niet over alle rijen).
            # ---------------------------------------------------
            rij_lange_pauze_namen = pp2_get_long_break_owners_on_row(
                ws_pp2,
                pv_row,
                pauze_cols_pp2
            )
    
            for kandidaat in rij_lange_pauze_namen:
                if kandidaat not in pp2_students_before_end_pending:
                    continue
    
                if not pp2_student_has_long_break_in_row(
                    kandidaat,
                    ws_pp2,
                    pv_row,
                    pauze_cols_pp2
                ):
                    continue
    
                is_minor_long_worker = (
                    pp2_is_minderjarig(kandidaat)
                    and student_totalen.get(kandidaat, 0) > 6
                )
    
                if is_minor_long_worker:
                    # Zoek eindkolom van het LAATSTE halfuur op DEZE specifieke rij
                    ankercol = None
                    for idx in range(len(pauze_cols_pp2) - 1):
                        col1 = pauze_cols_pp2[idx]
                        col2 = pauze_cols_pp2[idx + 1]
                        if (
                            ws_pp2.cell(pv_row, col1).value == kandidaat
                            and ws_pp2.cell(pv_row, col2).value == kandidaat
                        ):
                            ankercol = col2
                else:
                    ankercol = None
    
                cols = pp2_find_needed_short_cols_for_student_on_row(
                    naam=kandidaat,
                    pv_row=pv_row,
                    ws_sheet=ws_pp2,
                    pauze_cols=pauze_cols_pp2,
                    open_spots_set=pp2_open_spots,
                    min_col_exclusive=ankercol,
                    zoek_zo_laat_mogelijk=is_minor_long_worker
                )
    
                if not cols:
                    continue
    
                toegewezen_naam = kandidaat
                toegewezen_cols = cols
                break
    
            # ---------------------------------------------------
            # PRIORITEIT 2:
            # anders eerste geldige kandidaat uit de vaste lijst
            # ---------------------------------------------------
            if toegewezen_naam is None:
                for kandidaat in pp2_students_before_end_pending:
                    is_minor_lw_p2 = (
                        pp2_is_minderjarig(kandidaat)
                        and student_totalen.get(kandidaat, 0) > 6
                    )

                    ankercol_p2 = None
                    if is_minor_lw_p2:
                        for _pv2, pv_row2 in pv_rows_pp2:
                            for idx in range(len(pauze_cols_pp2) - 1):
                                col1 = pauze_cols_pp2[idx]
                                col2 = pauze_cols_pp2[idx + 1]
                                if (
                                    ws_pp2.cell(pv_row2, col1).value == kandidaat
                                    and ws_pp2.cell(pv_row2, col2).value == kandidaat
                                ):
                                    if ankercol_p2 is None or col2 > ankercol_p2:
                                        ankercol_p2 = col2

                    cols = pp2_find_needed_short_cols_for_student_on_row(
                        naam=kandidaat,
                        pv_row=pv_row,
                        ws_sheet=ws_pp2,
                        pauze_cols=pauze_cols_pp2,
                        open_spots_set=pp2_open_spots,
                        min_col_exclusive=ankercol_p2,
                        zoek_zo_laat_mogelijk=is_minor_lw_p2
                    )

                    if not cols:
                        continue

                    toegewezen_naam = kandidaat
                    toegewezen_cols = cols
                    break
    
            if toegewezen_naam and toegewezen_cols:
                pp2_place_short_break_cols_on_row(
                    naam=toegewezen_naam,
                    pv=pv,
                    pv_row=pv_row,
                    cols=toegewezen_cols
                )
    
                if toegewezen_naam in pp2_students_before_end_pending:
                    pp2_students_before_end_pending.remove(toegewezen_naam)
    
    # ---------------------------------------
    # 2) Daarna: lange pauzevlinders zelf
    #    - alleen die nog korte kwartieren nodig hebben
    #    - alleen in eigen rij
    #    - na korte werkers
    #    - voor andere lange werkers
    # ---------------------------------------
    pp2_lange_pv_short_breaks_placed = []
    
    for pv, pv_row in pv_rows_pp2:
        naam = pv["naam"]
    
        if not pp2_student_is_long_worker(naam):
            continue
    
        resterend = pp2_resterende_korte_kwartieren(
            naam=naam,
            ws_sheet=ws_pp2,
            pv_rows=pv_rows_pp2,
            pauze_cols=pauze_cols_pp2,
            lange_pauze_ontvangers=pp2_lange_pauze_ontvangers
        )
    
        if resterend <= 0:
            continue
    
        gekozen_cols = pp2_find_needed_short_cols_for_student_on_row(
            naam=naam,
            pv_row=pv_row,
            ws_sheet=ws_pp2,
            pauze_cols=pauze_cols_pp2,
            open_spots_set=pp2_open_spots
        )
    
        if not gekozen_cols:
            continue
    
        for col in gekozen_cols:
            pp2_write_short_break_for_pv(
                ws_sheet=ws_pp2,
                pv_row=pv_row,
                col=col,
                naam=naam
            )
    
        pp2_lange_pv_short_breaks_placed.append({
            "naam": naam,
            "tijden": [ws_pp2.cell(1, col).value for col in gekozen_cols]
        })
    
    
    # STAP 5 55555555555555555555555555555555555555555555555555555555555555555555555555555555555555555555555555555555
    
    
    # -----------------------------------
    # STAP 5 PP optie 2:
    # laatste resterende korte kwartieren invullen
    # - werkt met resterende kwartieren i.p.v. ja/nee
    # - minderjarigen >4u krijgen hier ook 2 opeenvolgende kwartieren
    # - eerst overige pending korte kwartieren
    # - pas daarna eindwerkers zonder lange pauze
    # -----------------------------------
    
    def pp2_get_long_break_students_on_row_in_order(ws_sheet, pv_row, pauze_cols):
        """
        Geef de studenten terug die op deze rij een lange pauze hebben,
        in dezelfde volgorde als de lange pauzes op de rij zelf:
        dus van links naar rechts.
    
        Dit zorgt ervoor dat korte pauzes later ook in een logische
        volgorde kunnen volgen, gelijklopend met de lange pauzes.
        """
        found = {}
    
        for idx in range(len(pauze_cols) - 1):
            col1 = pauze_cols[idx]
            col2 = pauze_cols[idx + 1]
    
            val1 = ws_sheet.cell(pv_row, col1).value
            val2 = ws_sheet.cell(pv_row, col2).value
    
            if val1 and val1 == val2:
                naam = str(val1).strip()
                # bewaar de startkolom van de eerste lange pauze op deze rij
                if naam not in found:
                    found[naam] = col1
    
        return [naam for naam, _col in sorted(found.items(), key=lambda x: x[1])]
    
    
    def pp2_student_has_long_break_in_row(naam, ws_sheet, pv_row, pauze_cols):
        """
        True als deze student op deze specifieke rij ergens een lange pauze heeft.
        """
        for idx in range(len(pauze_cols) - 1):
            col1 = pauze_cols[idx]
            col2 = pauze_cols[idx + 1]
    
            if (
                ws_sheet.cell(pv_row, col1).value == naam and
                ws_sheet.cell(pv_row, col2).value == naam
            ):
                return True
    
        return False
    
    
    def pp2_student_works_until_day_end(naam):
        """
        True als student werkt tot het einduur van de dag.
        """
        werk_uren = pp2_get_student_work_hours(naam)
        if not werk_uren or not open_uren:
            return False
        return max(werk_uren) == max(open_uren)
    
    
    def pp2_build_step5_pending_groups():
        """
        Splits alle NIET-pauzevlinders die nog korte kwartieren nodig hebben in:
        A) overige pending korte kwartieren
        B) eindwerkers zonder lange pauze
    
        Minderjarige vroege stoppers horen hier ook NIET meer in:
        die werden al eerder apart behandeld en mogen in stap 5
        niet opnieuw een kort kwartier krijgen.
    
        Pauzevlinders zelf horen hier ook niet meer in:
        - korte pauzevlinders werden al eerder verwerkt
        - lange pauzevlinders kregen in stap 4 hun eigen aparte fase,
          enkel in hun eigen rij
        """
        pauzevlinder_namen_set = {pv["naam"] for pv in selected}
        all_pending = []
    
        for s in studenten:
            naam = s["naam"]
    
            # Pauzevlinders hier NIET meer meenemen
            if naam in pauzevlinder_namen_set:
                continue
    
            # Minderjarige vroege stoppers hier ook NIET meer meenemen
            if (
                pp2_is_minderjarig(naam)
                and len(pp2_get_student_work_hours(naam)) >= 4
                and pp2_get_student_work_hours(naam)
                and max(pp2_get_student_work_hours(naam)) <= 15
            ):
                continue
    
            resterend = pp2_resterende_korte_kwartieren(
                naam=naam,
                ws_sheet=ws_pp2,
                pv_rows=pv_rows_pp2,
                pauze_cols=pauze_cols_pp2,
                lange_pauze_ontvangers=pp2_lange_pauze_ontvangers
            )
    
            if resterend > 0:
                all_pending.append(naam)
    
        endworkers_without_long_break = []
        other_pending_short_breaks = []
    
        for naam in all_pending:
            heeft_lange = (naam in pp2_lange_pauze_ontvangers)
            werkt_tot_einduur = pp2_student_works_until_day_end(naam)
    
            if werkt_tot_einduur and not heeft_lange:
                endworkers_without_long_break.append(naam)
            else:
                other_pending_short_breaks.append(naam)
    
        random.shuffle(other_pending_short_breaks)
        random.shuffle(endworkers_without_long_break)
    
        return other_pending_short_breaks, endworkers_without_long_break
    
    
    def pp2_try_assign_from_candidate_list_on_row(candidate_list, pv, pv_row, shuffle_candidates=False):
        """
        Probeer op deze rij een kandidaat te plaatsen uit de opgegeven lijst.
        Werkt met 1 of 2 kwartieren, afhankelijk van wat nog nodig is.
    
        Belangrijk:
        - de volgorde van candidate_list blijft behouden als dat een prioriteitslijst is
          (bv. dezelfde volgorde als de lange pauzes op de rij)
        - voor gewone fallback-lijsten kan shuffle_candidates=True gebruikt worden
          zodat niet-minderjarigen daar opnieuw randomer verdeeld worden
        """
        kandidaten = candidate_list[:]
    
        if shuffle_candidates and len(kandidaten) > 1:
            random.shuffle(kandidaten)
    
        for kandidaat in kandidaten:
            cols = pp2_find_needed_short_cols_for_student_on_row(
                naam=kandidaat,
                pv_row=pv_row,
                ws_sheet=ws_pp2,
                pauze_cols=pauze_cols_pp2,
                open_spots_set=pp2_open_spots
            )
    
            if not cols:
                continue
    
            pp2_place_short_break_cols_on_row(
                naam=kandidaat,
                pv=pv,
                pv_row=pv_row,
                cols=cols
            )
    
            return kandidaat, cols
    
        return None, []
    
    
    pp2_other_pending_short_breaks, pp2_endworkers_without_long_break = pp2_build_step5_pending_groups()
    
    pp2_step5_short_breaks_placed = []
    
    
    
    
    # -----------------------------------
    # 1B) Speciale behandeling:
    #     minderjarigen die <= 6u werken én pas beginnen na 13u
    #     => korte pauze zo laat mogelijk, bij voorkeur op rij van lange pauze
    # -----------------------------------
    pp2_late_start_minors_handled = set()
    
    for naam in list(pp2_other_pending_short_breaks):
        werk_uren = pp2_get_student_work_hours(naam)
        if not werk_uren:
            continue
    
        is_minor = pp2_is_minderjarig(naam)
        werkt_kort = student_totalen.get(naam, 0) <= 6
        begint_laat = min(werk_uren) > 13
    
        if not (is_minor and werkt_kort and begint_laat):
            continue
    
        # Zoek rij van de lange pauze van deze student
        lange_pauze_rij = None
        for _pv, pv_row in pv_rows_pp2:
            for idx in range(len(pauze_cols_pp2) - 1):
                col1 = pauze_cols_pp2[idx]
                col2 = pauze_cols_pp2[idx + 1]
                if (
                    ws_pp2.cell(pv_row, col1).value == naam
                    and ws_pp2.cell(pv_row, col2).value == naam
                ):
                    lange_pauze_rij = pv_row
                    break
            if lange_pauze_rij is not None:
                break
    
        # PV-rij volgorde: eerst rij van lange pauze, dan de rest
        pv_volgorde = (
            [(pv, pv_row) for pv, pv_row in pv_rows_pp2 if pv_row == lange_pauze_rij]
            + [(pv, pv_row) for pv, pv_row in pv_rows_pp2 if pv_row != lange_pauze_rij]
        )
    
        geplaatst = False
        for col in reversed(pauze_cols_pp2):
            for pv, pv_row in pv_volgorde:
                if (pv_row, col) in pp2_open_spots:
                    continue
                if not pp2_is_beschikbaar(ws_pp2, pv_row, col):
                    continue
                if not pp2_is_valid_short_break_for_student(naam, col, ws_pp2):
                    continue
    
                pp2_place_short_break_cols_on_row(
                    naam=naam,
                    pv=pv,
                    pv_row=pv_row,
                    cols=[col]
                )
                pp2_late_start_minors_handled.add(naam)
                geplaatst = True
                break
            if geplaatst:
                break
    
    # Verwijder deze studenten uit de gewone pending lijst
    pp2_other_pending_short_breaks = [
        naam for naam in pp2_other_pending_short_breaks
        if naam not in pp2_late_start_minors_handled
    ]

    # -----------------------------------             
    # 1C) Speciale behandeling:
    #     minderjarigen >6u die tot het einduur werken
    #     => korte pauze zo laat mogelijk, na het tweede halfuur
    # -----------------------------------
    pp2_minor_long_end_handled = set()

    for naam in list(pp2_other_pending_short_breaks):
        if not pp2_is_minderjarig(naam):
            continue
        if student_totalen.get(naam, 0) <= 6:
            continue
        if not pp2_student_works_until_day_end(naam):
            continue

        # Zoek ankercol = einde van het LAATSTE halfuur over alle rijen
        ankercol = None
        laatste_lange_rij = None
        for _pv, pv_row in pv_rows_pp2:
            for idx in range(len(pauze_cols_pp2) - 1):
                col1 = pauze_cols_pp2[idx]
                col2 = pauze_cols_pp2[idx + 1]
                if (
                    ws_pp2.cell(pv_row, col1).value == naam
                    and ws_pp2.cell(pv_row, col2).value == naam
                ):
                    if ankercol is None or col2 > ankercol:
                        ankercol = col2
                        laatste_lange_rij = pv_row

        # PV-rij volgorde: eerst rij van laatste halfuur, dan de rest
        pv_volgorde = (
            [(pv, pv_row) for pv, pv_row in pv_rows_pp2 if pv_row == laatste_lange_rij]
            + [(pv, pv_row) for pv, pv_row in pv_rows_pp2 if pv_row != laatste_lange_rij]
        )

        geplaatst = False
        for col in reversed(pauze_cols_pp2):
            if ankercol is not None and col <= ankercol:
                continue
            for pv, pv_row in pv_volgorde:
                if (pv_row, col) in pp2_open_spots:
                    continue
                if not pp2_is_beschikbaar(ws_pp2, pv_row, col):
                    continue
                if not pp2_is_valid_short_break_for_student(naam, col, ws_pp2):
                    continue

                pp2_place_short_break_cols_on_row(naam=naam, pv=pv, pv_row=pv_row, cols=[col])
                pp2_minor_long_end_handled.add(naam)
                geplaatst = True
                break
            if geplaatst:
                break

    # Verwijder uit de algemene pending lijst
    pp2_other_pending_short_breaks = [
        naam for naam in pp2_other_pending_short_breaks
        if naam not in pp2_minor_long_end_handled
    ]

  
    # -----------------------------------
    # 2) Eerst alle "gewone" resterende korte kwartieren invullen
    #    inclusief:
    #    - studenten met lange pauze die nog korte kwartieren missen
    #    - minderjarigen met dubbele korte kwartieren
    # -----------------------------------
    for col in pauze_cols_pp2:
        if not pp2_other_pending_short_breaks:
            break
    
        for pv, pv_row in pv_rows_pp2:
            if not pp2_other_pending_short_breaks:
                break
    
            # open spots overslaan
            if (pv_row, col) in pp2_open_spots:
                continue
    
            # vak moet leeg zijn
            if not pp2_is_beschikbaar(ws_pp2, pv_row, col):
                continue
    
            toegewezen_naam = None
            toegewezen_cols = []
    
            # -----------------------------------
            # PRIORITEIT 1:
            # studenten die in deze rij al een lange pauze kregen,
            # in dezelfde volgorde als hun lange pauzes op die rij
            # -----------------------------------
            rij_lange_pauze_namen = pp2_get_long_break_students_on_row_in_order(
                ws_sheet=ws_pp2,
                pv_row=pv_row,
                pauze_cols=pauze_cols_pp2
            )
    
            prioriteitslijst = [
                naam for naam in rij_lange_pauze_namen
                if naam in pp2_other_pending_short_breaks
            ]
    
            toegewezen_naam, toegewezen_cols = pp2_try_assign_from_candidate_list_on_row(
                candidate_list=prioriteitslijst,
                pv=pv,
                pv_row=pv_row,
                shuffle_candidates=False
            )
    
            # -----------------------------------
            # PRIORITEIT 2:
            # fallback naar overige nog open korte kwartieren
            # hier mag het randomer blijven voor niet-minderjarigen
            # -----------------------------------
            if toegewezen_naam is None:
                fallback_lijst = [
                    naam for naam in pp2_other_pending_short_breaks
                    if naam not in prioriteitslijst
                ]
    
                toegewezen_naam, toegewezen_cols = pp2_try_assign_from_candidate_list_on_row(
                    candidate_list=fallback_lijst,
                    pv=pv,
                    pv_row=pv_row,
                    shuffle_candidates=True
                )
    
            # schrijven/loggen indien kandidaat gevonden
            if toegewezen_naam:
                pp2_step5_short_breaks_placed.append({
                    "naam": toegewezen_naam,
                    "pauzevlinder": pv["naam"],
                    "tijden": [ws_pp2.cell(1, c).value for c in toegewezen_cols],
                    "via_lange_pauze_prioriteit": toegewezen_naam in rij_lange_pauze_namen
                })
    
                if pp2_resterende_korte_kwartieren(
                    naam=toegewezen_naam,
                    ws_sheet=ws_pp2,
                    pv_rows=pv_rows_pp2,
                    pauze_cols=pauze_cols_pp2,
                    lange_pauze_ontvangers=pp2_lange_pauze_ontvangers
                ) <= 0:
                    if toegewezen_naam in pp2_other_pending_short_breaks:
                        pp2_other_pending_short_breaks.remove(toegewezen_naam)
    
    
    # -----------------------------------
    # 3) Pas daarna:
    #    studenten die tot het einduur werken én geen lange pauze kregen
    # -----------------------------------
    for col in pauze_cols_pp2:
        if not pp2_endworkers_without_long_break:
            break
    
        for pv, pv_row in pv_rows_pp2:
            if not pp2_endworkers_without_long_break:
                break
    
            # open spots overslaan
            if (pv_row, col) in pp2_open_spots:
                continue
    
            # vak moet leeg zijn
            if not pp2_is_beschikbaar(ws_pp2, pv_row, col):
                continue
    
            toegewezen_naam = None
            toegewezen_cols = []
    
            rij_lange_pauze_namen = pp2_get_long_break_students_on_row_in_order(
                ws_sheet=ws_pp2,
                pv_row=pv_row,
                pauze_cols=pauze_cols_pp2
            )
    
            prioriteitslijst = [
                naam for naam in rij_lange_pauze_namen
                if naam in pp2_endworkers_without_long_break
            ]
    
            toegewezen_naam, toegewezen_cols = pp2_try_assign_from_candidate_list_on_row(
                candidate_list=prioriteitslijst,
                pv=pv,
                pv_row=pv_row,
                shuffle_candidates=False
            )
    
            if toegewezen_naam is None:
                fallback_lijst = [
                    naam for naam in pp2_endworkers_without_long_break
                    if naam not in prioriteitslijst
                ]
    
                toegewezen_naam, toegewezen_cols = pp2_try_assign_from_candidate_list_on_row(
                    candidate_list=fallback_lijst,
                    pv=pv,
                    pv_row=pv_row,
                    shuffle_candidates=True
                )
    
            if toegewezen_naam:
                pp2_step5_short_breaks_placed.append({
                    "naam": toegewezen_naam,
                    "pauzevlinder": pv["naam"],
                    "tijden": [ws_pp2.cell(1, c).value for c in toegewezen_cols],
                    "via_lange_pauze_prioriteit": toegewezen_naam in rij_lange_pauze_namen
                })
    
                if pp2_resterende_korte_kwartieren(
                    naam=toegewezen_naam,
                    ws_sheet=ws_pp2,
                    pv_rows=pv_rows_pp2,
                    pauze_cols=pauze_cols_pp2,
                    lange_pauze_ontvangers=pp2_lange_pauze_ontvangers
                ) <= 0:
                    if toegewezen_naam in pp2_endworkers_without_long_break:
                        pp2_endworkers_without_long_break.remove(toegewezen_naam)
    
    # -----------------------------------
    # 3) Pas daarna:
    #    studenten die tot het einduur werken én geen lange pauze kregen
    # -----------------------------------
    for col in pauze_cols_pp2:
        if not pp2_endworkers_without_long_break:
            break
    
        for pv, pv_row in pv_rows_pp2:
            if not pp2_endworkers_without_long_break:
                break
    
            # open spots overslaan
            if (pv_row, col) in pp2_open_spots:
                continue
    
            # vak moet leeg zijn
            if not pp2_is_beschikbaar(ws_pp2, pv_row, col):
                continue
    
            toegewezen_naam = None
            toegewezen_cols = []
    
            rij_lange_pauze_namen = pp2_get_long_break_students_on_row_in_order(
                ws_sheet=ws_pp2,
                pv_row=pv_row,
                pauze_cols=pauze_cols_pp2
            )
    
            prioriteitslijst = [
                naam for naam in rij_lange_pauze_namen
                if naam in pp2_endworkers_without_long_break
            ]
    
            toegewezen_naam, toegewezen_cols = pp2_try_assign_from_candidate_list_on_row(
                candidate_list=prioriteitslijst,
                pv=pv,
                pv_row=pv_row,
                shuffle_candidates=False
            )
    
            if toegewezen_naam is None:
                fallback_lijst = [
                    naam for naam in pp2_endworkers_without_long_break
                    if naam not in prioriteitslijst
                ]
    
                toegewezen_naam, toegewezen_cols = pp2_try_assign_from_candidate_list_on_row(
                    candidate_list=fallback_lijst,
                    pv=pv,
                    pv_row=pv_row,
                    shuffle_candidates=True
                )
    
            if toegewezen_naam:
                pp2_step5_short_breaks_placed.append({
                    "naam": toegewezen_naam,
                    "pauzevlinder": pv["naam"],
                    "tijden": [ws_pp2.cell(1, c).value for c in toegewezen_cols],
                    "via_lange_pauze_prioriteit": toegewezen_naam in rij_lange_pauze_namen
                })
    
                if pp2_resterende_korte_kwartieren(
                    naam=toegewezen_naam,
                    ws_sheet=ws_pp2,
                    pv_rows=pv_rows_pp2,
                    pauze_cols=pauze_cols_pp2,
                    lange_pauze_ontvangers=pp2_lange_pauze_ontvangers
                ) <= 0:
                    if toegewezen_naam in pp2_endworkers_without_long_break:
                        pp2_endworkers_without_long_break.remove(toegewezen_naam)
    
    
    
    
    #FEEDBACKKKKKKKKKKKKKKKKKKKKKKKKKKKKKKKKKKKKKKKKKKKKKKKKKKKKKKKKKKKKKKKKKKKKKKKKKKKKKKKKKKKKKKKKKKKKKKKKKKKKKKKKK
    # =============================
    # FEEDBACK SHEET - OPTIE 2
    # =============================
    ws_feedback2 = wb_arg.create_sheet("Feedback PP")
    
    groen_fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
    rood_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    
    row_fb2 = 1
    
    ws_feedback2.cell(row_fb2, 1, "Feedback pauzeplanning").font = Font(bold=True)
    row_fb2 += 2
    
    # -----------------------------------
    # 1) Lange pauzes controleren
    # Nieuwe logica PP optie 2:
    # - alleen studenten met >6 uur werk moeten een lange pauze hebben
    # -----------------------------------
    pp2_lange_pauze_ontbreekt = []
    
    for s in studenten:
        naam = s["naam"]
        gewerkte_uren = student_totalen.get(naam, 0)
    
        if gewerkte_uren > 6:
            if not pp2_heeft_al_lange_pauze(naam, ws_pp2, pv_rows_pp2, pauze_cols_pp2):
                pp2_lange_pauze_ontbreekt.append(naam)
    
    if not pp2_lange_pauze_ontbreekt:
        cel = ws_feedback2.cell(row_fb2, 1, "✓ Alle lange pauzes toegekend")
        cel.fill = groen_fill
        cel.font = Font(bold=True, color="006100")
        row_fb2 += 2
    else:
        cel = ws_feedback2.cell(row_fb2, 1, "✗ Ontbrekende lange pauzes:")
        cel.fill = rood_fill
        cel.font = Font(bold=True)
        row_fb2 += 1
    
        for naam in sorted(pp2_lange_pauze_ontbreekt):
            ws_feedback2.cell(row_fb2, 1, naam)
            row_fb2 += 1
    
        row_fb2 += 1
    
    # -----------------------------------
    # 2) Korte kwartieren controleren
    # Gebruik exact dezelfde logica als de planner zelf:
    # - pp2_benodigde_korte_kwartieren(...)
    # - pp2_resterende_korte_kwartieren(...)
    # Dus geen aparte feedbacktelling meer
    # -----------------------------------
    pp2_korte_kwartieren_ontbreekt = []
    
    for s in studenten:
        naam = s["naam"]
    
        nodig = pp2_benodigde_korte_kwartieren(naam)
        if nodig <= 0:
            continue
    
        resterend = pp2_resterende_korte_kwartieren(
            naam=naam,
            ws_sheet=ws_pp2,
            pv_rows=pv_rows_pp2,
            pauze_cols=pauze_cols_pp2,
            lange_pauze_ontvangers=pp2_lange_pauze_ontvangers
        )
    
        if resterend > 0:
            pp2_korte_kwartieren_ontbreekt.append((naam, resterend))
    
    if not pp2_korte_kwartieren_ontbreekt:
        cel = ws_feedback2.cell(row_fb2, 1, "✓ Alle korte kwartieren toegekend")
        cel.fill = groen_fill
        cel.font = Font(bold=True, color="006100")
        row_fb2 += 2
    else:
        cel = ws_feedback2.cell(row_fb2, 1, "✗ Ontbrekende korte kwartieren:")
        cel.fill = rood_fill
        cel.font = Font(bold=True)
        row_fb2 += 1
    
        for naam, resterend in sorted(pp2_korte_kwartieren_ontbreekt, key=lambda x: x[0].lower()):
            if resterend == 1:
                ws_feedback2.cell(row_fb2, 1, f"{naam} - nog 1 kwartier tekort")
            else:
                ws_feedback2.cell(row_fb2, 1, f"{naam} - nog {resterend} kwartieren tekort")
            row_fb2 += 1
    
        row_fb2 += 1
    
    # -----------------------------------
    # kolombreedte en opmaak
    # -----------------------------------
    ws_feedback2.column_dimensions["A"].width = 45
    
    for row in ws_feedback2.iter_rows():
        for cell in row:
            cell.alignment = Alignment(horizontal="left", vertical="center")
            cell.border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin")
            )

# Globals herstellen
    ws_planning     = _ws_planning_bak
    student_totalen = _student_totalen_bak


# ── Oorspronkelijke aanroep ──
maak_pp2_sheets(wb_out, assigned_map)


# PART 6 6666666666666666666666666666666666666666666666666666666666666666666666666666666666666666666666666666666666666666
# PART 6 666666666666666666666666666666666666666666666666666666666666666666666666666666666666666666666666666666666666666

# -----------------------------
# DEEL 6: Wissels detecteren, classificeren en exporteren
# -----------------------------

from collections import defaultdict, deque
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

# -----------------------------
# Helpers
# -----------------------------
def build_student_per_hour_map(assigned_map):
    student_per_uur = defaultdict(dict)
    for (uur, attr), namen in assigned_map.items():
        for naam in namen:
            student_per_uur[naam][uur] = attr
    return student_per_uur


def extract_hourly_changes(student_per_uur, open_uren):
    """
    Bouw per uur alle veranderingen op:
    - newcomers: studenten die op dit uur starten
    - movers: studenten die op dit uur van attractie wisselen
    - leavers: studenten die vorig uur wel werkten en nu niet meer
    - disappearing_sources: attractieplekken die verdwijnen tussen vorig uur en dit uur
    """
    changes_per_hour = {}

    def capaciteit_op_uur(uur, attr):
        if uur not in open_uren:
            return 0
        return max(0, aantallen.get(uur, {}).get(attr, 0))

    all_attrs = set()
    for uur2 in open_uren:
        all_attrs.update(aantallen.get(uur2, {}).keys())

    for uur in sorted(open_uren):
        # VERVANGEN DOOR:
        _sorted_open = sorted(open_uren)
        _idx = _sorted_open.index(uur)
        prev_uur = _sorted_open[_idx - 1] if _idx > 0 else None

        prev_students = {}
        curr_students = {}

        for naam, uren_dict in student_per_uur.items():
            if prev_uur in uren_dict:
                prev_students[naam] = uren_dict[prev_uur]
            if uur in uren_dict:
                curr_students[naam] = uren_dict[uur]

        newcomers = []
        movers = []
        leavers = []

        for naam, curr_attr in curr_students.items():
            if naam not in prev_students:
                newcomers.append({
                    "naam": naam,
                    "naar": curr_attr
                })
            else:
                prev_attr = prev_students[naam]
                if prev_attr != curr_attr:
                    movers.append({
                        "naam": naam,
                        "van": prev_attr,
                        "naar": curr_attr,
                        "uur": uur,
                        "type": "normaal"
                    })

        for naam, prev_attr in prev_students.items():
            if naam not in curr_students:
                leavers.append({
                    "naam": naam,
                    "van": prev_attr
                })

        disappearing_sources = []
        if prev_uur is not None and prev_uur in open_uren:
            for attr in sorted(all_attrs):
                prev_cap = capaciteit_op_uur(prev_uur, attr)
                curr_cap = capaciteit_op_uur(uur, attr)

                if curr_cap < prev_cap:
                    for pos in range(curr_cap + 1, prev_cap + 1):
                        disappearing_sources.append({
                            "attr": attr,
                            "pos": pos,
                            "reason": "capacity_drop"
                        })

        changes_per_hour[uur] = {
            "newcomers": newcomers,
            "movers": movers,
            "leavers": leavers,
            "disappearing_sources": disappearing_sources
        }

    return changes_per_hour




def classify_hourly_switches(uur, newcomers, movers, leavers=None, disappearing_sources=None):
    """
    Types:
    - volledig automatisch:
        een nieuwkomer komt toe op attractie A,
        daardoor kan een student van A weg,
        waardoor ketting verder loopt
    - half-automatisch:
        een ketting die start vanuit een verdwijnende plek
        of een logisch vervolg daarop is
    - normaal:
        losse wissels of resterende lussen zonder duidelijk startpunt

    Belangrijk:
    - de eerste edge van een ketting krijgt 'half-start'
    - de rest krijgt 'half-automatisch'
    - losse enkele wissels blijven 'normaal'

    Extra regels:
    - bij echte ronde lussen kiezen we het startpunt liefst op een attractie
      met 2 plekken op dit uur
    - niet-ronde kettingen komen vóór ronde lussen in de output
    - groene wissels starten altijd bij de attractie waar de nieuwkomer toekomt
    - enkel kettingen met lengte > 1 komen in de half-automatische output
      zodat er geen dubbels ontstaan
    """
    if not movers:
        return []

    if leavers is None:
        leavers = []

    if disappearing_sources is None:
        disappearing_sources = []

    # -----------------------------
    # Helpers
    # -----------------------------
    def stable_edge_key(edge):
        return (edge["van"], edge["naar"], edge["naam"])

    def next_edge_key(edge):
        return (edge["naar"], edge["naam"])

    def has_two_spots(attr):
        try:
            return aantallen[uur].get(attr, 1) >= 2
        except Exception:
            return False

    def roll_chain_from_start_edge(start_edge, edge_pool, used_ids):
        chain = []
        current = start_edge

        while current and current["id"] not in used_ids:
            chain.append(current)
            used_ids.add(current["id"])

            next_candidates = [
                e for e in edge_pool
                if e["id"] not in used_ids and e["van"] == current["naar"]
            ]
            next_candidates.sort(key=next_edge_key)
            current = next_candidates[0] if next_candidates else None

        return chain

    def classify_chain_shape(chain):
        """
        Geeft terug:
        - 'open' als begin en einde verschillen
        - 'cycle' als begin en einde terug sluiten
        """
        if len(chain) <= 1:
            return "single"

        eerste_van = chain[0]["van"]
        laatste_naar = chain[-1]["naar"]

        if eerste_van == laatste_naar:
            return "cycle"
        return "open"

    def add_chain_record_if_needed(chain_records, chain):
        """
        Enkel echte kettingen (lengte > 1) komen in chain_records.
        Singles blijven 'normaal' en worden later via normal_edges getoond.
        """
        if not chain:
            return

        shape = classify_chain_shape(chain)

        if len(chain) == 1:
            chain[0]["type"] = "normaal"
            return

        chain[0]["type"] = "half-start"
        for e in chain[1:]:
            e["type"] = "half-automatisch"

        chain_records.append({
            "shape": shape,
            "start_has_two_spots": has_two_spots(chain[0]["van"]),
            "edges": chain
        })

    # -----------------------------
    # Edges opbouwen
    # -----------------------------
    edges = []
    for idx, m in enumerate(movers):
        edges.append({
            "id": idx,
            "naam": m["naam"],
            "van": m["van"],
            "naar": m["naar"],
            "uur": uur,
            "type": "normaal"
        })

    # -----------------------------
    # Maps
    # -----------------------------
    outgoing = defaultdict(list)
    incoming = defaultdict(list)

    for e in edges:
        outgoing[e["van"]].append(e)
        incoming[e["naar"]].append(e)

    for attr in outgoing:
        outgoing[attr].sort(key=next_edge_key)
    for attr in incoming:
        incoming[attr].sort(key=lambda x: (x["van"], x["naam"]))

    # -----------------------------
    # 1. Volledig automatische kettingen
    # -----------------------------
    newcomers_by_attr = defaultdict(list)
    for n in newcomers:
        newcomers_by_attr[n["naar"]].append(n["naam"])

    auto_edge_ids = set()
    queue = deque()

    # Groen start ALTIJD bij de attractie waar de nieuwkomer toekomt
    # De nieuwkomer zet daar de ketting in gang.
    for attr in newcomers_by_attr.keys():
        for e in outgoing.get(attr, []):
            if e["id"] not in auto_edge_ids:
                auto_edge_ids.add(e["id"])
                queue.append(e)

    while queue:
        current = queue.popleft()
        next_attr = current["naar"]

        for next_edge in outgoing.get(next_attr, []):
            if next_edge["id"] not in auto_edge_ids:
                auto_edge_ids.add(next_edge["id"])
                queue.append(next_edge)

    for e in edges:
        if e["id"] in auto_edge_ids:
            e["type"] = "volledig automatisch"

    # -----------------------------
    # 2. Resterende edges
    # -----------------------------
    remaining_edges = [e for e in edges if e["id"] not in auto_edge_ids]

    if not remaining_edges:
        auto_edges = [e for e in edges if e["type"] == "volledig automatisch"]

        ordered_auto = []
        used_auto = set()

        # Volg exact de volgorde van newcomers, niet alfabetisch op attractie
        for newcomer in newcomers:
            start_attr = newcomer["naar"]

            start_candidates = [
                e for e in auto_edges
                if e["id"] not in used_auto and e["van"] == start_attr
            ]
            start_candidates.sort(key=next_edge_key)

            for start in start_candidates:
                current = start
                while current and current["id"] not in used_auto:
                    ordered_auto.append(current)
                    used_auto.add(current["id"])

                    next_candidates = [
                        e for e in auto_edges
                        if e["id"] not in used_auto and e["van"] == current["naar"]
                    ]
                    next_candidates.sort(key=next_edge_key)
                    current = next_candidates[0] if next_candidates else None

        leftovers_auto = [e for e in auto_edges if e["id"] not in used_auto]
        leftovers_auto.sort(key=stable_edge_key)
        ordered_auto.extend(leftovers_auto)

        return ordered_auto

    source_attrs = [x["attr"] for x in disappearing_sources]

    chain_records = []
    used_ids = set()

    # -----------------------------
    # 3. Eerst kettingen vanuit verdwijnende plekken
    # -----------------------------
    for start_attr in source_attrs:
        start_candidates = [
            e for e in remaining_edges
            if e["id"] not in used_ids and e["van"] == start_attr
        ]
        start_candidates.sort(key=stable_edge_key)

        for start_edge in start_candidates:
            if start_edge["id"] in used_ids:
                continue

            chain = roll_chain_from_start_edge(start_edge, remaining_edges, used_ids)
            add_chain_record_if_needed(chain_records, chain)

    # -----------------------------
    # 4. Restjes groeperen in componenten
    # -----------------------------
    leftovers = [e for e in remaining_edges if e["id"] not in used_ids]

    if leftovers:
        remaining_by_id = {e["id"]: e for e in leftovers}
        adjacency = defaultdict(set)

        for e1 in leftovers:
            for e2 in leftovers:
                if e1["id"] == e2["id"]:
                    continue
                if e1["naar"] == e2["van"] or e2["naar"] == e1["van"]:
                    adjacency[e1["id"]].add(e2["id"])
                    adjacency[e2["id"]].add(e1["id"])

        visited = set()
        components = []

        for e in leftovers:
            if e["id"] in visited:
                continue

            stack = [e["id"]]
            comp_ids = []

            while stack:
                curr = stack.pop()
                if curr in visited:
                    continue
                visited.add(curr)
                comp_ids.append(curr)

                for nb in adjacency[curr]:
                    if nb not in visited:
                        stack.append(nb)

            components.append([remaining_by_id[i] for i in comp_ids])

        for comp_edges in components:
            if not comp_edges:
                continue

            comp_used = set()

            start_candidates = []
            for e in comp_edges:
                has_prev = any(
                    other["id"] != e["id"] and other["naar"] == e["van"]
                    for other in comp_edges
                )
                if not has_prev:
                    start_candidates.append(e)

            # ---------------------------------
            # NIET-RONDE KETTINGEN
            # ---------------------------------
            if start_candidates:
                start_candidates.sort(key=stable_edge_key)

                for start_edge in start_candidates:
                    if start_edge["id"] in comp_used:
                        continue

                    chain = roll_chain_from_start_edge(start_edge, comp_edges, comp_used)
                    add_chain_record_if_needed(chain_records, chain)

                rest = [e for e in comp_edges if e["id"] not in comp_used]
                rest.sort(key=stable_edge_key)

                for edge in rest:
                    if edge["id"] in comp_used:
                        continue
                    chain = roll_chain_from_start_edge(edge, comp_edges, comp_used)
                    add_chain_record_if_needed(chain_records, chain)

            # ---------------------------------
            # ECHTE RONDE LUS
            # ---------------------------------
            else:
                two_spot_candidates = [e for e in comp_edges if has_two_spots(e["van"])]

                if two_spot_candidates:
                    two_spot_candidates.sort(key=stable_edge_key)
                    start_edge = two_spot_candidates[0]
                else:
                    comp_edges.sort(key=stable_edge_key)
                    start_edge = comp_edges[0]

                chain = roll_chain_from_start_edge(start_edge, comp_edges, comp_used)

                rest = [e for e in comp_edges if e["id"] not in comp_used]
                rest.sort(key=stable_edge_key)
                chain.extend(rest)

                add_chain_record_if_needed(chain_records, chain)

    # -----------------------------
    # 5. Definitieve volgorde
    # -----------------------------
    auto_edges = [e for e in edges if e["type"] == "volledig automatisch"]
    normal_edges = [e for e in edges if e["type"] == "normaal"]

    ordered_auto = []
    used_auto = set()

    # Groen start ALTIJD vanuit de attractie van de nieuwkomer
    # en volgt dan pas de ketting verder.
    for newcomer in newcomers:
        start_attr = newcomer["naar"]

        start_candidates = [
            e for e in auto_edges
            if e["id"] not in used_auto and e["van"] == start_attr
        ]
        start_candidates.sort(key=next_edge_key)

        for start in start_candidates:
            current = start
            while current and current["id"] not in used_auto:
                ordered_auto.append(current)
                used_auto.add(current["id"])

                next_candidates = [
                    e for e in auto_edges
                    if e["id"] not in used_auto and e["van"] == current["naar"]
                ]
                next_candidates.sort(key=next_edge_key)
                current = next_candidates[0] if next_candidates else None

    leftovers_auto = [e for e in auto_edges if e["id"] not in used_auto]
    leftovers_auto.sort(key=stable_edge_key)
    ordered_auto.extend(leftovers_auto)

    # niet-ronde kettingen eerst, dan ronde lussen
    chain_records.sort(
        key=lambda rec: (
            0 if rec["shape"] == "open" else 1,
            0 if rec["shape"] == "cycle" and rec["start_has_two_spots"] else 1,
            stable_edge_key(rec["edges"][0]) if rec["edges"] else ("", "", "")
        )
    )

    ordered_half = []
    for rec in chain_records:
        ordered_half.extend(rec["edges"])

    normal_edges.sort(key=stable_edge_key)

    # Extra veiligheid tegen dubbels
    seen_ids = set()
    final_order = []

    for e in ordered_auto + ordered_half + normal_edges:
        if e["id"] in seen_ids:
            continue
        seen_ids.add(e["id"])
        final_order.append(e)

    return final_order



# ─────────────────────────────────────────────────────────────────
# DEEL 6 uitvoer als herbruikbare functie
# ─────────────────────────────────────────────────────────────────
def maak_wisselplanning_sheet(wb_arg, am_arg):
    """
    Bouw het 'Wissels'-werkblad op basis van am_arg (assigned_map).
    Vervangt het bestaande sheet als het al bestaat.
    """
    # Verwijder oud sheet indien aanwezig
    if "Wissels" in wb_arg.sheetnames:
        del wb_arg["Wissels"]

    # Stap 1: student → uur → attractie
    student_per_uur = build_student_per_hour_map(am_arg)

    # Stap 2: veranderingen per uur opbouwen
    changes_per_hour = extract_hourly_changes(student_per_uur, open_uren)

    # Stap 3: per uur classificeren en ordenen
    wissels_per_uur = {}
    for uur in sorted(open_uren):
        newcomers        = changes_per_hour[uur]["newcomers"]
        movers           = changes_per_hour[uur]["movers"]
        leavers          = changes_per_hour[uur]["leavers"]
        disappearing_sources = changes_per_hour[uur]["disappearing_sources"]

        ordered_switches = classify_hourly_switches(
            uur, newcomers, movers, leavers, disappearing_sources
        )
        if ordered_switches:
            wissels_per_uur[uur] = ordered_switches

    # KPI berekenen
    totaal_wissels = 0
    aantal_auto    = 0
    for uur in wissels_per_uur:
        for w in wissels_per_uur[uur]:
            totaal_wissels += 1
            if w["type"] == "volledig automatisch":
                aantal_auto += 1
    niet_groen = totaal_wissels - aantal_auto

    # Stap 4: werkblad aanmaken
    ws_wissels = wb_arg.create_sheet(title="Wissels")

    # KPI rechts van de tabel (kolom G)
    ws_wissels.cell(1, 7, "KPI Wissels").font = Font(bold=True)
    ws_wissels.cell(2, 7, "Totaal wissels:")
    ws_wissels.cell(2, 8, totaal_wissels)
    ws_wissels.cell(3, 7, "Volledig automatisch:")
    ws_wissels.cell(3, 8, aantal_auto)
    ws_wissels.cell(4, 7, "Niet-groen (KPI):")
    ws_wissels.cell(4, 8, niet_groen)
    ws_wissels.cell(4, 8).font = Font(bold=True)

    _center = Alignment(horizontal="center", vertical="center")
    _border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"),  bottom=Side(style="thin")
    )
    green_fill  = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    orange_fill = PatternFill(start_color="F4B084", end_color="F4B084", fill_type="solid")

    current_row = 1
    for uur in sorted(wissels_per_uur.keys()):
        # Titelrij per uur
        title_cell = ws_wissels.cell(current_row, 1, f"Wissels om {formatteer_uur(uur)}")
        title_cell.font      = Font(bold=True)
        title_cell.alignment = _center
        current_row += 1

        # Headers
        for col_idx, header in enumerate(["Student", "Van", "Naar"], start=1):
            cell           = ws_wissels.cell(current_row, col_idx, header)
            cell.font      = Font(bold=True)
            cell.alignment = _center
            cell.border    = _border
        current_row += 1

        # Wissels
        for w in wissels_per_uur[uur]:
            ws_wissels.cell(current_row, 1, w["naam"])
            ws_wissels.cell(current_row, 2, w["van"])
            ws_wissels.cell(current_row, 3, w["naar"])

            for col_idx in range(1, 4):
                cell           = ws_wissels.cell(current_row, col_idx)
                cell.alignment = _center
                cell.border    = _border

            if w["type"] == "volledig automatisch":
                ws_wissels.cell(current_row, 2).fill = green_fill
                ws_wissels.cell(current_row, 3).fill = green_fill
            elif w["type"] == "half-automatisch":
                ws_wissels.cell(current_row, 2).fill = yellow_fill
                ws_wissels.cell(current_row, 3).fill = yellow_fill
            elif w["type"] == "half-start":
                ws_wissels.cell(current_row, 2).fill = orange_fill
                ws_wissels.cell(current_row, 3).fill = orange_fill

            current_row += 1
        current_row += 1  # lege rij tussen uren

    # Stap 5: kolombreedtes
    for col_idx, breedte in {1: 22, 2: 25, 3: 25, 7: 24, 8: 18}.items():
        ws_wissels.column_dimensions[get_column_letter(col_idx)].width = breedte


# ── oorspronkelijke aanroep (vervangt de oude losse code) ──
maak_wisselplanning_sheet(wb_out, assigned_map)

# -----------------------------
# Werkblad Heropleidingen
# -----------------------------
from openpyxl.styles.proxy import StyleProxy
from copy import copy

ws_bron = wb.worksheets[[ws.title for ws in wb.worksheets].index("Heropleidingen")] if "Heropleidingen" in wb.sheetnames else None
if ws_bron:
    ws_hero = wb_out.create_sheet(title="Heropleidingen")
    for rij in ws_bron.iter_rows():
        for cel in rij:
            nieuwe_cel = ws_hero.cell(row=cel.row, column=cel.column, value=cel.value)
            if cel.has_style:
                nieuwe_cel.font = copy(cel.font)
                nieuwe_cel.fill = copy(cel.fill)
                nieuwe_cel.border = copy(cel.border)
                nieuwe_cel.alignment = copy(cel.alignment)
                nieuwe_cel.number_format = cel.number_format
    for kol, breedte in ws_bron.column_dimensions.items():
        ws_hero.column_dimensions[kol].width = breedte.width
        ws_hero.column_dimensions["A"].width = 11
    for rij, hoogte in ws_bron.row_dimensions.items():
        ws_hero.row_dimensions[rij].height = hoogte.height



#NIEUWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWW
#NIEUWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWW

# -----------------------------
# Werkbladen altijd verbergen
# -----------------------------
for bladnaam in ["Pauzevlinders", "Feedback"]:
    if bladnaam in wb_out.sheetnames:
        ws_hide = wb_out[bladnaam]
        ws_hide.sheet_state = "veryHidden" 

# Snapshot voor last-minute (zonder dringende heropleidingen)
if "lm_base_bytes" not in st.session_state:
    _buf = BytesIO()
    wb_out.save(_buf)
    st.session_state["lm_base_bytes"] = _buf.getvalue()

output_lm_base = BytesIO(st.session_state["lm_base_bytes"])

# -----------------------------
# Dringende heropleidingen in Planning
# -----------------------------
ws_plan = wb_out["Planning"]
laatste_rij = ws_plan.max_row
invoegrij = laatste_rij + 2

for rij in ws_hero.iter_rows():
    if rij[0].value == "Belangrijk!":
        naam = rij[1].value or ""
        omschrijving = rij[2].value or ""
        ws_plan.cell(invoegrij, 1).value = f"Dringende heropleiding: {naam}: {omschrijving}"
        invoegrij += 1



#ooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooo


# -----------------------------
# Opslaan in hetzelfde unieke bestand als DEEL 3
# -----------------------------
output = BytesIO()
wb_out.save(output)
output.seek(0)
# st.success("Planning gegenereerd!")
st.download_button(
    "Download planning",
    data=output.getvalue(),
    file_name=f"Planning_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
)





#DEELLL 8 OFZOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOO
# ============================================================
# LAST-MINUTE AFWEZIGEN V5
# VOLLEDIGE VERVANGING van alle vorige last-minute patches
# Plakken ONDER de bestaande st.download_button("Download planning", ...)
# ============================================================

import copy
import random
from collections import defaultdict, Counter
from io import BytesIO
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# ------------------------------------------------------------
# Basishelpers
# ------------------------------------------------------------
def lm5_split_display_label(label):
    if not label:
        return "", 1
    s = str(label).strip()
    parts = s.rsplit(" ", 1)
    if len(parts) == 2 and parts[1].isdigit():
        return parts[0].strip(), int(parts[1])
    return s, 1

def lm5_is_pv_row(label):
    return str(label).strip().lower().startswith("pauzevlinder")

def lm5_is_extra_row(label):
    return str(label).strip().lower().startswith("extra")

def lm5_parse_output_hour(header):
    if not header:
        return None
    s = str(header).strip().lower()
    if s == "9u30-11u":
        return min(open_uren) if open_uren else 10
    if s == "18u-19u30":
        return max(open_uren) if open_uren else 18
    # Strip label-gedeelte tussen haakjes, bv. "9u30 (0,5h)" → "9u30"
    kern = s.split("(")[0].strip()
    if "u" in kern:
        delen = kern.split("u")
        try:
            uren = int(delen[0])
            min_str = delen[1].strip() if len(delen) > 1 else ""
            minuten = int(min_str) if min_str.isdigit() else 0
            return uren + minuten / 60
        except:
            return None
    if ":" in kern:
        delen = kern.split(":")
        try:
            return int(delen[0]) + int(delen[1]) / 60
        except:
            return None
    try:
        return float(kern)
    except:
        return None
        

def lm5_student_lookup():
    return {str(s["naam"]).strip(): s for s in studenten}

def lm5_get_student(naam):
    return lm5_student_lookup().get(str(naam).strip())

def lm5_copy_student_state(student):
    return {
        "naam": str(student["naam"]).strip(),
        "uren_beschikbaar": list(student["uren_beschikbaar"]),
        "attracties": list(student["attracties"]),
        "aantal_attracties": student["aantal_attracties"],
        "is_pauzevlinder": student["is_pauzevlinder"],
        "pv_number": student["pv_number"],
        "assigned_attracties": set(),
        "assigned_hours": []
    }

def lm5_student_can_attr(student, attr):
    if not student:
        return False
    return student_kan_attr(student, attr)

def lm5_get_hours_on_attr(ctx, student_name, attr):
    uren = []
    for (uur, a), namen in ctx["assigned_map"].items():
        if a == attr and student_name in namen:
            uren.append(uur)
    return sorted(set(uren))

def lm5_count_switches(ctx, student_name):
    uur_attr = []
    for uur in sorted(set(ctx["student_states"][student_name]["assigned_hours"])):
        attr = lm5_student_current_attr_on_hour(ctx, student_name, uur)
        if attr:
            uur_attr.append((uur, attr))
    if not uur_attr:
        return 0
    switches = 0
    prev = uur_attr[0][1]
    for _, attr in uur_attr[1:]:
        if attr != prev:
            switches += 1
        prev = attr
    return switches

# ------------------------------------------------------------
# Input + originele planning lezen
# ------------------------------------------------------------
def lm5_extract_base_maps(base_bytes):
    wb_tmp = load_workbook(BytesIO(base_bytes))
    ws_plan = wb_tmp["Planning"]

    uur_to_col = {}
    for col in range(2, ws_plan.max_column + 1):
        uur = lm5_parse_output_hour(ws_plan.cell(1, col).value)
        if uur in open_uren and uur not in uur_to_col:
            uur_to_col[uur] = col

    if not uur_to_col:
        raise ValueError("Geen geldige uurkolommen gevonden in blad 'Planning'.")

    attr_rows = []
    pv_rows = []
    extra_rows = []
    student_hour_attr = {}
    student_hour_row = defaultdict(str)
    row_labels = []

    for row in range(2, ws_plan.max_row + 1):
        label = ws_plan.cell(row, 1).value
        if not label:
            continue
        label = str(label).strip()
        row_labels.append((row, label))

        if lm5_is_pv_row(label):
            pv_rows.append((row, label))
        elif lm5_is_extra_row(label):
            extra_rows.append((row, label))
        else:
            attr_rows.append((row, label))

        for uur, col in uur_to_col.items():
            naam = ws_plan.cell(row, col).value
            if naam and str(naam).strip():
                naam = str(naam).strip()
                attr, _pos = lm5_split_display_label(label)
                student_hour_attr[(naam, uur)] = attr
                student_hour_row[(naam, uur)] = label

    return {
        "wb": wb_tmp,
        "ws_plan": ws_plan,
        "uur_to_col": uur_to_col,
        "attr_rows": attr_rows,
        "pv_rows": pv_rows,
        "extra_rows": extra_rows,
        "student_hour_attr": student_hour_attr,
        "student_hour_row": student_hour_row,
        "row_labels": row_labels,
    }

def lm5_working_students_today(base_maps):
    out = set()
    for (naam, uur), _attr in base_maps["student_hour_attr"].items():
        if uur in open_uren:
            out.add(naam)
    return sorted(out)

def lm5_present_students_on_hour(base_maps, uur, absentees_set):
    out = []
    seen = set()
    for (naam, uur2), _attr in base_maps["student_hour_attr"].items():
        if uur2 == uur and naam not in absentees_set and naam not in seen:
            out.append(naam)
            seen.add(naam)
    return out

def lm5_pv_names():
    return [str(pv["naam"]).strip() for pv in selected]

def lm5_bereken_pauze_counts(absentees_set, base_maps):
    """
    Bereken lange_pauzes en korte_pauzes voor de last-minute planning
    op basis van werkelijke werkuren (excl. afwezigen).

    lange_pauzes = (studenten > 6u) + (minderjarige studenten >= 4u)
    korte_pauzes = studenten >= 4u
    """
    uren_per_student = defaultdict(set)
    for (naam, uur), _attr in base_maps["student_hour_attr"].items():
        if naam not in absentees_set:
            uren_per_student[naam].add(uur)

    lange_pauzes = 0
    korte_pauzes = 0
    for naam, uren in uren_per_student.items():
        n = len(uren)
        is_minor = "-18" in str(naam)
        if n > 6:
            lange_pauzes += 1
        if is_minor and n >= 4:
            lange_pauzes += 1
        if n >= 4:
            korte_pauzes += 1

    return lange_pauzes, korte_pauzes


def lm5_extract_capacity_actions():
    result = []

    # L3:M12 in Aanpassingen
    for rij in range(3, 13):
        left_source = ws_aanpassingen.cell(rij, 12).value   # kolom L
        right_source = ws_aanpassingen.cell(rij, 13).value  # kolom M

        left = str(left_source).strip() if left_source is not None and str(left_source).strip() != "" else ""
        right = str(right_source).strip() if right_source is not None and str(right_source).strip() != "" else ""

        if not left and not right:
            continue

        if left and not right:
            result.append({"type": "disable", "attr": left, "source_row": rij})

        elif left and right:
            result.append({"type": "merge", "groep": [left, right], "source_row": rij})

    return result

def lm5_all_single_attrs():
    return [a for a in attracties_te_plannen if " + " not in str(a)]

def lm5_full_capable_students():
    singles = lm5_all_single_attrs()
    out = []
    for s in studenten:
        if all(lm5_student_can_attr(s, attr) for attr in singles):
            out.append(str(s["naam"]).strip())
    return out

# ------------------------------------------------------------
# Pauzevlinder-vervangers
# ------------------------------------------------------------
def lm5_pick_pv_replacements(absent_pv_names, start_uur, base_maps, absentees_set):
    full_capable = set(lm5_full_capable_students())
    pv_name_set = set(lm5_pv_names())

    future_workers = set()
    for (naam, uur), _attr in base_maps["student_hour_attr"].items():
        if uur >= start_uur and naam not in absentees_set:
            future_workers.add(naam)

    chosen = {}
    used = set()

    for pvnaam in absent_pv_names:
        kandidaten = []
        for naam in sorted(future_workers):
            if naam in pv_name_set:
                continue
            if naam in used:
                continue
            if naam not in full_capable:
                continue

            orig_extra_count = 0
            orig_total_count = 0
            for (n, uur), rowlabel in base_maps["student_hour_row"].items():
                if n == naam and uur >= start_uur:
                    orig_total_count += 1
                    if "extra" in str(rowlabel).lower():
                        orig_extra_count += 1

            student = lm5_get_student(naam)
            breedte = student["aantal_attracties"] if student else 999

            kandidaten.append((
                -orig_extra_count,
                orig_total_count,
                -breedte,
                naam
            ))

        if kandidaten:
            kandidaten.sort()
            top = [x[3] for x in kandidaten[:min(3, len(kandidaten))]]
            chosen[pvnaam] = random.choice(top)
            used.add(chosen[pvnaam])

    return chosen

def lm5_active_pv_assignment_for_hour(ctx, uur):
    if uur not in required_pauze_hours:
        return {}

    vrijgegeven_uren = ctx.get("vrijgegeven_pv_uren", set())
    vrijgegeven_pv   = ctx.get("vrijgegeven_pv_naam", None)

    result = {}
    for pvnaam in lm5_pv_names():
        if uur in vrijgegeven_uren and pvnaam == vrijgegeven_pv:
            continue
        if pvnaam in ctx["abs_set"]:
            vervanger = ctx["pv_replacements"].get(pvnaam)
            if vervanger:
                result[pvnaam] = vervanger
        else:
            result[pvnaam] = pvnaam
    return result

def lm5_present_attraction_students_on_hour(ctx, uur):
    bruto = lm5_present_students_on_hour(ctx["base_maps"], uur, ctx["abs_set"])
    pv_assignment = lm5_active_pv_assignment_for_hour(ctx, uur)
    bezette_pv_mensen = set(pv_assignment.values())
    return [naam for naam in bruto if naam not in bezette_pv_mensen]

# ------------------------------------------------------------
# Uurstaat herberekenen
# ------------------------------------------------------------
def lm5_rebuild_hour_state(uur, available_attraction_students, capacity_actions):
    counts = {}
    active = set()
    debug_actions = []

    # -----------------------------
    # Basis actieve attracties opbouwen voor DIT uur
    # -----------------------------
    for attr in attracties_te_plannen:
        if " + " in str(attr):
            continue

        if uur in dichte_uren_per_attr.get(normalize_attr(attr), set()):
            counts[attr] = 0
        else:
            if aantallen_raw.get(attr, 0) >= 1:
                counts[attr] = 1
                active.add(attr)
            else:
                counts[attr] = 0

    # -----------------------------
    # Vaste samenvoegingen van DIT uur eerst toepassen
    # -----------------------------
    groepen = []
    for groep in uur_samenvoegingen.get(uur, []):
        g = [str(x).strip() for x in groep if x and str(x).strip()]
        if len(g) < 2:
            continue

        groepen.append(g)
        sameng = " + ".join(g)
        counts[sameng] = 1
        active.add(sameng)

        for onderdeel in g:
            counts[onderdeel] = 0
            if onderdeel in active:
                active.remove(onderdeel)

    def min_spots():
        return sum(1 for a in active if counts.get(a, 0) >= 1)

    def samengestelde_naam_van_groep(g):
        return " + ".join(g)

    def groep_is_al_actief(g):
        sameng = samengestelde_naam_van_groep(g)
        return sameng in active and all(counts.get(a, 0) == 0 for a in g)

    def merge_is_mogelijk_en_geeft_reductie(g):
        # Alleen als ALLE losse onderdelen nog actief zijn
        # Dan krijg je een echte reductie van minstens 1 plek
        if len(g) < 2:
            return False
        if not all(a in active for a in g):
            return False

        before = min_spots()
        after = before - len(g) + 1
        return after < before

    def disable_is_mogelijk_en_geeft_reductie(attr):
        if attr not in active:
            return False
        if counts.get(attr, 0) < 1:
            return False

        before = min_spots()
        after = before - 1
        return after < before

    debug_actions.append(
        f"Uur {uur}: START | available={available_attraction_students} | min_spots={min_spots()} | active_start={sorted(active)}"
    )

    # -----------------------------
    # PER UUR: rij voor rij door CD/CE
    # -----------------------------
    while min_spots() > available_attraction_students:
        found_reduction = False

        for entry in capacity_actions:
            before = min_spots()

            if before <= available_attraction_students:
                debug_actions.append(
                    f"Uur {uur}: STOP acties | min_spots={before} <= available={available_attraction_students}"
                )
                break

            if entry["type"] == "merge":
                g = [str(x).strip() for x in entry["groep"] if x and str(x).strip()]
                sameng = samengestelde_naam_van_groep(g)
                source_row = entry.get("source_row", "?")

                if groep_is_al_actief(g):
                    debug_actions.append(
                        f"Uur {uur}: RIJ {source_row} MERGE SKIP | groep={g} | reden=al_actief"
                    )
                    continue

                if not merge_is_mogelijk_en_geeft_reductie(g):
                    debug_actions.append(
                        f"Uur {uur}: RIJ {source_row} MERGE SKIP | groep={g} | reden=niet_mogelijk_op_dit_uur | current_active={sorted(active)}"
                    )
                    continue

                for onderdeel in g:
                    counts[onderdeel] = 0
                    if onderdeel in active:
                        active.remove(onderdeel)

                counts[sameng] = 1
                active.add(sameng)

                if g not in groepen:
                    groepen.append(g)

                after = min_spots()

                if after < before:
                    found_reduction = True
                    debug_actions.append(
                        f"Uur {uur}: RIJ {source_row} MERGE OK | groep={g} | min_spots {before}->{after} | new_active={sorted(active)}"
                    )
                    break
                else:
                    debug_actions.append(
                        f"Uur {uur}: RIJ {source_row} MERGE GEEN EFFECT | groep={g}"
                    )

            elif entry["type"] == "disable":
                attr = str(entry["attr"]).strip()
                source_row = entry.get("source_row", "?")

                if not disable_is_mogelijk_en_geeft_reductie(attr):
                    debug_actions.append(
                        f"Uur {uur}: RIJ {source_row} DISABLE SKIP | attr={attr} | reden=niet_mogelijk_op_dit_uur"
                    )
                    continue

                counts[attr] = 0
                if attr in active:
                    active.remove(attr)

                after = min_spots()

                if after < before:
                    found_reduction = True
                    debug_actions.append(
                        f"Uur {uur}: RIJ {source_row} DISABLE OK | attr={attr} | min_spots {before}->{after} | new_active={sorted(active)}"
                    )
                    break
                else:
                    debug_actions.append(
                        f"Uur {uur}: RIJ {source_row} DISABLE GEEN EFFECT | attr={attr}"
                    )

        if not found_reduction:
            debug_actions.append(
                f"Uur {uur}: GEEN VERDERE REDUCTIE MOGELIJK | min_spots={min_spots()} | available={available_attraction_students}"
            )
            break

    # -----------------------------
    # Tweede plaatsen pas NA merge/disable
    # -----------------------------
    second_spot_blocked_lm = set()
    base_spots = min_spots()
    extra_spots = available_attraction_students - base_spots

    debug_actions.append(
        f"Uur {uur}: VOOR 2DE PLEKKEN | base_spots={base_spots} | extra_spots={extra_spots}"
    )

    for attr in second_priority_order:
        if attr in active and aantallen_raw.get(attr, 0) == 2:
            if extra_spots > 0:
                counts[attr] = 2
                extra_spots -= 1
                debug_actions.append(
                    f"Uur {uur}: 2DE PLEK OPEN | attr={attr} | resterend_extra_spots={extra_spots}"
                )
            else:
                counts[attr] = 1
                second_spot_blocked_lm.add(attr)
                debug_actions.append(
                    f"Uur {uur}: 2DE PLEK DICHT | attr={attr}"
                )

    # -----------------------------
    # Red spots opnieuw opbouwen
    # -----------------------------
    red_spots_lm = set()

    samengestelde_actief = set(" + ".join(g) for g in groepen)
    losse_in_samenvoeging = set(a for g in groepen for a in g)

    for attr in losse_in_samenvoeging:
        red_spots_lm.add(attr)

    for samengestelde_attr in samengevoegde_attracties:
        if samengestelde_attr not in samengestelde_actief:
            red_spots_lm.add(samengestelde_attr)

    for attr in list(counts.keys()):
        if " + " not in str(attr):
            if uur in dichte_uren_per_attr.get(normalize_attr(attr), set()):
                red_spots_lm.add(attr)

    debug_actions.append(
        f"Uur {uur}: EINDE | active_end={sorted(active)} | counts={{{', '.join(f'{k}: {v}' for k, v in sorted(counts.items()))}}}"
    )

    return {
        "counts": counts,
        "active": active,
        "groepen": groepen,
        "red_spots": red_spots_lm,
        "second_spot_blocked": second_spot_blocked_lm,
        "debug_actions": debug_actions,
    }


# ------------------------------------------------------------
# Toewijzingshelpers
# ------------------------------------------------------------
def lm5_student_current_attr_on_hour(ctx, naam, uur):
    for (h, attr), namen in ctx["assigned_map"].items():
        if h == uur and naam in namen:
            return attr
    return None

def lm5_can_place_student_on_attr(ctx, student, attr, uren):
    if not lm5_student_can_attr(student, attr):
        return False

    for uur in uren:
        hstate = ctx["hour_states"].get(uur)
        if not hstate:
            return False

        if attr not in hstate["active"]:
            return False
        if attr in hstate["red_spots"]:
            return False

        max_spots = hstate["counts"].get(attr, 0)
        if attr in hstate["second_spot_blocked"]:
            max_spots = min(max_spots, 1)

        if ctx["per_hour_assigned_counts"][uur][attr] >= max_spots:
            return False

        if uur in student["assigned_hours"]:
            return False

        pv_assignment_now = lm5_active_pv_assignment_for_hour(ctx, uur)
        if student["naam"] in set(pv_assignment_now.values()):
            return False

    bestaande = sorted([
        h for h in set(student["assigned_hours"])
        if student["naam"] in ctx["assigned_map"].get((h, attr), [])
    ])
    totaal = sorted(set(bestaande) | set(uren))

    if len(totaal) > 6:
        return False

    return True

def lm5_place_student_on_attr(ctx, student, attr, uren):
    for uur in uren:
        ctx["assigned_map"][(uur, attr)].append(student["naam"])
        ctx["per_hour_assigned_counts"][uur][attr] += 1
        if uur not in student["assigned_hours"]:
            student["assigned_hours"].append(uur)
    student["assigned_attracties"].add(attr)
    ctx["prev_attr"][student["naam"]] = attr

def lm5_remove_student_from_attr_hours(ctx, student, attr, uren):
    for uur in uren:
        namen = ctx["assigned_map"].get((uur, attr), [])
        if student["naam"] in namen:
            namen.remove(student["naam"])
            if ctx["per_hour_assigned_counts"][uur][attr] > 0:
                ctx["per_hour_assigned_counts"][uur][attr] -= 1

def lm5_rebuild_student_attracties(ctx, student):
    attrs = set()
    for (uur, attr), namen in ctx["assigned_map"].items():
        if student["naam"] in namen:
            attrs.add(attr)
    for a in list(student["assigned_attracties"]):
        if a in ["Extra", "Pauzevlinder-vervanging"]:
            attrs.add(a)
    student["assigned_attracties"] = attrs

def lm5_original_attr_score(ctx, naam, attr, uren):
    same = 0
    mismatches = 0
    for uur in uren:
        orig_attr = ctx["base_maps"]["student_hour_attr"].get((naam, uur), "")
        if normalize_attr(orig_attr) == normalize_attr(attr):
            same += 1
        else:
            mismatches += 1
    return (-same, mismatches)

def lm5_candidate_attr_score(ctx, student, attr, uren):
    naam = student["naam"]
    orig_score = lm5_original_attr_score(ctx, naam, attr, uren)

    prev_penalty = 0
    prev_attr = ctx["prev_attr"].get(naam, "")
    if prev_attr and normalize_attr(prev_attr) != normalize_attr(attr):
        prev_penalty = 1

    bestaande = sorted([
        h for h in set(student["assigned_hours"])
        if naam in ctx["assigned_map"].get((h, attr), [])
    ])
    totaal = sorted(set(bestaande) | set(uren))

    run_penalty = 0 if len(uren) >= 3 else (1 if len(uren) == 2 else 2)
    over4_penalty = 1 if len(totaal) > 4 else 0

    return (
        run_penalty,
        prev_penalty,
        over4_penalty,
        orig_score[0],
        orig_score[1],
        ctx["changes_count"][naam],
        attr
    )

def lm5_try_place_best_block(ctx, student, future_hours, start_idx):
    if start_idx >= len(future_hours):
        return False, start_idx + 1

    for block_size in [3, 2, 1]:
        uren = future_hours[start_idx:start_idx + block_size]
        if len(uren) < block_size:
            continue
        if any(uur in student["assigned_hours"] for uur in uren):
            continue

        candidate_attrs = []
        for attr in attracties_te_plannen:
            if lm5_can_place_student_on_attr(ctx, student, attr, uren):
                candidate_attrs.append(attr)

        if candidate_attrs:
            candidate_attrs.sort(key=lambda a: lm5_candidate_attr_score(ctx, student, a, uren))
            best_attr = candidate_attrs[0]
            lm5_place_student_on_attr(ctx, student, best_attr, uren)
            return True, start_idx + block_size

    return False, start_idx + 1

# ------------------------------------------------------------
# Uur-per-uur seed
# ------------------------------------------------------------
def lm5_seed_same_place_first(ctx, uur, target_slots, present_attraction_students):
    used_students = set()
    assigned_rows = set()

    for attr, pos, rijlabel in target_slots:
        orig_names = []
        for (naam, uur2), rowlabel in ctx["base_maps"]["student_hour_row"].items():
            if uur2 == uur and rowlabel == rijlabel and naam in present_attraction_students:
                orig_names.append(naam)

        for naam in orig_names:
            if naam in used_students:
                continue
            student = ctx["student_states"][naam]
            if lm5_can_place_student_on_attr(ctx, student, attr, [uur]):
                lm5_place_student_on_attr(ctx, student, attr, [uur])
                used_students.add(naam)
                assigned_rows.add(rijlabel)
                break

    return used_students, assigned_rows

def lm5_fill_remaining_hour(ctx, uur, target_slots, present_attraction_students, used_students, assigned_rows):
    remaining_slots = [(attr, pos, rijlabel) for attr, pos, rijlabel in target_slots if rijlabel not in assigned_rows]
    remaining_students = [n for n in present_attraction_students if n not in used_students]

    for attr, pos, rijlabel in remaining_slots:
        kandidaten = []
        for naam in remaining_students:
            student = ctx["student_states"][naam]
            if not lm5_can_place_student_on_attr(ctx, student, attr, [uur]):
                continue

            orig_attr = ctx["base_maps"]["student_hour_attr"].get((naam, uur), "")
            same_orig = 0 if normalize_attr(orig_attr) == normalize_attr(attr) else 1
            same_prev = 0 if normalize_attr(ctx["prev_attr"].get(naam, "")) == normalize_attr(attr) else 1

            kandidaten.append((
                same_orig,
                same_prev,
                ctx["changes_count"][naam],
                student["aantal_attracties"],
                naam
            ))

        if kandidaten:
            kandidaten.sort()
            gekozen = kandidaten[0][4]
            student = ctx["student_states"][gekozen]
            lm5_place_student_on_attr(ctx, student, attr, [uur])

            orig_attr = ctx["base_maps"]["student_hour_attr"].get((gekozen, uur), "")
            if normalize_attr(orig_attr) != normalize_attr(attr):
                ctx["changes_count"][gekozen] += 1

            used_students.add(gekozen)
            remaining_students.remove(gekozen)

# ------------------------------------------------------------
# Vrijgekomen studenten + lege plaatsen
# ------------------------------------------------------------
def lm5_collect_released_students_and_missing_slots(ctx, base_maps, start_uur):
    released_students = defaultdict(list)
    missing_slots_by_hour = defaultdict(list)

    for uur in sorted(open_uren):
        if uur < start_uur:
            continue

        present_attraction_students = lm5_present_attraction_students_on_hour(ctx, uur)
        hstate = ctx["hour_states"][uur]
        target_slots, _inactive = lm5_build_target_slots_for_hour(base_maps["attr_rows"], hstate)

        for attr, pos, rijlabel in target_slots:
            namen = ctx["assigned_map"].get((uur, attr), [])
            if len(namen) < pos:
                missing_slots_by_hour[uur].append((attr, pos, rijlabel))

        for naam in present_attraction_students:
            if uur not in ctx["student_states"][naam]["assigned_hours"]:
                released_students[uur].append(naam)

    return released_students, missing_slots_by_hour

# ------------------------------------------------------------
# Eerst directe invulling
# ------------------------------------------------------------
def lm5_try_direct_fill_from_released_students(ctx, released_students, missing_slots_by_hour):
    any_change = False

    for uur in sorted(missing_slots_by_hour.keys()):
        slots = list(missing_slots_by_hour[uur])
        remaining_released = list(released_students.get(uur, []))
        new_slots = []

        for (attr, pos, rijlabel) in slots:
            gevuld = False
            kandidaten = []

            for naam in remaining_released:
                student = ctx["student_states"][naam]
                if lm5_can_place_student_on_attr(ctx, student, attr, [uur]):
                    orig_attr = ctx["base_maps"]["student_hour_attr"].get((naam, uur), "")
                    kandidaten.append((
                        0 if normalize_attr(orig_attr) == normalize_attr(attr) else 1,
                        0 if normalize_attr(ctx["prev_attr"].get(naam, "")) == normalize_attr(attr) else 1,
                        ctx["changes_count"][naam],
                        naam
                    ))

            if kandidaten:
                kandidaten.sort()
                gekozen = kandidaten[0][3]
                student = ctx["student_states"][gekozen]
                lm5_place_student_on_attr(ctx, student, attr, [uur])

                if gekozen in remaining_released:
                    remaining_released.remove(gekozen)

                orig_attr = ctx["base_maps"]["student_hour_attr"].get((gekozen, uur), "")
                if normalize_attr(orig_attr) != normalize_attr(attr):
                    ctx["changes_count"][gekozen] += 1

                gevuld = True
                any_change = True

            if not gevuld:
                new_slots.append((attr, pos, rijlabel))

        released_students[uur] = remaining_released
        missing_slots_by_hour[uur] = new_slots

    return any_change

# ------------------------------------------------------------
# Kettingwissels op exact hetzelfde blok
# ------------------------------------------------------------
def lm5_student_has_exact_block_on_attr(ctx, naam, attr, block_hours):
    for uur in block_hours:
        huidige_attr = lm5_student_current_attr_on_hour(ctx, naam, uur)
        if huidige_attr != attr:
            return False
    return True

def lm5_find_same_block_students(ctx, block_hours, exclude_names=None):
    if exclude_names is None:
        exclude_names = set()

    kandidaten = []
    alle_namen = sorted(ctx["student_states"].keys())

    for naam in alle_namen:
        if naam in exclude_names:
            continue

        attrs = set()
        ok = True
        for uur in block_hours:
            attr = lm5_student_current_attr_on_hour(ctx, naam, uur)
            if not attr:
                ok = False
                break
            attrs.add(attr)

        if ok and len(attrs) == 1:
            kandidaten.append((naam, list(attrs)[0]))

    return kandidaten

def lm5_can_student_take_attr_block(ctx, student, attr, block_hours):
    if not lm5_student_can_attr(student, attr):
        return False

    for uur in block_hours:
        hstate = ctx["hour_states"].get(uur)
        if not hstate:
            return False
        if attr not in hstate["active"]:
            return False
        if attr in hstate["red_spots"]:
            return False

        pv_assignment_now = lm5_active_pv_assignment_for_hour(ctx, uur)
        if student["naam"] in set(pv_assignment_now.values()):
            return False

        huidige_attr = lm5_student_current_attr_on_hour(ctx, student["naam"], uur)
        if huidige_attr and huidige_attr != attr:
            return False

        max_spots = hstate["counts"].get(attr, 0)
        if attr in hstate["second_spot_blocked"]:
            max_spots = min(max_spots, 1)

        current_count = ctx["per_hour_assigned_counts"][uur][attr]
        already_here = student["naam"] in ctx["assigned_map"].get((uur, attr), [])
        effective_count = current_count if already_here else current_count + 1

        if effective_count > max_spots:
            return False

    bestaande = sorted([
        h for h in set(student["assigned_hours"])
        if student["naam"] in ctx["assigned_map"].get((h, attr), [])
    ])
    totaal = sorted(set(bestaande) | set(block_hours))
    if len(totaal) > 6:
        return False
    return True

def lm5_try_chain_swap_for_block(ctx, released_student_name, target_attr, block_hours):
    released_student = ctx["student_states"][released_student_name]

    kandidaten = lm5_find_same_block_students(
        ctx=ctx,
        block_hours=block_hours,
        exclude_names={released_student_name}
    )

    for andere_naam, attr_b in kandidaten:
        if attr_b == target_attr:
            continue

        andere_student = ctx["student_states"][andere_naam]

        if not lm5_student_can_attr(released_student, attr_b):
            continue
        if not lm5_student_can_attr(andere_student, target_attr):
            continue

        orig_switches_r = lm5_count_switches(ctx, released_student_name)
        orig_switches_o = lm5_count_switches(ctx, andere_naam)

        saved_assigned_map = copy.deepcopy(ctx["assigned_map"])
        saved_counts = copy.deepcopy(ctx["per_hour_assigned_counts"])
        saved_hours_r = list(released_student["assigned_hours"])
        saved_hours_o = list(andere_student["assigned_hours"])
        saved_attrs_r = set(released_student["assigned_attracties"])
        saved_attrs_o = set(andere_student["assigned_attracties"])
        saved_prev = dict(ctx["prev_attr"])

        lm5_remove_student_from_attr_hours(ctx, andere_student, attr_b, block_hours)

        if not lm5_can_student_take_attr_block(ctx, released_student, attr_b, block_hours):
            ctx["assigned_map"] = saved_assigned_map
            ctx["per_hour_assigned_counts"] = saved_counts
            released_student["assigned_hours"] = saved_hours_r
            andere_student["assigned_hours"] = saved_hours_o
            released_student["assigned_attracties"] = saved_attrs_r
            andere_student["assigned_attracties"] = saved_attrs_o
            ctx["prev_attr"] = saved_prev
            continue

        lm5_place_student_on_attr(ctx, released_student, attr_b, block_hours)

        if not lm5_can_student_take_attr_block(ctx, andere_student, target_attr, block_hours):
            ctx["assigned_map"] = saved_assigned_map
            ctx["per_hour_assigned_counts"] = saved_counts
            released_student["assigned_hours"] = saved_hours_r
            andere_student["assigned_hours"] = saved_hours_o
            released_student["assigned_attracties"] = saved_attrs_r
            andere_student["assigned_attracties"] = saved_attrs_o
            ctx["prev_attr"] = saved_prev
            continue

        lm5_place_student_on_attr(ctx, andere_student, target_attr, block_hours)

        lm5_rebuild_student_attracties(ctx, released_student)
        lm5_rebuild_student_attracties(ctx, andere_student)

        # harde 6u/4u check op betrokken attracties
        if len(lm5_get_hours_on_attr(ctx, released_student_name, attr_b)) > 6:
            ctx["assigned_map"] = saved_assigned_map
            ctx["per_hour_assigned_counts"] = saved_counts
            released_student["assigned_hours"] = saved_hours_r
            andere_student["assigned_hours"] = saved_hours_o
            released_student["assigned_attracties"] = saved_attrs_r
            andere_student["assigned_attracties"] = saved_attrs_o
            ctx["prev_attr"] = saved_prev
            continue

        if len(lm5_get_hours_on_attr(ctx, andere_naam, target_attr)) > 6:
            ctx["assigned_map"] = saved_assigned_map
            ctx["per_hour_assigned_counts"] = saved_counts
            released_student["assigned_hours"] = saved_hours_r
            andere_student["assigned_hours"] = saved_hours_o
            released_student["assigned_attracties"] = saved_attrs_r
            andere_student["assigned_attracties"] = saved_attrs_o
            ctx["prev_attr"] = saved_prev
            continue

        new_switches_r = lm5_count_switches(ctx, released_student_name)
        new_switches_o = lm5_count_switches(ctx, andere_naam)
        extra_switches = (new_switches_r - orig_switches_r) + (new_switches_o - orig_switches_o)

        if extra_switches > 1:
            ctx["assigned_map"] = saved_assigned_map
            ctx["per_hour_assigned_counts"] = saved_counts
            released_student["assigned_hours"] = saved_hours_r
            andere_student["assigned_hours"] = saved_hours_o
            released_student["assigned_attracties"] = saved_attrs_r
            andere_student["assigned_attracties"] = saved_attrs_o
            ctx["prev_attr"] = saved_prev
            continue

        ctx["changes_count"][released_student_name] += 1
        ctx["changes_count"][andere_naam] += 1
        return True

    return False

def lm5_try_fill_missing_with_chain_swaps(ctx, released_students, missing_slots_by_hour, start_uur):
    any_change = False
    future_hours = [u for u in sorted(open_uren) if u >= start_uur]

    for block_size in [3, 2, 1]:
        for i in range(len(future_hours) - block_size + 1):
            block_hours = future_hours[i:i + block_size]

            target_attrs = None
            for uur in block_hours:
                attrs_this_hour = set(attr for attr, _pos, _rijlabel in missing_slots_by_hour.get(uur, []))
                if target_attrs is None:
                    target_attrs = attrs_this_hour
                else:
                    target_attrs &= attrs_this_hour

            if not target_attrs:
                continue

            for target_attr in sorted(target_attrs):
                kandidaten_released = []
                for naam in sorted(set(n for uur in block_hours for n in released_students.get(uur, []))):
                    if all(naam in released_students.get(uur, []) for uur in block_hours):
                        kandidaten_released.append(naam)

                for released_name in kandidaten_released:
                    released_student = ctx["student_states"][released_name]

                    # eerst directe blokinvulling
                    if lm5_can_student_take_attr_block(ctx, released_student, target_attr, block_hours):
                        lm5_place_student_on_attr(ctx, released_student, target_attr, block_hours)
                        for uur in block_hours:
                            if released_name in released_students[uur]:
                                released_students[uur].remove(released_name)
                            removed = False
                            nieuwe = []
                            for item in missing_slots_by_hour[uur]:
                                if item[0] == target_attr and not removed:
                                    removed = True
                                    continue
                                nieuwe.append(item)
                            missing_slots_by_hour[uur] = nieuwe
                        any_change = True
                        break

                    # anders kettingwissel
                    if lm5_try_chain_swap_for_block(ctx, released_name, target_attr, block_hours):
                        for uur in block_hours:
                            if released_name in released_students[uur]:
                                released_students[uur].remove(released_name)
                            removed = False
                            nieuwe = []
                            for item in missing_slots_by_hour[uur]:
                                if item[0] == target_attr and not removed:
                                    removed = True
                                    continue
                                nieuwe.append(item)
                            missing_slots_by_hour[uur] = nieuwe
                        any_change = True
                        break

    return any_change

# ------------------------------------------------------------
# Nadien blokken 3/2/1 maken voor overige gaten
# ------------------------------------------------------------
def lm5_assign_future_blocks(ctx, start_uur):
    future_students = sorted({
        naam for (naam, uur), _attr in ctx["base_maps"]["student_hour_attr"].items()
        if uur >= start_uur and naam not in ctx["abs_set"]
    })

    for naam in future_students:
        student = ctx["student_states"][naam]
        future_hours = sorted({
            uur for (n, uur), _attr in ctx["base_maps"]["student_hour_attr"].items()
            if n == naam and uur >= start_uur
        })

        if not future_hours:
            continue

        i = 0
        while i < len(future_hours):
            uur = future_hours[i]

            if uur in student["assigned_hours"]:
                i += 1
                continue

            pv_assignment_now = lm5_active_pv_assignment_for_hour(ctx, uur)
            if naam in set(pv_assignment_now.values()):
                student["assigned_hours"].append(uur)
                student["assigned_attracties"].add("Pauzevlinder-vervanging")
                ctx["prev_attr"][naam] = "Pauzevlinder-vervanging"
                i += 1
                continue

            placed, next_i = lm5_try_place_best_block(ctx, student, future_hours, i)
            if placed:
                i = next_i
            else:
                i += 1

# ------------------------------------------------------------
# Exact 1 keer per gewerkt uur
# ------------------------------------------------------------
def lm5_force_exactly_one_assignment_per_hour(ctx, start_uur):
    for uur in sorted(open_uren):
        if uur < start_uur:
            continue

        present_students = lm5_present_students_on_hour(ctx["base_maps"], uur, ctx["abs_set"])
        pv_assignment_now = lm5_active_pv_assignment_for_hour(ctx, uur)
        pv_reserved_now = set(pv_assignment_now.values())

        for naam in present_students:
            student = ctx["student_states"][naam]

            if uur in student["assigned_hours"]:
                continue

            if naam in pv_reserved_now:
                student["assigned_hours"].append(uur)
                student["assigned_attracties"].add("Pauzevlinder-vervanging")
                ctx["prev_attr"][naam] = "Pauzevlinder-vervanging"
                continue

            ctx["extra_assignments"][uur].append(naam)
            student["assigned_hours"].append(uur)
            student["assigned_attracties"].add("Extra")
            ctx["prev_attr"][naam] = "Extra"

# ------------------------------------------------------------
# Extra -> attractie verplaatsing
# ------------------------------------------------------------
def lm5_try_fill_empty_slots_from_extras(ctx, start_uur):
    for uur in sorted(open_uren):
        if uur < start_uur:
            continue

        hstate = ctx["hour_states"][uur]
        extras_now = list(ctx["extra_assignments"][uur])

        for attr in hstate["active"]:
            max_spots = hstate["counts"].get(attr, 0)
            if attr in hstate["second_spot_blocked"]:
                max_spots = min(max_spots, 1)

            current = list(ctx["assigned_map"].get((uur, attr), []))

            while len(current) < max_spots:
                kandidaten = []

                for naam in extras_now:
                    student = ctx["student_states"][naam]

                    if not lm5_student_can_attr(student, attr):
                        continue

                    bestaande = sorted([
                        h for h in set(student["assigned_hours"])
                        if student["naam"] in ctx["assigned_map"].get((h, attr), [])
                    ])
                    totaal = sorted(set(bestaande) | {uur})

                    if len(totaal) > 6:
                        continue

                    orig_attr = ctx["base_maps"]["student_hour_attr"].get((naam, uur), "")
                    kandidaten.append((
                        0 if normalize_attr(orig_attr) == normalize_attr(attr) else 1,
                        0 if normalize_attr(ctx["prev_attr"].get(naam, "")) == normalize_attr(attr) else 1,
                        ctx["changes_count"][naam],
                        naam
                    ))

                if not kandidaten:
                    break

                kandidaten.sort()
                gekozen = kandidaten[0][3]

                ctx["extra_assignments"][uur].remove(gekozen)
                extras_now.remove(gekozen)

                ctx["assigned_map"][(uur, attr)].append(gekozen)
                ctx["per_hour_assigned_counts"][uur][attr] += 1
                ctx["student_states"][gekozen]["assigned_attracties"].add(attr)
                ctx["prev_attr"][gekozen] = attr

                current = list(ctx["assigned_map"].get((uur, attr), []))

# ------------------------------------------------------------
# Complete build
# ------------------------------------------------------------
def lm5_build_target_slots_for_hour(attr_rows, hour_state):
    slots = []
    inactive_rows = set()

    for _row, rijlabel in attr_rows:
        attr, pos = lm5_split_display_label(rijlabel)
        allowed = hour_state["counts"].get(attr, 0)

        if attr in hour_state["red_spots"]:
            allowed = 0

        if attr in hour_state["second_spot_blocked"] and pos == 2:
            allowed = 0

        if allowed >= pos:
            slots.append((attr, pos, rijlabel))
        else:
            inactive_rows.add(rijlabel)

    return slots, inactive_rows


# ------------------------------------------------------------
# Context
# ------------------------------------------------------------
def lm5_init_context(base_maps, absentees, start_uur):
    abs_set = {str(x).strip() for x in absentees}
    absent_pv = [n for n in abs_set if n in set(lm5_pv_names())]

    pv_replacements = lm5_pick_pv_replacements(
        absent_pv_names=absent_pv,
        start_uur=start_uur,
        base_maps=base_maps,
        absentees_set=abs_set
    )

    student_states = {}
    for s in studenten:
        student_states[str(s["naam"]).strip()] = lm5_copy_student_state(s)

    return {
        "base_maps": base_maps,
        "abs_set": abs_set,
        "absent_pv": absent_pv,
        "pv_replacements": pv_replacements,
        "student_states": student_states,
        "assigned_map": defaultdict(list),
        "extra_assignments": defaultdict(list),
        "per_hour_assigned_counts": {uur: defaultdict(int) for uur in open_uren},
        "hour_states": {},
        "changes_count": defaultdict(int),
        "prev_attr": {},
    }


def lm5_vrijgeven_afgekapte_pv_uren(ctx, start_uur):
    """
    Last-minute only: als de afgekapte PV-uren minstens 2 aaneensluitende
    uren bevatten, worden die uren vrijgegeven zodat de laatste PV (of diens
    vervanger) in die uren gewoon op de planning kan worden ingezet.
    """
    if not afgekapte_pv_uren or not selected:
        return

    # Alleen runs van >= 2 aaneensluitende uren vrijgeven
    runs = contiguous_runs(sorted(afgekapte_pv_uren))
    geschikte_uren = set()
    for run in runs:
        if len(run) >= 2:
            geschikte_uren.update(run)

    if not geschikte_uren:
        return

    laatste_pv_naam = str(selected[-1]["naam"]).strip()
    # Bepaal de effectieve persoon (zelf of vervanger)
    effectief_naam = ctx["pv_replacements"].get(laatste_pv_naam, laatste_pv_naam)

    # Sla op in ctx zodat lm5_active_pv_assignment_for_hour het kan gebruiken
    ctx["vrijgegeven_pv_uren"] = geschikte_uren
    ctx["vrijgegeven_pv_naam"] = laatste_pv_naam   # originele PV-naam, niet vervanger

    # Voeg de uren terug toe aan uren_beschikbaar van de effectieve persoon
    if effectief_naam in ctx["student_states"]:
        s = ctx["student_states"][effectief_naam]
        for uur in sorted(geschikte_uren):
            if uur >= start_uur and uur not in s["uren_beschikbaar"]:
                s["uren_beschikbaar"].append(uur)


def lm5_seed_hours_before_start(ctx, start_uur):
    for (naam, uur), attr in ctx["base_maps"]["student_hour_attr"].items():
        if uur >= start_uur:
            continue
        if naam in ctx["abs_set"]:
            continue

        ctx["assigned_map"][(uur, attr)].append(naam)
        ctx["per_hour_assigned_counts"][uur][attr] += 1
        ctx["student_states"][naam]["assigned_hours"].append(uur)
        ctx["student_states"][naam]["assigned_attracties"].add(attr)
        ctx["prev_attr"][naam] = attr


# ------------------------------------------------------------
# Complete build
# ------------------------------------------------------------

def lm5_extend_attr_rows_with_dynamic_merges(base_maps, ctx, start_uur):
    attr_rows = list(base_maps["attr_rows"])
    pv_rows = list(base_maps.get("pv_rows", []))
    extra_rows = list(base_maps.get("extra_rows", []))

    bestaande_attr_pos = set()
    for row, rijlabel in attr_rows:
        attr, pos = lm5_split_display_label(rijlabel)
        bestaande_attr_pos.add((attr, pos))

    # Verzamel alle dynamische samengestelde attracties die effectief voorkomen
    dynamische_attrs = {}

    for uur in sorted(open_uren):
        if uur < start_uur:
            continue
        if uur not in ctx["hour_states"]:
            continue

        hstate = ctx["hour_states"][uur]

        for attr in hstate["active"]:
            if " + " not in str(attr):
                continue

            max_pos = hstate["counts"].get(attr, 0)
            if attr in hstate["second_spot_blocked"]:
                max_pos = min(max_pos, 1)
            max_pos = max(1, max_pos)

            dynamische_attrs[attr] = max(dynamische_attrs.get(attr, 0), max_pos)

        for attr, count in hstate["counts"].items():
            if " + " not in str(attr):
                continue
            if count < 1:
                continue

            max_pos = count
            if attr in hstate["second_spot_blocked"]:
                max_pos = min(max_pos, 1)
            max_pos = max(1, max_pos)

            dynamische_attrs[attr] = max(dynamische_attrs.get(attr, 0), max_pos)

    if not dynamische_attrs:
        return

    # Bepaal waar de nieuwe rijen moeten komen:
    # liefst ONDER de extra-rijen, anders onder de pauzevlinderrijen,
    # anders onder de bestaande attractierijen.
    alle_bestaande_rijen = [row for row, _ in attr_rows] + [row for row, _ in pv_rows] + [row for row, _ in extra_rows]
    if alle_bestaande_rijen:
        insert_after = max(alle_bestaande_rijen)
    else:
        insert_after = 1

    nieuwe_attr_rows = []

    current_row = insert_after + 1

    for attr in sorted(dynamische_attrs.keys(), key=lambda x: str(x).lower()):
        nodig = dynamische_attrs[attr]

        for pos in range(1, nodig + 1):
            if (attr, pos) in bestaande_attr_pos:
                continue

            current_row += 1

            if nodig > 1:
                rijlabel = f"{attr} {pos}"
            else:
                rijlabel = attr

            nieuwe_attr_rows.append((current_row, rijlabel))
            bestaande_attr_pos.add((attr, pos))

    # Voeg de nieuwe rijen toe aan attr_rows
    attr_rows.extend(nieuwe_attr_rows)

    # Sorteer op fysieke rij
    attr_rows.sort(key=lambda x: x[0])

    # Schrijf terug naar base_maps
    base_maps["attr_rows"] = attr_rows


def lm5_extend_extra_rows_if_needed(base_maps, ctx):
    """
    Als ctx meer extra-studenten heeft dan er extra-rijen bestaan,
    voeg dan de ontbrekende 'Extra N'-rijen toe aan base_maps.
    """
    extra_rows = list(base_maps.get("extra_rows", []))

    # Hoeveel extra-plekken zijn er maximaal nodig op één uur?
    max_nodig = 0
    for uur, namen in ctx["extra_assignments"].items():
        max_nodig = max(max_nodig, len(namen))

    extra_tekort = max_nodig - len(extra_rows)
    if extra_tekort <= 0:
        return

    # Bepaal na welke rij de nieuwe extra-rijen komen:
    # 1 lege rij na de laagste bestaande rij in het sheet
    attr_rows  = list(base_maps.get("attr_rows", []))
    pv_rows    = list(base_maps.get("pv_rows", []))
    alle_rijen = [row for row, _ in attr_rows + pv_rows + extra_rows]
    insert_after = max(alle_rijen) if alle_rijen else 1

    # +2: 1 lege rij spatie, dan de eerste nieuwe extra-rij
    current_row = insert_after + 2

    for i in range(extra_tekort):
        nieuw_idx = len(extra_rows) + i + 1
        label = f"Extra {nieuw_idx}"
        extra_rows.append((current_row, label))
        current_row += 1

    base_maps["extra_rows"] = extra_rows


# ------------------------------------------------------------
# Post-processing: 5/6-uursblokken wisselen (last-minute versie)
# ------------------------------------------------------------
def lm5_postprocess_long_blocks(ctx, start_uur):
    """
    Equivalent van de hoofdplanning-postprocessing, maar voor de ctx-structuur.
    Werkt alleen op uren >= start_uur (uren daarvoor zijn ingezaaid en vast).
    """

    am = ctx["assigned_map"]

    # --- Hulpfuncties die werken op ctx ---

    def lm5_pp_get_attr_on_hour(naam, uur):
        for (u, a), namen in am.items():
            if u == uur and naam in namen:
                return a
        return None

    def lm5_pp_get_hours_on_attr(naam, attr):
        return sorted(u for (u, a), namen in am.items() if a == attr and naam in namen)

    def lm5_pp_get_runs_on_attr(naam, attr):
        return contiguous_runs(lm5_pp_get_hours_on_attr(naam, attr))

    def lm5_pp_count_attr_switches(naam):
        uur_attr = sorted(
            (u, lm5_pp_get_attr_on_hour(naam, u))
            for u in sorted(set(ctx["student_states"][naam]["assigned_hours"]))
            if lm5_pp_get_attr_on_hour(naam, u)
        )
        if not uur_attr:
            return 0
        switches = 0
        prev = uur_attr[0][1]
        for _, a in uur_attr[1:]:
            if a != prev:
                switches += 1
            prev = a
        return switches

    def lm5_pp_count_problem_attrs(naam):
        return sum(
            1 for (_, a), namen in am.items()
            if naam in namen and len(lm5_pp_get_hours_on_attr(naam, a)) > 4
        )

    def lm5_pp_total_overflow(naam):
        seen = set()
        overflow = 0
        for (_, a), namen in am.items():
            if naam in namen and a not in seen:
                seen.add(a)
                uren = len(lm5_pp_get_hours_on_attr(naam, a))
                if uren > 4:
                    overflow += uren - 4
        return overflow

    def lm5_pp_remove(naam, uur, attr):
        key = (uur, attr)
        if naam in am.get(key, []):
            am[key].remove(naam)
        s = ctx["student_states"][naam]
        if uur in s["assigned_hours"]:
            s["assigned_hours"].remove(uur)

    def lm5_pp_add(naam, uur, attr):
        am[(uur, attr)].append(naam)
        ctx["student_states"][naam]["assigned_hours"].append(uur)
        ctx["student_states"][naam]["assigned_attracties"].add(attr)

    def lm5_pp_rebuild_attrs(naam):
        s = ctx["student_states"][naam]
        s["assigned_attracties"] = {
            lm5_pp_get_attr_on_hour(naam, u)
            for u in set(s["assigned_hours"])
            if lm5_pp_get_attr_on_hour(naam, u)
        }

    def lm5_pp_is_valid(naam, attr, uren):
        """Student mag attr op al die uren doen, en uren >= start_uur."""
        if not all(u >= start_uur for u in uren):
            return False
        s = ctx["student_states"].get(naam)
        if not s or not lm5_student_can_attr(s, attr):
            return False
        for u in uren:
            hstate = ctx["hour_states"].get(u, {})
            if attr not in hstate.get("active", set()):
                return False
        return True

    def lm5_pp_respects_rules(naam, attr):
        uren = lm5_pp_get_hours_on_attr(naam, attr)
        if len(uren) > 6:
            return False
        if max_consecutive_hours(uren) > 4:
            return False
        return True

    def lm5_pp_is_swap_target(naam, attr, block_hours):
        """Staat naam op exact al die uren op attr?"""
        return all(naam in am.get((u, attr), []) for u in block_hours)

    def lm5_pp_try_swap_block(naam_a, attr_a, block_hours, all_names):
        if len(block_hours) not in [2, 3]:
            return False

        orig_sw_a  = lm5_pp_count_attr_switches(naam_a)
        orig_pr_a  = lm5_pp_count_problem_attrs(naam_a)
        orig_ov_a  = lm5_pp_total_overflow(naam_a)
        eerste_uur = block_hours[0]

        for naam_b in all_names:
            if naam_b == naam_a:
                continue

            attr_b = lm5_pp_get_attr_on_hour(naam_b, eerste_uur)
            if not attr_b or attr_b == attr_a:
                continue
            if not lm5_pp_is_swap_target(naam_b, attr_b, block_hours):
                continue
            if not lm5_pp_is_valid(naam_a, attr_b, block_hours):
                continue
            if not lm5_pp_is_valid(naam_b, attr_a, block_hours):
                continue

            orig_sw_b = lm5_pp_count_attr_switches(naam_b)
            orig_pr_b = lm5_pp_count_problem_attrs(naam_b)
            orig_ov_b = lm5_pp_total_overflow(naam_b)

            # Wissel uitvoeren
            for u in block_hours:
                lm5_pp_remove(naam_a, u, attr_a)
                lm5_pp_remove(naam_b, u, attr_b)
            for u in block_hours:
                lm5_pp_add(naam_a, u, attr_b)
                lm5_pp_add(naam_b, u, attr_a)
            lm5_pp_rebuild_attrs(naam_a)
            lm5_pp_rebuild_attrs(naam_b)

            # Geen geïsoleerde 1-uursblokken laten ontstaan
            def heeft_geisoleerd_uur(naam, attr):
                runs = lm5_pp_get_runs_on_attr(naam, attr)
                return any(len(r) == 1 for r in runs)

            valid = all(
                lm5_pp_respects_rules(n, a)
                for n, a in [(naam_a, attr_a), (naam_a, attr_b),
                             (naam_b, attr_a), (naam_b, attr_b)]
            )

            if heeft_geisoleerd_uur(naam_a, attr_a): valid = False
            if heeft_geisoleerd_uur(naam_a, attr_b): valid = False
            if heeft_geisoleerd_uur(naam_b, attr_a): valid = False
            if heeft_geisoleerd_uur(naam_b, attr_b): valid = False

            extra_sw = (
                (lm5_pp_count_attr_switches(naam_a) - orig_sw_a) +
                (lm5_pp_count_attr_switches(naam_b) - orig_sw_b)
            )
            if extra_sw > 2:
                valid = False

            new_pr = lm5_pp_count_problem_attrs(naam_a) + lm5_pp_count_problem_attrs(naam_b)
            new_ov = lm5_pp_total_overflow(naam_a)      + lm5_pp_total_overflow(naam_b)
            orig_pr = orig_pr_a + orig_pr_b
            orig_ov = orig_ov_a + orig_ov_b

            if new_pr > orig_pr:
                valid = False
            if new_pr == orig_pr and new_ov > orig_ov:
                valid = False

            verbetering = (
                new_pr < orig_pr
                or (new_pr == orig_pr and new_ov < orig_ov)
            )
            if not verbetering:
                valid = False

            if valid:
                return True

            # Rollback
            for u in block_hours:
                lm5_pp_remove(naam_a, u, attr_b)
                lm5_pp_remove(naam_b, u, attr_a)
            for u in block_hours:
                lm5_pp_add(naam_a, u, attr_a)
                lm5_pp_add(naam_b, u, attr_b)
            lm5_pp_rebuild_attrs(naam_a)
            lm5_pp_rebuild_attrs(naam_b)

        return False

    def lm5_pp_try_swap_long_attr(naam, attr, all_names):
        if len(lm5_pp_get_hours_on_attr(naam, attr)) <= 4:
            return False

        runs = lm5_pp_get_runs_on_attr(naam, attr)
        if not runs:
            return False

        def kandidaat_blokken(run):
            blokken = []
            if len(run) >= 3:
                blokken.append(run[-3:])
            if len(run) >= 2:
                blokken.append(run[-2:])
            if len(run) >= 3 and run[:3] != run[-3:]:
                blokken.append(run[:3])
            if len(run) >= 2 and run[:2] != run[-2:]:
                blokken.append(run[:2])
            return blokken

        laatste_run = runs[-1]
        eerste_run  = runs[0]

        for blok in kandidaat_blokken(laatste_run):
            if lm5_pp_try_swap_block(naam, attr, blok, all_names):
                return True
        if eerste_run != laatste_run:
            for blok in kandidaat_blokken(eerste_run):
                if lm5_pp_try_swap_block(naam, attr, blok, all_names):
                    return True
        return False

    # --- Hoofdlus ---
    all_names = [
        naam for naam, s in ctx["student_states"].items()
        if any(u >= start_uur for u in s["assigned_hours"])
        and not s.get("is_pauzevlinder")
    ]

    for _ in range(7):
        wijziging = False
        for naam in all_names:
            s = ctx["student_states"][naam]
            probleem_attrs = sorted(
                {a for a in s["assigned_attracties"]
                 if len(lm5_pp_get_hours_on_attr(naam, a)) > 4},
                key=lambda a: -len(lm5_pp_get_hours_on_attr(naam, a))
            )
            for attr in probleem_attrs:
                if lm5_pp_try_swap_long_attr(naam, attr, all_names):
                    wijziging = True
                    break
        if not wijziging:
            break



def lm5_build_lastminute_context(base_bytes, absentees, start_uur):
    base_maps = lm5_extract_base_maps(base_bytes)
    ctx = lm5_init_context(base_maps, absentees, start_uur)
    herbereken_afgekapte_pv_uren(absentees_set=ctx["abs_set"], base_maps=base_maps)
    lm5_vrijgeven_afgekapte_pv_uren(ctx, start_uur)
    lm5_seed_hours_before_start(ctx, start_uur)
    capacity_actions = lm5_extract_capacity_actions()


    # STAP 1: per uur echte capaciteit herberekenen
    for uur in sorted(open_uren):
        if uur < start_uur:
            continue

        present_attraction_students = lm5_present_attraction_students_on_hour(ctx, uur)

        hour_state = lm5_rebuild_hour_state(
            uur=uur,
            available_attraction_students=len(present_attraction_students),
            capacity_actions=capacity_actions
        )

        ctx["hour_states"][uur] = hour_state

    # NIEUW: zorg dat nieuw samengestelde attracties ook echte rijen krijgen
    lm5_extend_attr_rows_with_dynamic_merges(base_maps, ctx, start_uur)

  

    # STAP 2: eerst zoveel mogelijk exact dezelfde plek houden
    for uur in sorted(open_uren):
        if uur < start_uur:
            continue

        hstate = ctx["hour_states"][uur]
        target_slots, _inactive_rows = lm5_build_target_slots_for_hour(base_maps["attr_rows"], hstate)
        present_attraction_students = lm5_present_attraction_students_on_hour(ctx, uur)

        used_students, assigned_rows = lm5_seed_same_place_first(
            ctx=ctx,
            uur=uur,
            target_slots=target_slots,
            present_attraction_students=present_attraction_students
        )

        lm5_fill_remaining_hour(
            ctx=ctx,
            uur=uur,
            target_slots=target_slots,
            present_attraction_students=present_attraction_students,
            used_students=used_students,
            assigned_rows=assigned_rows
        )

    # STAP 3: vrijgekomen studenten en lege plekken verzamelen
    released_students, missing_slots_by_hour = lm5_collect_released_students_and_missing_slots(
        ctx=ctx,
        base_maps=base_maps,
        start_uur=start_uur
    )

    # STAP 4: directe invulling
    lm5_try_direct_fill_from_released_students(
        ctx=ctx,
        released_students=released_students,
        missing_slots_by_hour=missing_slots_by_hour
    )

    # STAP 5: kettingwissels op blokken van 3/2/1
    for _ in range(5):
        changed = lm5_try_fill_missing_with_chain_swaps(
            ctx=ctx,
            released_students=released_students,
            missing_slots_by_hour=missing_slots_by_hour,
            start_uur=start_uur
        )
        if not changed:
            break

    # STAP 6: resterende gaten met blokvoorkeur 3/2/1
    lm5_assign_future_blocks(ctx, start_uur)

    # STAP 7: exact 1 keer per gewerkt uur
    lm5_force_exactly_one_assignment_per_hour(ctx, start_uur)

    # STAP 8: lege attractieplekken opvullen door Extra -> attractie
    lm5_try_fill_empty_slots_from_extras(ctx, start_uur)

    # STAP 9: post-processing 5/6-uursblokken wegwisselen
    lm5_postprocess_long_blocks(ctx, start_uur)

    lm5_extend_extra_rows_if_needed(base_maps, ctx)

    return ctx, base_maps


# ------------------------------------------------------------
# Output schrijven
# ------------------------------------------------------------

def lm5_reconstruct_studenten(ctx_assigned_map):
    """Maak een kopie van studenten met assigned_hours vanuit de nieuwe ctx."""
    import copy as _copy
    hours_per_student = defaultdict(list)
    for (uur, attr), namen in ctx_assigned_map.items():
        for naam in namen:
            hours_per_student[str(naam).strip()].append(uur)

    lm_studenten = _copy.deepcopy(studenten)
    for s in lm_studenten:
        s["assigned_hours"] = hours_per_student.get(str(s["naam"]).strip(), [])
    return lm_studenten

def lm5_write_lastminute_workbook(base_bytes, ctx, base_maps, start_uur, absentees):
    wb_lm = load_workbook(BytesIO(base_bytes))
    ws_plan = wb_lm["Planning"]
    # Last-minute planning toont altijd datum van vandaag, ongeacht W4
    ws_plan.cell(1, 1).value = vandaag_altijd_vandaag
    ws_pauze_lm = wb_lm["Pauzevlinders"]
    ws_pauze_lm.cell(1, 1).value = vandaag_altijd_vandaag

    gray_fill  = PatternFill(start_color="808080", end_color="808080", fill_type="solid")
    white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

    uur_to_col = base_maps["uur_to_col"]
    attr_rows  = list(base_maps["attr_rows"])
    pv_rows    = base_maps["pv_rows"]
    extra_rows = base_maps["extra_rows"]
    # Samengevoegde attractierijen fysiek boven pauzevlinders plaatsen
    if pv_rows and attr_rows:
        eerste_pv_rij = min(row for row, _ in pv_rows)
        nieuwe_attr = [(row, label) for row, label in attr_rows if row >= eerste_pv_rij]
        if nieuwe_attr:
            aantal = len(nieuwe_attr)
            ws_plan.insert_rows(eerste_pv_rij, amount=aantal + 1)
            verschuiving = aantal + 1
            pv_rows    = [(row + verschuiving, label) for row, label in pv_rows]
            extra_rows = [(row + verschuiving, label) for row, label in extra_rows]
            hernummer  = {old: eerste_pv_rij + i for i, (old, _) in enumerate(nieuwe_attr)}
            attr_rows  = [(hernummer[row], label) if row in hernummer else (row, label) for row, label in attr_rows]
            base_maps["pv_rows"]    = pv_rows
            base_maps["extra_rows"] = extra_rows
            base_maps["attr_rows"]  = attr_rows

    center_align = Alignment(horizontal="center", vertical="center")
    thin_border  = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    # Zorg dat alle dynamische attractierijen echt bestaan in kolom A
    for row, rijlabel in attr_rows:
        huidige_waarde = ws_plan.cell(row, 1).value
        if huidige_waarde is None or str(huidige_waarde).strip() == "":
            ws_plan.cell(row, 1).value     = rijlabel
            ws_plan.cell(row, 1).font      = Font(bold=True)
            ws_plan.cell(row, 1).fill      = white_fill
            ws_plan.cell(row, 1).alignment = center_align
            ws_plan.cell(row, 1).border    = thin_border

        for uur in sorted(open_uren):
            if uur not in uur_to_col:
                continue
            col  = uur_to_col[uur]
            cell = ws_plan.cell(row, col)
            cell.alignment = center_align
            cell.border    = thin_border
            if cell.fill.fill_type is None:
                cell.fill = white_fill

    # Zorg dat ook nieuwe extra-rijen in kolom A staan
    for row, rijlabel in extra_rows:
        huidige_waarde = ws_plan.cell(row, 1).value
        if huidige_waarde is None or str(huidige_waarde).strip() == "":
            ws_plan.cell(row, 1).value     = rijlabel
            ws_plan.cell(row, 1).font      = Font(bold=True)
            ws_plan.cell(row, 1).fill      = white_fill
            ws_plan.cell(row, 1).alignment = center_align
            ws_plan.cell(row, 1).border    = thin_border

            for uur in sorted(open_uren):
                if uur not in uur_to_col:
                    continue
                col  = uur_to_col[uur]
                cell = ws_plan.cell(row, col)
                cell.alignment = center_align
                cell.border    = thin_border
                cell.fill      = white_fill

    for uur in sorted(open_uren):
        if uur < start_uur:
            continue
        if uur not in uur_to_col:
            continue

        col    = uur_to_col[uur]
        hstate = ctx["hour_states"][uur]

        # Attractierijen
        for row, rijlabel in attr_rows:
            cell       = ws_plan.cell(row, col)
            attr, pos  = lm5_split_display_label(rijlabel)
            allowed    = hstate["counts"].get(attr, 0)
            inactive   = False

            if attr in hstate["red_spots"]:
                inactive = True
            if attr in hstate["second_spot_blocked"] and pos == 2:
                inactive = True
            if allowed < pos:
                inactive = True

            namen = list(ctx["assigned_map"].get((uur, attr), []))
            naam  = namen[pos - 1] if pos - 1 < len(namen) else ""

            cell.value     = naam
            cell.alignment = center_align
            cell.border    = thin_border

            if inactive:
                cell.fill = gray_fill
            elif naam and naam in student_kleuren:
                cell.fill = PatternFill(
                    start_color=student_kleuren[naam],
                    end_color=student_kleuren[naam],
                    fill_type="solid"
                )
            else:
                cell.fill = white_fill

        # PV-rijen
        pv_assignment_now = lm5_active_pv_assignment_for_hour(ctx, uur)
        for idx, (row, _label) in enumerate(pv_rows, start=1):
            cell   = ws_plan.cell(row, col)
            pvnaam = str(selected[idx - 1]["naam"]).strip() if idx <= len(selected) else ""
            naam   = pv_assignment_now.get(pvnaam, "")

            cell.value     = naam
            cell.alignment = center_align
            cell.border    = thin_border

            if naam and naam in student_kleuren:
                cell.fill = PatternFill(
                    start_color=student_kleuren[naam],
                    end_color=student_kleuren[naam],
                    fill_type="solid"
                )
            else:
                cell.fill = white_fill

        # Extra-rijen
        extras_now = list(ctx["extra_assignments"][uur])
        for idx, (row, _label) in enumerate(extra_rows):
            cell = ws_plan.cell(row, col)
            naam = extras_now[idx] if idx < len(extras_now) else ""

            cell.value     = naam
            cell.alignment = center_align
            cell.border    = thin_border

            if naam and naam in student_kleuren:
                cell.fill = PatternFill(
                    start_color=student_kleuren[naam],
                    end_color=student_kleuren[naam],
                    fill_type="solid"
                )
            else:
                cell.fill = white_fill

    # Reconstruct assigned_hours per student vanuit ctx
    import copy as _copy
    _hours_per_student = defaultdict(list)
    for (uur, attr), namen in ctx["assigned_map"].items():
        for naam in namen:
            _hours_per_student[str(naam).strip()].append(uur)

    studenten_lm = _copy.deepcopy(studenten)
    for s in studenten_lm:
        s["assigned_hours"] = _hours_per_student.get(str(s["naam"]).strip(), [])

    # Herwerk Analyse
    # Bouw actieve attracties per uur vanuit last-minute hour_states
    lm_actieve_attracties = {}
    for uur in open_uren:
        if uur in ctx["hour_states"]:
            lm_actieve_attracties[uur] = ctx["hour_states"][uur]["active"]
        else:
            lm_actieve_attracties[uur] = actieve_attracties_per_uur.get(uur, set())
    maak_analyse_sheet(
        wb_lm,
        ctx["assigned_map"],
        ctx["extra_assignments"],
        studenten_lm,
        actieve_attracties_override=lm_actieve_attracties
    )
    # Herwerk Wisselplanning
    maak_wisselplanning_sheet(wb_lm, ctx["assigned_map"])

    # Herwerk PP optie 2:
    # Vervang eerst kolom A in Pauzevlinders én tijdelijk selected,
    # zodat maak_pp2_sheets de vervanger als echte PV behandelt
    # (juiste werkuren, eigen pauze op eigen rij).
    selected_bak = [dict(pv) for pv in selected]
    if ctx.get("pv_replacements"):
        ws_pauze_lm = wb_lm["Pauzevlinders"]
        for i, pv in enumerate(selected):
            pvnaam    = str(pv["naam"]).strip()
            vervanger = ctx["pv_replacements"].get(pvnaam)
            if not vervanger:
                continue
            vervanger_student = next(
                (s for s in studenten if str(s["naam"]).strip() == vervanger), None
            )
            if not vervanger_student:
                continue
            # Kolom A in Pauzevlinders bijwerken
            for r in range(2, ws_pauze_lm.max_row + 1):
                if str(ws_pauze_lm.cell(r, 1).value or "").strip() == pvnaam:
                    ws_pauze_lm.cell(r, 1).value = vervanger
                    break
            # selected tijdelijk aanpassen zodat maak_pp2_sheets
            # de vervanger als PV herkent en diens werkuren gebruikt
            selected[i] = dict(vervanger_student)
            selected[i]["is_pauzevlinder"] = True
            selected[i]["pv_number"]       = pv.get("pv_number", i + 1)

    herbereken_afgekapte_pv_uren(absentees_set=ctx["abs_set"], base_maps=base_maps)
    maak_pp2_sheets(wb_lm, ctx["assigned_map"])

    for i in range(len(selected)):
        if i < len(selected_bak):
            selected[i] = selected_bak[i]
    herbereken_afgekapte_pv_uren()  # ← zonder args: terug naar originele staat

    return wb_lm
    
    
# ------------------------------------------------------------
# UI
# ------------------------------------------------------------
st.markdown("### Last-minute afwezigen")


@st.cache_data
def _cached_base_maps(base_bytes):
    return lm5_extract_base_maps(base_bytes)
    
base_bytes_lm5 = st.session_state["lm_base_bytes"]
base_maps_lm5 = _cached_base_maps(base_bytes_lm5)   # ← slaat nu WEL aan
werkende_studenten_vandaag_lm5 = lm5_working_students_today(base_maps_lm5)

with st.expander("Last-minute afwezigen", expanded=False):
    with st.form("lm5_form"):
        gekozen_afwezigen_lm5 = st.multiselect(
            "Kies 1 tot 5 afwezige studenten",
            options=werkende_studenten_vandaag_lm5,
            default=[],
        )
        start_uur_lm5 = st.selectbox(
            "Vanaf welk uur moet de nieuwe planning starten?",
            options=sorted(open_uren),
            format_func=formatteer_uur,
        )
        submitted = st.form_submit_button("Maak last-minute planning")

    if submitted:
        if not gekozen_afwezigen_lm5:
            st.warning("Kies eerst minstens 1 afwezige student.")
        elif len(gekozen_afwezigen_lm5) > 5:
            st.warning("Je mag maximaal 5 studenten kiezen.")
        else:
            try:
                ctx_lm5, base_maps_lm5_build = lm5_build_lastminute_context(
                    base_bytes=base_bytes_lm5,
                    absentees=gekozen_afwezigen_lm5,
                    start_uur=start_uur_lm5
                )

                lm5_result = lm5_write_lastminute_workbook(
                    base_bytes=base_bytes_lm5,
                    ctx=ctx_lm5,
                    base_maps=base_maps_lm5_build,
                    start_uur=start_uur_lm5,
                    absentees=gekozen_afwezigen_lm5
                )

                if isinstance(lm5_result, (bytes, bytearray)):
                    lm5_file_bytes = bytes(lm5_result)
                else:
                    lm5_output = BytesIO()
                    lm5_result.save(lm5_output)
                    lm5_output.seek(0)
                    lm5_file_bytes = lm5_output.getvalue()

                st.session_state["lm5_result_bytes"] = lm5_file_bytes
                st.session_state["lm5_result_filename"] = f"Planning_last_minute_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

            except Exception as e:
                st.error(f"Fout in last-minute planner: {e}")

    if "lm5_result_bytes" in st.session_state:
        st.success("Last-minute planning gemaakt.")
        st.download_button(
            "Download last-minute planning",
            data=st.session_state["lm5_result_bytes"],
            file_name=st.session_state.get("lm5_result_filename", "Planning_last_minute.xlsx"),
            key="lm5_download_button"
        )
